#!/usr/bin/env python3
"""Inspect all sheets in inventario-limpo.xlsx — headers, sample data, merged cells."""

import openpyxl
from openpyxl.utils import get_column_letter
import sys

XLSX = "/Users/marko/Desktop/CLAUDE/PROJETOS/mmd/data/inventario-limpo.xlsx"

CATEGORY_SHEETS = [
    "ILUMINACAO", "AUDIO", "CABO", "ENERGIA",
    "ESTRUTURA", "EFEITO", "VIDEO", "ACESSORIO",
]

wb = openpyxl.load_workbook(XLSX, data_only=True)

print(f"Total sheets: {len(wb.sheetnames)}")
print(f"Sheet names: {wb.sheetnames}")
print("=" * 100)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'=' * 100}")
    print(f"SHEET: {sheet_name}")
    print(f"  Dimensions: {ws.dimensions}")
    print(f"  Max row: {ws.max_row}, Max col: {ws.max_column}")

    # Report merged cells
    if ws.merged_cells.ranges:
        print(f"  Merged cells: {[str(m) for m in ws.merged_cells.ranges]}")
    else:
        print(f"  Merged cells: None")

    # For category sheets, headers are at row 4
    if sheet_name in CATEGORY_SHEETS:
        header_row = 4
        print(f"\n  --- Headers (row {header_row}) ---")
        headers = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=header_row, column=col).value
            letter = get_column_letter(col)
            headers.append((letter, val))
        for letter, val in headers:
            print(f"    Col {letter}: {val!r}")

        # Show first data row (row 5)
        print(f"\n  --- Sample data row (row 5) ---")
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=5, column=col).value
            letter = get_column_letter(col)
            h = headers[col-1][1] if col-1 < len(headers) else "?"
            print(f"    Col {letter} [{h}]: {val!r}")

        # Check rows 1-3 for any content
        print(f"\n  --- Rows 1-3 (pre-header) ---")
        for r in range(1, 4):
            row_vals = []
            for col in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=col).value
                if v is not None:
                    row_vals.append(f"{get_column_letter(col)}={v!r}")
            if row_vals:
                print(f"    Row {r}: {', '.join(row_vals)}")
            else:
                print(f"    Row {r}: (empty)")

        # Count data rows
        data_rows = 0
        for r in range(5, ws.max_row + 1):
            if ws.cell(row=r, column=1).value is not None:
                data_rows += 1
        print(f"\n  Data rows (with col A populated): {data_rows}")

    elif sheet_name == "LOTES":
        print(f"\n  --- Scanning for header row (first 10 rows) ---")
        header_row = None
        for r in range(1, min(11, ws.max_row + 1)):
            row_vals = []
            for col in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=col).value
                if v is not None:
                    row_vals.append(f"{get_column_letter(col)}={v!r}")
            if row_vals:
                print(f"    Row {r}: {', '.join(row_vals)}")
                if header_row is None and len(row_vals) >= 3:
                    header_row = r
            else:
                print(f"    Row {r}: (empty)")

        if header_row:
            print(f"\n  --- Likely header row: {header_row} ---")
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=header_row, column=col).value
                letter = get_column_letter(col)
                print(f"    Col {letter}: {val!r}")

            # Sample data
            data_row = header_row + 1
            print(f"\n  --- Sample data row (row {data_row}) ---")
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=data_row, column=col).value
                letter = get_column_letter(col)
                print(f"    Col {letter}: {val!r}")

    elif sheet_name == "FORA DE OPERACAO":
        print(f"\n  --- Scanning for header row (first 10 rows) ---")
        header_row = None
        for r in range(1, min(11, ws.max_row + 1)):
            row_vals = []
            for col in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=col).value
                if v is not None:
                    row_vals.append(f"{get_column_letter(col)}={v!r}")
            if row_vals:
                print(f"    Row {r}: {', '.join(row_vals)}")
                if header_row is None and len(row_vals) >= 3:
                    header_row = r
            else:
                print(f"    Row {r}: (empty)")

        if header_row:
            print(f"\n  --- Likely header row: {header_row} ---")
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=header_row, column=col).value
                letter = get_column_letter(col)
                print(f"    Col {letter}: {val!r}")

            data_row = header_row + 1
            print(f"\n  --- Sample data row (row {data_row}) ---")
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=data_row, column=col).value
                letter = get_column_letter(col)
                print(f"    Col {letter}: {val!r}")

    elif sheet_name == "DASHBOARD":
        print(f"\n  --- Full content scan (all non-empty cells) ---")
        for r in range(1, ws.max_row + 1):
            row_vals = []
            for col in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=col).value
                if v is not None:
                    row_vals.append(f"{get_column_letter(col)}={v!r}")
            if row_vals:
                print(f"    Row {r}: {', '.join(row_vals)}")

    elif sheet_name == "MANUAL":
        print(f"\n  --- First 20 rows ---")
        for r in range(1, min(21, ws.max_row + 1)):
            row_vals = []
            for col in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=col).value
                if v is not None:
                    row_vals.append(f"{get_column_letter(col)}={v!r}")
            if row_vals:
                print(f"    Row {r}: {', '.join(row_vals)}")
            else:
                print(f"    Row {r}: (empty)")

    elif sheet_name == "REF CATEGORIAS":
        print(f"\n  --- First 20 rows ---")
        for r in range(1, min(21, ws.max_row + 1)):
            row_vals = []
            for col in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=col).value
                if v is not None:
                    row_vals.append(f"{get_column_letter(col)}={v!r}")
            if row_vals:
                print(f"    Row {r}: {', '.join(row_vals)}")
            else:
                print(f"    Row {r}: (empty)")

    else:
        # Unknown sheet — dump first 15 rows
        print(f"\n  --- First 15 rows ---")
        for r in range(1, min(16, ws.max_row + 1)):
            row_vals = []
            for col in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=col).value
                if v is not None:
                    row_vals.append(f"{get_column_letter(col)}={v!r}")
            if row_vals:
                print(f"    Row {r}: {', '.join(row_vals)}")
            else:
                print(f"    Row {r}: (empty)")

wb.close()
print("\n\nDone.")
