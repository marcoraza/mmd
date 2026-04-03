#!/usr/bin/env python3
from __future__ import annotations

import argparse
import math
import re
import shutil
import statistics
import time
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from difflib import SequenceMatcher
from html import unescape
from pathlib import Path
from typing import Iterable
from urllib.error import HTTPError, URLError
from urllib.parse import quote
from urllib.request import Request, urlopen

from openpyxl import load_workbook


THIS_DIR = Path(__file__).resolve().parent
PROJECT_DIR = THIS_DIR.parent
DEFAULT_CLEANED = PROJECT_DIR / "data" / "inventario-limpo.xlsx"
DEFAULT_ORIGINAL = PROJECT_DIR / "data" / "inventario-original.xlsx"
AUDIT_SHEET = "REF Valores"

STOPWORDS = {
    "A",
    "AO",
    "AS",
    "COM",
    "COR",
    "DA",
    "DE",
    "DO",
    "DOS",
    "E",
    "EM",
    "KIT",
    "MODEL",
    "MODELO",
    "NO",
    "NA",
    "O",
    "OS",
    "PARA",
    "POR",
    "SEM",
    "THE",
    "UM",
    "UMA",
}

NEGATIVE_TOKENS = {
    "ADAPTADOR",
    "BAG",
    "BATERIA",
    "BOLSA",
    "CABO",
    "CAPA",
    "CAPAS",
    "CASE",
    "COLMEIA",
    "FONTE",
    "LAMPADA",
    "LAMPADAS",
    "MICROFONE",
    "MOTOR",
    "PECA",
    "PEÇA",
    "REPARO",
    "REPAROS",
    "REFIL",
    "RECEPTOR",
    "RESERVATORIO",
    "SUPORTE",
    "TRANSCODIFICADOR",
}

GENERIC_QUERIES = {
    "ACESSORIO",
    "AUDIO",
    "CABO",
    "EFEITO",
    "ENERGIA",
    "ESTRUTURA",
    "GERAL",
    "ILUMINACAO",
    "ILUMINAÇÃO",
    "OUTRO",
    "OUTROS",
    "VIDEO",
}

SUBCATEGORY_ESTIMATE_ALIASES = {
    "BASE DE MICROFONE": "WIRELESS",
    "BASE WIRELESS": "WIRELESS",
    "BODYPACK": "WIRELESS",
    "CAIXA SOM": "CAIXA DE SOM",
    "CAIXA DE SOM": "CAIXA DE SOM",
    "BASE MIC": "WIRELESS",
    "IN EAR S / FIO": "IN EAR S/ FIO",
    "MESA LUZ": "MESA DE LUZ",
    "MESA SOM": "MESA DE SOM",
    "MICROFONE SEM FIO": "WIRELESS",
    "MICROFONE WIRELESS": "WIRELESS",
    "MIFROFONE SEM FIO": "WIRELESS",
    "MIROFONE SEM FIO": "WIRELESS",
    "RX MIC": "WIRELESS",
    "RECEPTOR DE ANTENA": "WIRELESS",
    "RECEPTOR MICROFONE": "WIRELESS",
    "RIBALTA": "OUTDOOR",
    "OUTRAS LUZES": "INDOOR",
    "PAR LEDS": "OUTDOOR",
    "REGUA": "REGUA",
    "REGUA 6 TOMADAS": "REGUA",
    "REGUA 12 TOMADAS": "REGUA",
    "RÉGUA": "REGUA",
    "TRIPE DE CAIXA": "TRIPE DE CAIXA",
    "TRIPE DE MICROFONE": "TRIPE DE MICROFONE",
    "TRIPE CAIXA": "TRIPE DE CAIXA",
    "TRIPE MIC": "TRIPE DE MICROFONE",
    "TRIPÉ DE CAIXA": "TRIPE DE CAIXA",
    "TRIPÉ DE MICROFONE": "TRIPE DE MICROFONE",
    "SUB": "SUBWOOFER",
}

MODEL_TOKEN_PATTERN = re.compile(r"[A-Z0-9]+(?:[-/][A-Z0-9]+)+|[A-Z]*\d+[A-Z0-9-]*")

MANUAL_QUERY_OVERRIDES = {
    "APPLE APPLE ROUTER BRANCO": ["apple airport express branco"],
    "SENDER BOX SB-8": ["senderbox sb 8", "sender box sb-8"],
    "MOTOROLA TALK ABOUT KIT COMPLETO COM FONTE E FONE": ["motorola talkabout radio kit"],
    "ROLLAND GR-33": ["roland gr-33 guitar synthesizer"],
    "SENNHEISER EWD SKM S FREQ Q1 6": ["sennheiser ew-d skm-s q1-6"],
    "SENNHEISER SKM100S G4 FREQ G": ["sennheiser skm 100 g4"],
    "MICWL": ["microfone sem fio"],
    "PARA MICROFONE SM58": ["shure sm58"],
    "KICK CAIXA TOM 1 TOM 2 SURDO RIDE CHIMBAL": ["kit microfone bateria shure pga drum"],
    "SHOWTECH ST X251LAYX2": ["showtech st-x251lay"],
    "GLOBO ESPELHADO COM MOTOR": ["globo espelhado com motor"],
    "S MODELO": [],
    "ETIQUETA LIDER COR COBRE": ["transformador fiolux 5000va"],
    "SEM ETIQUETA": [],
    "HAYONIK": ["extensor hayonik"],
    "LEACS": ["caixa leacs ativa"],
    "SENNHEISER": ["microfone sem fio sennheiser"],
    "MEGA LIGHT": ["par led mega light"],
    "STAR LIGHT DIVISION": ["star light division led spot rgbw 54x3w"],
}


@dataclass
class ItemRow:
    row_idx: int
    item_id: int
    nome: str | None
    categoria: str | None
    subcategoria: str | None
    marca: str | None
    modelo: str | None
    rastreamento: str | None
    quantidade: int | float | None
    valor: float | None

    @property
    def key(self) -> tuple[str | None, str | None, str | None, str | None, str | None]:
        return (self.nome, self.categoria, self.subcategoria, self.marca, self.modelo)


@dataclass
class SourceMatch:
    value: float | None
    method: str
    source_title: str
    source_url: str
    confidence: float
    notes: str = ""


def fold_text(value: str | None) -> str:
    if not value:
        return ""
    value = unicodedata.normalize("NFKD", str(value))
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = value.upper()
    value = re.sub(r"[^A-Z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def compact_text(value: str | None) -> str:
    return fold_text(value).replace(" ", "")


def tokenize(value: str | None) -> list[str]:
    return [token for token in fold_text(value).split() if token and token not in STOPWORDS]


def slugify_query(value: str) -> str:
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = value.lower()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    value = value.strip("-")
    return re.sub(r"-+", "-", value)


def safe_float(value: object) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 2)

    cleaned = str(value).strip().upper()
    if not cleaned or "N ACHEI" in cleaned or "NAO ACHEI" in cleaned or "PRECO NAO ENCONTRADO" in cleaned:
        return None

    cleaned = cleaned.replace("R$", "").replace(".", "").replace(",", ".")
    cleaned = re.sub(r"[^0-9.]+", "", cleaned)
    if not cleaned:
        return None

    try:
        return round(float(cleaned), 2)
    except ValueError:
        return None


def dedupe_parts(parts: Iterable[str | None]) -> str:
    tokens: list[str] = []
    seen: set[str] = set()
    for part in parts:
        if not part:
            continue
        for raw_token in str(part).split():
            token = raw_token.strip()
            if not token:
                continue
            key = fold_text(token)
            if key and key not in seen:
                seen.add(key)
                tokens.append(token)
    return " ".join(tokens).strip()


def canonical_subcategory(value: str | None) -> str:
    folded = fold_text(value)
    return SUBCATEGORY_ESTIMATE_ALIASES.get(folded, folded)


def strip_brand_tokens(name: str | None, brand: str | None) -> str:
    name_tokens = tokenize(name)
    brand_tokens = set(tokenize(brand))
    if not name_tokens:
        return ""
    if not brand_tokens:
        return " ".join(name_tokens)
    stripped = [token for token in name_tokens if token not in brand_tokens]
    return " ".join(stripped).strip()


def strip_brand_preserving_tokens(name: str | None, brand: str | None) -> str:
    folded_name = fold_text(name)
    folded_brand = fold_text(brand)
    if not folded_name:
        return ""
    if not folded_brand:
        return folded_name
    brand_tokens = set(folded_brand.split())
    stripped = [token for token in folded_name.split() if token not in brand_tokens]
    return " ".join(stripped).strip()


def extract_model_candidates(*values: str | None) -> list[str]:
    candidates: list[str] = []
    seen: set[str] = set()
    for value in values:
        folded = fold_text(value)
        if not folded:
            continue
        for token in MODEL_TOKEN_PATTERN.findall(folded):
            compact = token.replace("-", "").replace("/", "")
            if len(compact) < 4:
                continue
            if not re.search(r"[A-Z]", compact) or not re.search(r"\d", compact):
                continue
            normalized = token.replace("-", " ").replace("/", " ").strip()
            if normalized and normalized not in seen:
                seen.add(normalized)
                candidates.append(normalized)
    return candidates


def build_query_variants(item: ItemRow) -> list[str]:
    base = dedupe_parts([item.marca, item.modelo, item.nome])
    variants = [base] if base else []

    if item.nome and item.nome not in variants:
        variants.append(item.nome)
    if item.modelo and item.modelo not in variants:
        variants.append(item.modelo)
    stripped_name = strip_brand_preserving_tokens(item.nome, item.marca)
    if stripped_name and stripped_name not in variants:
        variants.append(stripped_name)
    for candidate in extract_model_candidates(item.nome, item.modelo):
        if candidate and candidate not in variants:
            variants.append(candidate)
    if item.marca and item.modelo:
        combined = dedupe_parts([item.marca, item.modelo])
        if combined and combined not in variants:
            variants.append(combined)
    sparse_identity = len(tokenize(base)) < 2
    if sparse_identity and item.nome and item.subcategoria:
        combined = dedupe_parts([item.nome, item.subcategoria])
        if combined and combined not in variants:
            variants.append(combined)
    if sparse_identity and item.marca and item.subcategoria:
        combined = dedupe_parts([item.marca, item.subcategoria])
        if combined and combined not in variants:
            variants.append(combined)
    if sparse_identity and item.subcategoria and item.subcategoria not in variants:
        variants.append(item.subcategoria)

    override_key = fold_text(item.nome or item.modelo or item.marca)
    for override in MANUAL_QUERY_OVERRIDES.get(override_key, []):
        if override and override not in variants:
            variants.insert(0, override)

    clean_variants: list[str] = []
    seen: set[str] = set()
    for variant in variants:
        if not variant:
            continue
        variant = re.sub(r"\s+", " ", variant).strip()
        if not variant:
            continue
        key = fold_text(variant)
        if key in GENERIC_QUERIES:
            continue
        if key and key not in seen:
            seen.add(key)
            clean_variants.append(variant)
    return clean_variants


def fetch_url(url: str, retries: int = 2, sleep_s: float = 0.4, timeout_s: int = 8) -> str:
    last_error: Exception | None = None
    for attempt in range(retries):
        request = Request(
            url,
            headers={
                "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
                "(KHTML, like Gecko) Chrome/124.0 Safari/537.36",
                "Accept-Language": "pt-BR,pt;q=0.9,en;q=0.8",
            },
        )
        try:
            with urlopen(request, timeout=timeout_s) as response:
                return response.read().decode("utf-8", errors="ignore")
        except HTTPError as exc:
            if exc.code == 404:
                return ""
            last_error = exc
            if attempt < retries - 1:
                time.sleep(sleep_s * (attempt + 1))
        except (URLError, TimeoutError) as exc:
            last_error = exc
            if attempt < retries - 1:
                time.sleep(sleep_s * (attempt + 1))
    raise RuntimeError(f"Falha ao buscar {url}: {last_error}") from last_error


def parse_mercadolivre_results(html: str) -> list[dict[str, object]]:
    blocks = re.findall(r'<li class="ui-search-layout__item[^>]*>(.*?)</li>', html, re.S)
    results: list[dict[str, object]] = []
    for block in blocks:
        title_match = re.search(r'class="poly-component__title"[^>]*>(.*?)</a>', block, re.S)
        href_match = re.search(r'<a href="([^"]+)"[^>]*class="poly-component__title"', block, re.S)
        money_values = []
        for fraction_raw in re.findall(r'andes-money-amount__fraction"[^>]*>(.*?)</span>', block, re.S):
            cleaned_fraction = re.sub(r"[^0-9]", "", fraction_raw)
            if cleaned_fraction:
                money_values.append(float(cleaned_fraction))

        if not (title_match and href_match and money_values):
            continue

        title = unescape(re.sub(r"<[^>]+>", " ", title_match.group(1)))
        title = re.sub(r"\s+", " ", title).strip()

        price = max(money_values)

        results.append(
            {
                "title": title,
                "url": unescape(href_match.group(1).split("#")[0]),
                "price": round(price, 2),
            }
        )
    return results


def score_candidate(query: str, item: ItemRow, title: str, price: float) -> float:
    norm_query = fold_text(query)
    norm_title = fold_text(title)
    compact_query = compact_text(query)
    compact_title = compact_text(title)
    if not norm_query or not norm_title:
        return -1.0

    query_tokens = [token for token in tokenize(norm_query) if len(token) > 1]
    title_tokens = [token for token in tokenize(norm_title) if len(token) > 1]
    if not query_tokens or not title_tokens:
        return -1.0

    query_set = set(query_tokens)
    title_set = set(title_tokens)
    overlap = query_set & title_set
    numeric_query = {token for token in query_set if any(ch.isdigit() for ch in token)}
    numeric_overlap = sum(1 for token in overlap if any(ch.isdigit() for ch in token))

    ratio = SequenceMatcher(None, norm_query, norm_title).ratio()
    coverage = len(overlap) / max(len(query_set), 1)
    precision = len(overlap) / max(len(title_set), 1)

    score = (ratio * 0.50) + (coverage * 0.32) + (precision * 0.08) + (min(numeric_overlap, 4) * 0.05)

    if numeric_query:
        numeric_coverage = numeric_overlap / len(numeric_query)
        score += numeric_coverage * 0.08
        if numeric_overlap == 0:
            score -= 0.25

    if len(compact_query) >= 6 and compact_query in compact_title:
        score += 0.10
    for hint in (item.modelo, item.nome):
        compact_hint = compact_text(hint)
        if compact_hint and len(compact_hint) >= 5 and compact_hint in compact_title:
            score += 0.08
            break

    title_only_negatives = {token for token in title_set if token in NEGATIVE_TOKENS and token not in query_set}
    score -= 0.18 * len(title_only_negatives)

    if price <= 0:
        score -= 1
    elif price < 20 and item.categoria != "CABO":
        score -= 0.15

    if item.categoria == "CABO" and "CABO" in title_set:
        score += 0.05

    if item.subcategoria:
        sub_tokens = set(tokenize(item.subcategoria))
        if sub_tokens & title_set:
            score += 0.04

    if item.marca:
        brand_tokens = set(tokenize(item.marca))
        if brand_tokens and brand_tokens <= title_set:
            score += 0.06

    return score


def choose_market_match(item: ItemRow, query: str, results: list[dict[str, object]]) -> SourceMatch | None:
    if not results:
        return None

    scored: list[tuple[float, dict[str, object]]] = []
    for result in results:
        score = score_candidate(query, item, str(result["title"]), float(result["price"]))
        scored.append((score, result))

    scored.sort(key=lambda pair: pair[0], reverse=True)
    best_score, best_result = scored[0]
    if best_score < 0.70:
        return None

    top_band = [pair for pair in scored if pair[0] >= max(best_score - 0.08, 0.70)]
    prices = [float(result["price"]) for _, result in top_band[:3]]
    if not prices:
        return None

    if len(prices) >= 2:
        value = round(statistics.median(prices), 2)
    else:
        value = round(prices[0], 2)

    notes = f"consulta={query}; resultados_considerados={len(top_band[:3])}; melhor_score={best_score:.3f}"
    return SourceMatch(
        value=value,
        method="mercadolivre_busca",
        source_title=str(best_result["title"]),
        source_url=str(best_result["url"]),
        confidence=round(min(best_score, 0.99), 3),
        notes=notes,
    )


def load_local_reference_sources(original_path: Path) -> tuple[dict[str, list[SourceMatch]], list[tuple[str, SourceMatch]]]:
    wb = load_workbook(original_path, data_only=True)
    exact_map: dict[str, list[SourceMatch]] = defaultdict(list)
    fuzzy_pool: list[tuple[str, SourceMatch]] = []

    def add_entry(label: str | None, value: object, sheet_name: str) -> None:
        price = safe_float(value)
        if not label or price is None:
            return
        normalized = fold_text(label)
        if not normalized:
            return
        source = SourceMatch(
            value=price,
            method=f"fonte_local_{sheet_name.lower().replace(' ', '_')}",
            source_title=str(label).strip(),
            source_url=f"arquivo_local:{original_path.name}#{sheet_name}",
            confidence=0.96 if sheet_name != "EQUIPAMENTOS" else 0.92,
        )
        exact_map[normalized].append(source)
        fuzzy_pool.append((normalized, source))

    ws = wb["EQUIPAMENTOS"]
    for row in ws.iter_rows(min_row=3, values_only=True):
        marca = row[4] if len(row) > 4 else None
        modelo = row[5] if len(row) > 5 else None
        valor = row[10] if len(row) > 10 else None
        label = dedupe_parts([marca, modelo])
        add_entry(label, valor, "EQUIPAMENTOS")
        add_entry(modelo, valor, "EQUIPAMENTOS")

    ws = wb["EQUIPAMENTOS - MAIO"]
    for row in ws.iter_rows(min_row=4, values_only=True):
        add_entry(row[0] if len(row) > 0 else None, row[3] if len(row) > 3 else None, "EQUIPAMENTOS - MAIO")

    ws = wb["LISTA PARA MANUS"]
    for row in ws.iter_rows(min_row=3, values_only=True):
        add_entry(row[0] if len(row) > 0 else None, row[1] if len(row) > 1 else None, "LISTA PARA MANUS")

    return exact_map, fuzzy_pool


def pick_best_local_match(item: ItemRow, exact_map: dict[str, list[SourceMatch]], fuzzy_pool: list[tuple[str, SourceMatch]]) -> SourceMatch | None:
    queries = build_query_variants(item)

    for query in queries:
        normalized = fold_text(query)
        if normalized in exact_map:
            candidates = exact_map[normalized]
            best = max(candidates, key=lambda match: match.confidence)
            return SourceMatch(
                value=best.value,
                method=best.method,
                source_title=best.source_title,
                source_url=best.source_url,
                confidence=best.confidence,
                notes=f"consulta={query}; match=exato",
            )

    best_ratio = 0.0
    best_match: SourceMatch | None = None
    for query in queries[:3]:
        normalized = fold_text(query)
        numeric_query = {token for token in tokenize(normalized) if any(ch.isdigit() for ch in token)}
        for label, source in fuzzy_pool:
            ratio = SequenceMatcher(None, normalized, label).ratio()
            if numeric_query:
                numeric_label = {token for token in tokenize(label) if any(ch.isdigit() for ch in token)}
                if numeric_query and numeric_query - numeric_label:
                    ratio -= 0.12
            if ratio > best_ratio:
                best_ratio = ratio
                best_match = source

    if best_match and best_ratio >= 0.88:
        return SourceMatch(
            value=best_match.value,
            method=f"{best_match.method}_fuzzy",
            source_title=best_match.source_title,
            source_url=best_match.source_url,
            confidence=round(best_ratio, 3),
            notes="match=fuzzy_fonte_local",
        )
    return None


def infer_from_known_items(item: ItemRow, known_values: list[tuple[ItemRow, float]]) -> SourceMatch | None:
    brand = fold_text(item.marca)
    subcat = canonical_subcategory(item.subcategoria)
    category = fold_text(item.categoria)

    same_brand_subcat = [
        value
        for other, value in known_values
        if fold_text(other.marca) == brand and canonical_subcategory(other.subcategoria) == subcat
    ]
    same_subcat = [value for other, value in known_values if canonical_subcategory(other.subcategoria) == subcat]
    same_category = [value for other, value in known_values if fold_text(other.categoria) == category]

    for label, candidates, confidence in [
        ("estimativa_brand_subcategoria", same_brand_subcat, 0.55),
        ("estimativa_subcategoria", same_subcat, 0.46),
        ("estimativa_categoria", same_category, 0.34),
    ]:
        minimum = 1 if label != "estimativa_categoria" else 2
        if len(candidates) >= minimum:
            value = round(statistics.median(candidates), 2)
            return SourceMatch(
                value=value,
                method=label,
                source_title=f"{item.subcategoria or item.categoria}",
                source_url="estimativa_interna",
                confidence=confidence,
                notes=f"estimativa baseada em {len(candidates)} itens conhecidos",
            )
    return None


def load_items_sheet(workbook_path: Path) -> tuple[object, object, list[ItemRow]]:
    wb = load_workbook(workbook_path)
    ws_items = wb["ITENS"]
    items: list[ItemRow] = []
    for row_idx, row in enumerate(ws_items.iter_rows(min_row=2, values_only=True), start=2):
        items.append(
            ItemRow(
                row_idx=row_idx,
                item_id=int(row[0]),
                nome=row[1],
                categoria=row[2],
                subcategoria=row[3],
                marca=row[4],
                modelo=row[5],
                rastreamento=row[6],
                quantidade=row[7],
                valor=safe_float(row[8]),
            )
        )
    return wb, ws_items, items


def write_audit_sheet(wb, audit_rows: list[dict[str, object]]) -> None:
    if AUDIT_SHEET in wb.sheetnames:
        del wb[AUDIT_SHEET]
    ws = wb.create_sheet(AUDIT_SHEET)
    headers = [
        "Item ID",
        "Nome",
        "Categoria",
        "Subcategoria",
        "Marca",
        "Modelo",
        "Valor anterior (R$)",
        "Valor final (R$)",
        "Metodo",
        "Confianca",
        "Fonte titulo",
        "Fonte URL",
        "Notas",
        "Consultado em",
    ]
    ws.append(headers)
    for row in audit_rows:
        ws.append([row.get(header) for header in headers])

    for column in ("A", "G", "H", "J"):
        ws.column_dimensions[column].width = 16
    for column in ("B", "C", "D", "E", "F", "I"):
        ws.column_dimensions[column].width = 28
    ws.column_dimensions["K"].width = 48
    ws.column_dimensions["L"].width = 64
    ws.column_dimensions["M"].width = 40
    ws.column_dimensions["N"].width = 22


def main() -> int:
    parser = argparse.ArgumentParser(description="Preenche valores unitarios faltantes no inventario MMD.")
    parser.add_argument("--cleaned", type=Path, default=DEFAULT_CLEANED)
    parser.add_argument("--original", type=Path, default=DEFAULT_ORIGINAL)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--limit", type=int, default=None, help="Processa apenas os primeiros N itens sem valor.")
    args = parser.parse_args()

    cleaned_path = args.cleaned.resolve()
    original_path = args.original.resolve()

    if not cleaned_path.exists():
        raise FileNotFoundError(f"Planilha limpa nao encontrada: {cleaned_path}")
    if not original_path.exists():
        raise FileNotFoundError(f"Planilha original nao encontrada: {original_path}")

    wb, ws_items, items = load_items_sheet(cleaned_path)
    ws_serial = wb["SERIAL NUMBERS"]

    exact_map, fuzzy_pool = load_local_reference_sources(original_path)
    search_cache: dict[str, list[dict[str, object]]] = {}

    known_values: list[tuple[ItemRow, float]] = [(item, item.valor) for item in items if item.valor is not None]
    item_value_map: dict[tuple[str | None, str | None, str | None, str | None, str | None], float] = {
        item.key: item.valor for item in items if item.valor is not None
    }
    audit_rows: list[dict[str, object]] = []

    missing_items = [item for item in items if item.valor is None]
    if args.limit is not None:
        missing_items = missing_items[: args.limit]

    consulted_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for index, item in enumerate(missing_items, start=1):
        current_match = pick_best_local_match(item, exact_map, fuzzy_pool)
        match = current_match

        if match is None:
            for query in build_query_variants(item):
                if not query:
                    continue
                cache_key = slugify_query(query)
                if cache_key not in search_cache:
                    url = f"https://lista.mercadolivre.com.br/{quote(cache_key, safe='-')}"
                    try:
                        html = fetch_url(url)
                    except RuntimeError:
                        search_cache[cache_key] = []
                    else:
                        search_cache[cache_key] = parse_mercadolivre_results(html) if html else []
                market_match = choose_market_match(item, query, search_cache[cache_key])
                if market_match is not None:
                    match = market_match
                    break

        if match is None:
            match = infer_from_known_items(item, known_values)

        if match is None or match.value is None:
            audit_rows.append(
                {
                    "Item ID": item.item_id,
                    "Nome": item.nome,
                    "Categoria": item.categoria,
                    "Subcategoria": item.subcategoria,
                    "Marca": item.marca,
                    "Modelo": item.modelo,
                    "Valor anterior (R$)": item.valor,
                    "Valor final (R$)": None,
                    "Metodo": "sem_match",
                    "Confianca": 0,
                    "Fonte titulo": "",
                    "Fonte URL": "",
                    "Notas": "nenhuma fonte automatica encontrada",
                    "Consultado em": consulted_at,
                }
            )
            continue

        value = round(float(match.value), 2)
        item_value_map[item.key] = value
        known_values.append((item, value))
        ws_items.cell(row=item.row_idx, column=9).value = value

        audit_rows.append(
            {
                "Item ID": item.item_id,
                "Nome": item.nome,
                "Categoria": item.categoria,
                "Subcategoria": item.subcategoria,
                "Marca": item.marca,
                "Modelo": item.modelo,
                "Valor anterior (R$)": item.valor,
                "Valor final (R$)": value,
                "Metodo": match.method,
                "Confianca": match.confidence,
                "Fonte titulo": match.source_title,
                "Fonte URL": match.source_url,
                "Notas": match.notes,
                "Consultado em": consulted_at,
            }
        )

        if index % 25 == 0:
            print(f"[fill-inventory-values] Processados {index}/{len(missing_items)} itens sem valor...")

    # Replica valor por item nos seriais que ainda estao vazios.
    serial_updates = 0
    for row_idx, row in enumerate(ws_serial.iter_rows(min_row=2, values_only=True), start=2):
        key = (row[1], row[2], row[3], row[4], row[5])
        current_value = safe_float(row[10])
        if current_value is not None:
            continue
        mapped_value = item_value_map.get(key)
        if mapped_value is None:
            continue
        ws_serial.cell(row=row_idx, column=11).value = mapped_value
        serial_updates += 1

    write_audit_sheet(wb, audit_rows)

    if args.dry_run:
        unresolved = sum(1 for row in audit_rows if row["Metodo"] == "sem_match")
        print(f"DRY RUN. Itens processados: {len(missing_items)}. Sem match: {unresolved}. Seriais atualizados: {serial_updates}.")
        return 0

    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_path = cleaned_path.with_name(f"{cleaned_path.stem}.backup-{timestamp}{cleaned_path.suffix}")
    shutil.copy2(cleaned_path, backup_path)
    wb.save(cleaned_path)

    unresolved = sum(1 for row in audit_rows if row["Metodo"] == "sem_match")
    print(f"Backup criado em: {backup_path}")
    print(f"Itens sem valor processados: {len(missing_items)}")
    print(f"Itens ainda sem match: {unresolved}")
    print(f"Seriais atualizados: {serial_updates}")
    print(f"Aba de auditoria: {AUDIT_SHEET}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
