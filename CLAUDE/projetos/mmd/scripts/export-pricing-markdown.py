#!/usr/bin/env python3
from __future__ import annotations

import csv
import importlib.util
import io
import statistics
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


THIS_DIR = Path(__file__).resolve().parent
PROJECT_DIR = THIS_DIR.parent
WORKBOOK_PATH = PROJECT_DIR / "data" / "inventario-limpo.xlsx"
ORIGINAL_PATH = PROJECT_DIR / "data" / "inventario-original.xlsx"
OUTPUT_PATH = PROJECT_DIR / "data" / "valores-para-importacao.md"
HELPER_PATH = THIS_DIR / "fill-inventory-values.py"
HEADER_ROW = 4
DATA_START_ROW = 5
SKIP_SHEETS = {"MANUAL", "DASHBOARD", "LOTES", "REF CATEGORIAS"}
STATE_FACTORS = {
    "NOVO": 1.0,
    "SEMI_NOVO": 0.85,
    "USADO": 0.65,
    "RECONDICIONADO": 0.50,
}
MANUAL_SOURCE_OVERRIDES = [
    {
        "sheet": "ILUMINACAO",
        "name": "Showtech ST-XQDFS24 - BLINDADA",
        "subcategory": "Fixa",
        "brand": "Showtech",
        "value": 1200.00,
        "method": "manual_web_curado",
        "source_title": "Ribalta Led RGBW 24x12w Pixel Quadriled 4n1 Outdoor IP67",
        "source_url": "https://www.mercadolivre.com.br/ribalta-led-rgbw-24x12w-pixel-quadriled-4n1-outdoor-ip67/p/MLB50404325",
        "confidence": 0.96,
    },
    {
        "sheet": "ILUMINACAO",
        "name": "Tomate MLG-119",
        "subcategory": "Outdoor",
        "brand": "Tomate",
        "value": 192.14,
        "method": "manual_web_curado",
        "source_title": "Luminaria Holofote Lampiao Lanterna Tomate MLG-119",
        "source_url": "https://www.mercadolivre.com.br/luminaria-holofote-lampiao-lanterna-tomate-mlg-119/p/MLB24716343",
        "confidence": 0.93,
    },
    {
        "sheet": "ILUMINACAO",
        "name": "Luatek LK192",
        "subcategory": "Mesa De Luz",
        "brand": "Luatek",
        "value": 412.95,
        "method": "manual_web_curado",
        "source_title": "Controladora Mesa Dmx 192 Canais Luatek LK-192",
        "source_url": "https://www.aliexpress.com/i/1005004883502576.html",
        "confidence": 0.90,
    },
    {
        "sheet": "ILUMINACAO",
        "name": "Showtech ST-135BY-CS",
        "subcategory": "5R",
        "brand": "Showtech",
        "value": 780.00,
        "method": "manual_web_curado",
        "source_title": "Moving Spot 120w Profissional ST-135BY",
        "source_url": "https://www.importlight.com.br/laser-show-e-efeitos-led/moving-spot-120w-profissional-showtech-st-135by",
        "confidence": 0.82,
    },
    {
        "sheet": "ACESSORIO",
        "name": "Showtech ST-135BY-CS",
        "subcategory": "Moving Beam",
        "brand": "Showtech",
        "value": 780.00,
        "method": "manual_web_curado",
        "source_title": "Moving Spot 120w Profissional ST-135BY",
        "source_url": "https://www.importlight.com.br/laser-show-e-efeitos-led/moving-spot-120w-profissional-showtech-st-135by",
        "confidence": 0.82,
    },
    {
        "sheet": "ENERGIA",
        "name": "Wireconnex WPD-5",
        "subcategory": "Regua",
        "brand": "Wireconnex",
        "value": 215.14,
        "method": "manual_web_curado",
        "source_title": "Regua Wireconex WPD-5/A com 5 tomadas",
        "source_url": "https://www.magazineluiza.com.br/regua-wireconex-wpd-5-a-com-5-tomadas/p/hf50662937/in/fdla/",
        "confidence": 0.94,
    },
    {
        "sheet": "ESTRUTURA",
        "name": "Staner",
        "subcategory": "Tripe De Caixa",
        "brand": "Staner",
        "value": 270.00,
        "method": "manual_web_curado",
        "source_title": "Suporte tripe para caixa de som Staner extensivel",
        "source_url": "https://www.krunner.com.br/suporte-tripe-para-caixa-de-som-staner-extensivel",
        "confidence": 0.92,
    },
    {
        "sheet": "ESTRUTURA",
        "name": "RMV",
        "subcategory": "Tripe De Microfone",
        "brand": "RMV",
        "value": 137.90,
        "method": "manual_web_curado",
        "source_title": "Suporte pedestal PSSU00142 tripe para microfone girafa cachimbo RMV",
        "source_url": "https://www.casadamusica.com.br/suporte-pedestal-pssu00142-trip%C3%A9-para-microfone-girafa-cachimbo-rmv",
        "confidence": 0.92,
    },
    {
        "sheet": "AUDIO",
        "name": "Pioneer FLX4",
        "subcategory": "Cdj/Xdj/Ddj",
        "brand": "Pioneer",
        "value": 3258.00,
        "method": "manual_web_curado",
        "source_title": "Controlador Pioneer DJ DDJ FLX4 preto",
        "source_url": "https://lista.mercadolivre.com.br/ddj-flx4",
        "confidence": 0.93,
    },
    {
        "sheet": "AUDIO",
        "name": "RCF Art 712-A",
        "subcategory": "Caixa De Som",
        "brand": "RCF",
        "value": 6349.00,
        "method": "manual_web_curado",
        "source_title": "Alto falante RCF ART 712-A MK4 black 127V",
        "source_url": "https://www.mercadolivre.com.br/alto-falante-rcf-art-712-a-mk4-black-127v/p/MLB15541412",
        "confidence": 0.94,
    },
    {
        "sheet": "FORA DE OPERACAO",
        "name": "Mark Audio CA-1000",
        "subcategory": "Caixa De Som",
        "brand": "Mark Audio",
        "value": 850.00,
        "method": "manual_web_curado",
        "source_title": "mark audio ca 1000",
        "source_url": "https://somaovivo.org/forum/threads/mark-audio-ca-1000.3382/",
        "confidence": 0.72,
    },
    {
        "sheet": "AUDIO",
        "name": "Mackie PROFX6V3",
        "subcategory": "Effects",
        "brand": "Mackie",
        "value": 1790.00,
        "method": "manual_web_curado",
        "source_title": "Mesa de som analogica Mackie ProFX6v3",
        "source_url": "https://www.audiovideoecia.com.br/MLB-3984812867-mesa-de-som-analogica-mackie-profx6v3-com-efeitos-e-usb-_JM",
        "confidence": 0.91,
    },
    {
        "sheet": "AUDIO",
        "name": "Yamaha YAMAHA MG10XU",
        "subcategory": "Mesa De Som",
        "brand": "Yamaha",
        "value": 1849.00,
        "method": "manual_web_curado",
        "source_title": "Mesa Yamaha Mixer MG10XU 10 canais com efeitos",
        "source_url": "https://lista.mercadolivre.com.br/yamaha-mg10xu",
        "confidence": 0.90,
    },
    {
        "sheet": "AUDIO",
        "name": "RCF Evox J8",
        "subcategory": "Line Vertical",
        "brand": "RCF",
        "value": 10990.00,
        "method": "manual_web_curado",
        "source_title": "Caixa de som ativa RCF Evox J8 1400W com bag",
        "source_url": "https://www.mossmusic.com.br/produtos/caixa-de-som-ativa-rcf-evox-j8-1400w-1x12-1x8-bag",
        "confidence": 0.90,
    },
    {
        "sheet": "AUDIO",
        "name": "Shure SM58",
        "subcategory": "Microfone Com Fio",
        "brand": "Shure",
        "value": 1119.00,
        "method": "manual_web_curado",
        "source_title": "Microfone Shure SM58 LC preto",
        "source_url": "https://lista.mercadolivre.com.br/shure-sm58",
        "confidence": 0.92,
    },
    {
        "sheet": "AUDIO",
        "name": "Shure GLXD4R",
        "subcategory": "Base Wireless",
        "brand": "Shure",
        "value": 4199.00,
        "method": "manual_web_curado",
        "source_title": "Receptor sem fio GLXD4R Plus dual band rackmount Shure",
        "source_url": "https://www.ciadosom.com.br/produto/microfone-shure-glxd4r-wireless-receiver-banda-z2",
        "confidence": 0.90,
    },
    {
        "sheet": "AUDIO",
        "name": "Phenyx Pro PHENYX PRO PTU-7000",
        "subcategory": "Microfone Sem Fio",
        "brand": "Phenyx Pro",
        "value": 2480.00,
        "method": "manual_web_curado",
        "source_title": "Microfone sem fio UHF 4 canais Phenyx Pro PTU-7000-4H",
        "source_url": "https://produto.mercadolivre.com.br/MLB-4115887984-microfone-sem-fio-uhf-4-canais-profissional-dj-karaoke-_JM",
        "confidence": 0.86,
    },
    {
        "sheet": "AUDIO",
        "name": "PARA MICROFONE SM58",
        "subcategory": "Capsula Mic",
        "brand": "",
        "value": 1600.00,
        "method": "manual_web_curado",
        "source_title": "Capsula para microfone sem fio Shure Beta 58A GLXD2",
        "source_url": "https://www.mercadolivre.com.br/capsula-para-microfone-sem-fio-shure-beta-58a-glxd2-z2/p/MLB26825105",
        "confidence": 0.83,
    },
    {
        "sheet": "AUDIO",
        "name": "Shure Beta 58A + GLXD2 Z2",
        "subcategory": "Microfone Wireless",
        "brand": "Shure",
        "value": 3290.00,
        "method": "manual_web_curado",
        "source_title": "Microfone transmissor Shure GLXD2 Beta58A",
        "source_url": "https://www.cariocainstrumentos.com.br/microfone-transmissor-shure-glxd2-beta58a-z2",
        "confidence": 0.88,
    },
]


@dataclass
class SerialRecord:
    sheet: str
    row_idx: int
    codigo: str
    nome: str
    subcategoria: str | None
    marca: str | None
    estado: str | None
    desgaste: float | int | None
    valor_original: float | None

    @property
    def key(self) -> tuple[str, str, str | None, str | None]:
        return (self.sheet, self.nome, self.subcategoria, self.marca)


def load_helpers():
    spec = importlib.util.spec_from_file_location("fill_inventory_values", HELPER_PATH)
    module = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def find_header_row(ws) -> int | None:
    for row_idx in range(1, 8):
        values = [cell.value for cell in ws[row_idx]]
        if "Codigo" in values and "Nome" in values and "Valor Original" in values:
            return row_idx
    return None


def parse_serial_rows(path: Path, helpers) -> list[SerialRecord]:
    wb = load_workbook(path, data_only=True)
    serials: list[SerialRecord] = []
    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        header_row = find_header_row(ws)
        if header_row is None:
            continue
        headers = [cell.value for cell in ws[header_row]]
        if "Codigo" not in headers or "Nome" not in headers:
            continue

        idx = {name: headers.index(name) for name in headers if name}
        data_start = header_row + 1
        for row_idx, row in enumerate(ws.iter_rows(min_row=data_start, values_only=True), start=data_start):
            codigo = row[idx["Codigo"]] if idx.get("Codigo") is not None else None
            nome = row[idx["Nome"]] if idx.get("Nome") is not None else None
            if not codigo or not nome:
                continue
            serials.append(
                SerialRecord(
                    sheet=sheet_name,
                    row_idx=row_idx,
                    codigo=str(codigo).strip(),
                    nome=str(nome).strip(),
                    subcategoria=row[idx["Subcategoria"]] if idx.get("Subcategoria") is not None else None,
                    marca=row[idx["Marca"]] if idx.get("Marca") is not None else None,
                    estado=row[idx["Estado"]] if idx.get("Estado") is not None else None,
                    desgaste=row[idx["Desgaste"]] if idx.get("Desgaste") is not None else None,
                    valor_original=helpers.safe_float(row[idx["Valor Original"]]) if idx.get("Valor Original") is not None else None,
                )
            )
    return serials


def build_unique_items(serials: list[SerialRecord], helpers):
    grouped: dict[tuple[str, str, str | None, str | None], list[SerialRecord]] = defaultdict(list)
    for record in serials:
        grouped[record.key].append(record)

    items = []
    for item_id, (key, rows) in enumerate(grouped.items(), start=1):
        sheet, nome, subcategoria, marca = key
        existing_values = [row.valor_original for row in rows if row.valor_original is not None]
        item = helpers.ItemRow(
            row_idx=rows[0].row_idx,
            item_id=item_id,
            nome=nome,
            categoria=sheet,
            subcategoria=subcategoria,
            marca=marca,
            modelo=nome,
            rastreamento="INDIVIDUAL",
            quantidade=len(rows),
            valor=statistics.median(existing_values) if existing_values else None,
        )
        items.append((key, item, rows))
    return items


def normalize_identity(sheet: str | None, name: str | None, subcategory: str | None, brand: str | None, helpers) -> tuple[str, str, str, str]:
    return (
        helpers.fold_text(sheet),
        helpers.fold_text(name),
        helpers.fold_text(subcategory),
        helpers.fold_text(brand),
    )


def build_manual_override_map(helpers) -> dict[tuple[str, str, str, str], object]:
    overrides = {}
    for entry in MANUAL_SOURCE_OVERRIDES:
        key = normalize_identity(entry["sheet"], entry["name"], entry["subcategory"], entry["brand"], helpers)
        overrides[key] = helpers.SourceMatch(
            value=entry["value"],
            method=entry["method"],
            source_title=entry["source_title"],
            source_url=entry["source_url"],
            confidence=entry["confidence"],
            notes="override manual curado",
        )
    return overrides


def estimate_cable_value(name: str, subcategory: str | None, helpers) -> float:
    folded_name = helpers.fold_text(name)
    folded_subcat = helpers.fold_text(subcategory)

    match = helpers.re.search(r"(\d+)\s*[.,]?\s*(\d*)\s*(M|CM)", folded_name)
    meters = 1.0
    if match:
        whole = match.group(1)
        decimal = match.group(2)
        unit = match.group(3)
        value = float(f"{whole}.{decimal}") if decimal else float(whole)
        meters = value / 100 if unit == "CM" else value
    elif "10M" in helpers.compact_text(name):
        meters = 10

    base_by_subcat = {
        "AC": 12,
        "AC 4X": 28,
        "AC 8": 45,
        "ADAPTADOR DIVISOR P10": 20,
        "DISPLAY PORT": 26,
        "DMX": 35,
        "DVI": 24,
        "DVI HDMI": 28,
        "EXTENSOR AC": 18,
        "EXTENSOR HDMI": 32,
        "EXTENSOR USB": 28,
        "EXTENSOR USB A": 28,
        "GERAL": 25,
        "HDMI": 20,
        "HDMI DISPLAY PORT": 30,
        "MIDI DIN": 20,
        "MINI DISPLAYPORT": 26,
        "MINI DPPORT P DPPORT": 26,
        "P10": 20,
        "P10 P10": 24,
        "P10 P2": 22,
        "P10 XLRF": 30,
        "P10 XLRM": 30,
        "P2 P2": 18,
        "P2 USB A": 16,
        "PARALELO": 12,
        "POWERCON AC": 36,
        "POWERCON PIAL": 42,
        "RCA P10": 20,
        "RCA P2": 18,
        "RCA RCA": 18,
        "SPEAK ON": 42,
        "USB A MICRO USB B": 18,
        "USB A USB A": 18,
        "USB A USB B": 18,
        "USB A USB C": 18,
        "USB B USB A": 18,
        "USB B USB C": 20,
        "USB USB": 18,
        "VGA": 22,
        "XLR": 36,
        "XLR P10": 30,
        "XLR P2": 24,
        "XLR USB C": 26,
        "XLRM P10": 30,
        "XLRM RCA": 24,
        "XLRM USB C": 26,
    }
    per_meter_by_subcat = {
        "AC": 7,
        "AC 4X": 14,
        "AC 8": 18,
        "DISPLAY PORT": 9,
        "DMX": 10,
        "DVI": 8,
        "DVI HDMI": 9,
        "EXTENSOR AC": 8,
        "EXTENSOR HDMI": 10,
        "EXTENSOR USB": 10,
        "EXTENSOR USB A": 10,
        "HDMI": 8,
        "HDMI DISPLAY PORT": 10,
        "MIDI DIN": 8,
        "MINI DISPLAYPORT": 9,
        "MINI DPPORT P DPPORT": 9,
        "P10": 8,
        "P10 P10": 8,
        "P10 P2": 7,
        "P10 XLRF": 10,
        "P10 XLRM": 10,
        "P2 P2": 6,
        "P2 USB A": 6,
        "PARALELO": 5,
        "POWERCON AC": 12,
        "POWERCON PIAL": 14,
        "RCA P10": 7,
        "RCA P2": 6,
        "RCA RCA": 6,
        "SPEAK ON": 12,
        "USB A MICRO USB B": 6,
        "USB A USB A": 6,
        "USB A USB B": 6,
        "USB A USB C": 6,
        "USB B USB A": 6,
        "USB B USB C": 6,
        "USB USB": 6,
        "VGA": 8,
        "XLR": 11,
        "XLR P10": 10,
        "XLR P2": 8,
        "XLR USB C": 8,
        "XLRM P10": 10,
        "XLRM RCA": 8,
        "XLRM USB C": 8,
    }

    base = base_by_subcat.get(folded_subcat, 25)
    per_meter = per_meter_by_subcat.get(folded_subcat, 7)
    value = round(base + (meters * per_meter), 2)
    if "BRANCO" in folded_name or "CINZA" in folded_name:
        value += 2
    return value


def fallback_estimate(item, helpers, known_values, global_values) -> helpers.SourceMatch:
    inferred = helpers.infer_from_known_items(item, known_values)
    if inferred and inferred.value is not None:
        return inferred
    value = round(statistics.median(global_values), 2) if global_values else 100.0
    return helpers.SourceMatch(
        value=value,
        method="estimativa_global",
        source_title=item.subcategoria or item.categoria,
        source_url="estimativa_interna",
        confidence=0.2,
        notes="fallback global para nao deixar campo vazio",
    )


def compute_current_value(valor_original: float, estado: str | None, desgaste: float | int | None) -> tuple[float, float]:
    factor = STATE_FACTORS.get((estado or "USADO").strip().upper(), 0.65)
    wear = float(desgaste if desgaste is not None else 3)
    wear = max(1.0, min(wear, 5.0))
    dep_pct = round((wear / 5.0) * factor * 100.0, 2)
    valor_atual = round(valor_original * dep_pct / 100.0, 2)
    return valor_atual, dep_pct


def export_markdown(output_path: Path, rows: list[dict[str, object]], summary: dict[str, object], review_rows: list[dict[str, object]]) -> None:
    csv_buffer = io.StringIO()
    writer = csv.DictWriter(
        csv_buffer,
        fieldnames=[
            "Aba",
            "Codigo",
            "Nome",
            "Subcategoria",
            "Marca",
            "Valor Original (R$)",
            "Valor Atual (R$)",
            "Deprec. (%)",
            "Metodo",
            "Confianca",
            "Fonte Titulo",
            "Fonte URL",
        ],
    )
    writer.writeheader()
    for row in rows:
        writer.writerow(row)

    review_buffer = io.StringIO()
    review_writer = csv.DictWriter(
        review_buffer,
        fieldnames=[
            "Aba",
            "Nome",
            "Subcategoria",
            "Marca",
            "Valor Original (R$)",
            "Metodo",
            "Confianca",
            "Fonte Titulo",
            "Fonte URL",
        ],
    )
    review_writer.writeheader()
    for row in review_rows:
        review_writer.writerow(row)

    lines = [
        "# Valores Para Importacao",
        "",
        f"Gerado em: {summary['generated_at']}",
        f"Arquivo-base: `{summary['workbook']}`",
        "",
        "## Resumo",
        f"- Seriais exportados: {summary['serial_count']}",
        f"- Itens unicos avaliados: {summary['unique_items']}",
        f"- Valor Original ja existente na planilha: {summary['existing_original']}",
        f"- Valor Original preenchido via busca local/web/estimativa: {summary['filled_original']}",
        f"- Distribuicao por metodo: {summary['methods']}",
        "",
        "## CSV Para Importacao",
        "",
        "```csv",
        csv_buffer.getvalue().rstrip(),
        "```",
        "",
        "## Revisao Prioritaria",
        "",
        "Linhas abaixo sao as que vieram de estimativa ou confianca abaixo de 0.70.",
        "",
        "```csv",
        review_buffer.getvalue().rstrip(),
        "```",
        "",
    ]
    output_path.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    helpers = load_helpers()
    serials = parse_serial_rows(WORKBOOK_PATH, helpers)
    unique_items = build_unique_items(serials, helpers)
    exact_map, fuzzy_pool = helpers.load_local_reference_sources(ORIGINAL_PATH)
    manual_overrides = build_manual_override_map(helpers)
    search_cache: dict[str, list[dict[str, object]]] = {}

    known_values = []
    global_values = []
    resolved_values: dict[tuple[str, str, str | None, str | None], tuple[float, object]] = {}

    for _, item, _ in unique_items:
        if item.valor is not None:
            known_values.append((item, item.valor))
            global_values.append(item.valor)
            resolved_values[(item.categoria, item.nome, item.subcategoria, item.marca)] = (
                item.valor,
                helpers.SourceMatch(
                    value=item.valor,
                    method="valor_original_existente",
                    source_title="valor já presente na planilha",
                    source_url="arquivo_local:inventario-limpo.xlsx",
                    confidence=1.0,
                    notes="mantido",
                ),
            )

    for _, item, _ in unique_items:
        key = (item.categoria, item.nome, item.subcategoria, item.marca)
        if key in resolved_values:
            continue

        match = manual_overrides.get(normalize_identity(item.categoria, item.nome, item.subcategoria, item.marca, helpers))

        if match is None:
            match = helpers.pick_best_local_match(item, exact_map, fuzzy_pool)

        if match is None and item.categoria == "CABO":
            value = estimate_cable_value(item.nome or "", item.subcategoria, helpers)
            match = helpers.SourceMatch(
                value=value,
                method="estimativa_cabo_regra",
                source_title=item.subcategoria or "CABO",
                source_url="estimativa_interna",
                confidence=0.35,
                notes="regra por tipo de cabo e metragem",
            )

        if match is None and item.categoria != "CABO":
            for query in helpers.build_query_variants(item):
                cache_key = helpers.slugify_query(query)
                if cache_key not in search_cache:
                    url = f"https://lista.mercadolivre.com.br/{helpers.quote(cache_key, safe='-')}"
                    try:
                        html = helpers.fetch_url(url)
                    except RuntimeError:
                        search_cache[cache_key] = []
                    else:
                        search_cache[cache_key] = helpers.parse_mercadolivre_results(html) if html else []
                market_match = helpers.choose_market_match(item, query, search_cache[cache_key])
                if market_match is not None:
                    match = market_match
                    break

        if match is None:
            match = fallback_estimate(item, helpers, known_values, global_values)

        assert match.value is not None
        value = round(float(match.value), 2)
        resolved_values[key] = (value, match)
        known_values.append((item, value))
        global_values.append(value)

    export_rows = []
    review_rows = []
    seen_review = set()
    method_counter = Counter()

    for record in serials:
        key = (record.sheet, record.nome, record.subcategoria, record.marca)
        valor_original, match = resolved_values[key]
        valor_atual, dep_pct = compute_current_value(valor_original, record.estado, record.desgaste)
        row = {
            "Aba": record.sheet,
            "Codigo": record.codigo,
            "Nome": record.nome,
            "Subcategoria": record.subcategoria or "",
            "Marca": record.marca or "",
            "Valor Original (R$)": f"{valor_original:.2f}",
            "Valor Atual (R$)": f"{valor_atual:.2f}",
            "Deprec. (%)": f"{dep_pct:.2f}",
            "Metodo": match.method,
            "Confianca": f"{match.confidence:.3f}",
            "Fonte Titulo": match.source_title,
            "Fonte URL": match.source_url,
        }
        export_rows.append(row)
        method_counter[match.method] += 1
        if match.confidence < 0.70 or str(match.method).startswith("estimativa"):
            review_key = (
                record.sheet,
                record.nome,
                record.subcategoria or "",
                record.marca or "",
                f"{valor_original:.2f}",
                match.method,
                f"{match.confidence:.3f}",
                match.source_title,
                match.source_url,
            )
            if review_key not in seen_review:
                seen_review.add(review_key)
                review_rows.append(
                    {
                        "Aba": record.sheet,
                        "Nome": record.nome,
                        "Subcategoria": record.subcategoria or "",
                        "Marca": record.marca or "",
                        "Valor Original (R$)": f"{valor_original:.2f}",
                        "Metodo": match.method,
                        "Confianca": f"{match.confidence:.3f}",
                        "Fonte Titulo": match.source_title,
                        "Fonte URL": match.source_url,
                    }
                )

    summary = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "workbook": str(WORKBOOK_PATH),
        "serial_count": len(serials),
        "unique_items": len(unique_items),
        "existing_original": sum(1 for _, item, _ in unique_items if item.valor is not None),
        "filled_original": sum(1 for _, item, _ in unique_items if item.valor is None),
        "methods": ", ".join(f"{name}={count}" for name, count in method_counter.most_common()),
    }
    export_markdown(OUTPUT_PATH, export_rows, summary, review_rows)
    print(f"Markdown gerado em: {OUTPUT_PATH}")
    print(f"Seriais exportados: {len(serials)}")
    print(f"Itens unicos: {len(unique_items)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
