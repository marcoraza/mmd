#!/usr/bin/env python3
from __future__ import annotations

import csv
import importlib.util
import io
import re
import statistics
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


THIS_DIR = Path(__file__).resolve().parent
PROJECT_DIR = THIS_DIR.parent
WORKBOOK_PATH = PROJECT_DIR / "data" / "inventario-limpo.xlsx"
ORIGINAL_PATH = PROJECT_DIR / "data" / "inventario-original.xlsx"
OUTPUT_PATH = PROJECT_DIR / "data" / "valores-para-importacao.md"
CURRENT_MARKET_OVERRIDES_PATH = PROJECT_DIR / "data" / "current-market-overrides.csv"
HELPER_PATH = THIS_DIR / "fill-inventory-values.py"
HEADER_ROW = 4
DATA_START_ROW = 5
SKIP_SHEETS = {"MANUAL", "DASHBOARD", "LOTES", "REF CATEGORIAS"}
STATE_FACTORS = {
    "NOVO": 1.0,
    "SEMI_NOVO": 0.85,
    "USADO": 0.65,
    "RECONDICIONADO": 0.50,
}
MANUAL_SOURCE_OVERRIDES = [
    {
        "sheet": "ILUMINACAO",
        "name": "Showtech ST-XQDFS24 - BLINDADA",
        "subcategory": "Fixa",
        "brand": "Showtech",
        "value": 1200.00,
        "method": "manual_web_curado",
        "source_title": "Ribalta Led RGBW 24x12w Pixel Quadriled 4n1 Outdoor IP67",
        "source_url": "https://www.mercadolivre.com.br/ribalta-led-rgbw-24x12w-pixel-quadriled-4n1-outdoor-ip67/p/MLB50404325",
        "confidence": 0.96,
    },
    {
        "sheet": "ILUMINACAO",
        "name": "Tomate MLG-119",
        "subcategory": "Outdoor",
        "brand": "Tomate",
        "value": 192.14,
        "method": "manual_web_curado",
        "source_title": "Luminaria Multifuncional Recarregavel Tomate MLG-119",
        "source_url": "https://www.magazineluiza.com.br/luminaria-multifuncional-recarregavel-tomate-mlg-119-iluminacao-versatil-e-eficiente/p/cd551acd91/cj/lumt/",
        "confidence": 0.95,
    },
    {
        "sheet": "ILUMINACAO",
        "name": "Luatek LK192",
        "subcategory": "Mesa De Luz",
        "brand": "Luatek",
        "value": 412.95,
        "method": "manual_web_curado",
        "source_title": "Controladora Mesa Dmx 192 Canais Luatek LK-192",
        "source_url": "https://www.magazineluiza.com.br/mesa-profissional-dmx-512-controladora-de-efeitos-festa-acompanha-fonte-12v-lk192-luatek/p/hbc68c64kg/cj/meci/",
        "confidence": 0.95,
    },
    {
        "sheet": "ILUMINACAO",
        "name": "Showtech ST-135BY-CS",
        "subcategory": "5R",
        "brand": "Showtech",
        "value": 950.00,
        "method": "manual_web_curado",
        "source_title": "Moving Beam LED 120W com Borda de LED ST-135BY",
        "source_url": "https://www.vimatlux.com.br/moving-beam-led-100w-com-borda-de-led-st-135by",
        "confidence": 0.92,
    },
    {
        "sheet": "ILUMINACAO",
        "name": "Showtech ST-184N1",
        "subcategory": "Indoor",
        "brand": "Showtech",
        "value": 289.00,
        "method": "manual_web_curado",
        "source_title": "Par Led 18x12w RGBW Showtech ST-184N1",
        "source_url": "https://dplaymusic.com.br/index.php?product_id=730638184&route=product%2Fproduct",
        "confidence": 0.95,
    },
    {
        "sheet": "ILUMINACAO",
        "name": "ST-LS6",
        "subcategory": "Laser",
        "brand": "Showtech",
        "value": 2700.00,
        "method": "manual_web_curado",
        "source_title": "Super Ribalta Moving Laser 6 Saidas RGB Ribalta Tilt Pan",
        "source_url": "https://www.09iluminacao.com.br/super-ribalta-moving-laser-6-saidas-rgb-ribalta-tilt-pan/prod-9320384/",
        "confidence": 0.90,
    },
    {
        "sheet": "ILUMINACAO",
        "name": "ST-960PS",
        "subcategory": "Strobo",
        "brand": "Showtech",
        "value": 590.00,
        "method": "manual_web_curado",
        "source_title": "Super Strobo Led Atomic 1000w RGBW Showtech ST-960P9",
        "source_url": "https://www.mercadolivre.com.br/super-strobo-led-atomic-1000w-rgbw-showtech-st-960p9/p/MLB47884745",
        "confidence": 0.92,
    },
    {
        "sheet": "ILUMINACAO",
        "name": "LP-354",
        "subcategory": "Outdoor",
        "brand": "Hot Machine",
        "value": 450.00,
        "method": "manual_web_curado",
        "source_title": "Kit 10 Par Led Hot Machine LP 354 usado na OLX",
        "source_url": "https://sp.olx.com.br/sao-paulo-e-regiao/audio/equipamentos-e-acessorios-de-som/kit-10-par-led-hot-machine-lp-354-1480290069",
        "confidence": 0.88,
    },
    {
        "sheet": "ILUMINACAO",
        "name": "ST-32T",
        "subcategory": "Outdoor",
        "brand": "Showtech",
        "value": 664.99,
        "method": "manual_web_curado",
        "source_title": "Canhao Refletor Outdoor 12x15w IP65 RGBWA 6n1 DMX, equivalente de mercado",
        "source_url": "https://lista.mercadolivre.com.br/par-led-outdoor",
        "confidence": 0.72,
    },
    {
        "sheet": "ILUMINACAO",
        "name": "NE-117G-I",
        "subcategory": "Indoor",
        "brand": "BDL",
        "value": 171.00,
        "method": "manual_web_curado",
        "source_title": "Canhao PAR Showtec 18x12w 18 LEDs RGBW DMX com Flash",
        "source_url": "https://www.mercadolivre.com.br/canho-par-showtec-18x12w-18-leds-rgbw-dmx-com-flash/p/MLB22509686",
        "confidence": 0.78,
    },
    {
        "sheet": "ILUMINACAO",
        "name": "MEGALIGHT NE-117G-I",
        "subcategory": "Indoor",
        "brand": "BDL",
        "value": 171.00,
        "method": "manual_web_curado",
        "source_title": "Canhao PAR Showtec 18x12w 18 LEDs RGBW DMX com Flash",
        "source_url": "https://www.mercadolivre.com.br/canho-par-showtec-18x12w-18-leds-rgbw-dmx-com-flash/p/MLB22509686",
        "confidence": 0.74,
    },
    {
        "sheet": "ILUMINACAO",
        "name": "SEM IDENTIFICACAO (MAS PARECE NE-117G-I)",
        "subcategory": "Indoor",
        "brand": "",
        "value": 171.00,
        "method": "manual_web_curado",
        "source_title": "Canhao PAR Showtec 18x12w 18 LEDs RGBW DMX com Flash",
        "source_url": "https://www.mercadolivre.com.br/canho-par-showtec-18x12w-18-leds-rgbw-dmx-com-flash/p/MLB22509686",
        "confidence": 0.70,
    },
    {
        "sheet": "ACESSORIO",
        "name": "Showtech ST-135BY-CS",
        "subcategory": "Moving Beam",
        "brand": "Showtech",
        "value": 950.00,
        "method": "manual_web_curado",
        "source_title": "Moving Beam LED 120W com Borda de LED ST-135BY",
        "source_url": "https://www.vimatlux.com.br/moving-beam-led-100w-com-borda-de-led-st-135by",
        "confidence": 0.92,
    },
    {
        "sheet": "ENERGIA",
        "name": "Wireconnex WPD-5",
        "subcategory": "Regua",
        "brand": "Wireconnex",
        "value": 215.14,
        "method": "manual_web_curado",
        "source_title": "Regua Wireconex WPD-5/A com 5 tomadas",
        "source_url": "https://www.magazineluiza.com.br/regua-wireconex-wpd-5-a-com-5-tomadas/p/hf50662937/in/fdla/",
        "confidence": 0.94,
    },
    {
        "sheet": "ESTRUTURA",
        "name": "Staner",
        "subcategory": "Tripe De Caixa",
        "brand": "Staner",
        "value": 270.00,
        "method": "manual_web_curado",
        "source_title": "Suporte tripe para caixa de som Staner extensivel",
        "source_url": "https://www.krunner.com.br/suporte-tripe-para-caixa-de-som-staner-extensivel",
        "confidence": 0.92,
    },
    {
        "sheet": "ESTRUTURA",
        "name": "RMV",
        "subcategory": "Tripe De Microfone",
        "brand": "RMV",
        "value": 137.90,
        "method": "manual_web_curado",
        "source_title": "Suporte pedestal PSSU00142 tripe para microfone girafa cachimbo RMV",
        "source_url": "https://www.casadamusica.com.br/suporte-pedestal-pssu00142-trip%C3%A9-para-microfone-girafa-cachimbo-rmv",
        "confidence": 0.92,
    },
    {
        "sheet": "ESTRUTURA",
        "name": "SUP TV",
        "subcategory": "SUP TV",
        "brand": "",
        "value": 346.29,
        "method": "manual_web_curado",
        "source_title": "Suporte movel pedestal para TV de 32 a 65 com 2 bandejas",
        "source_url": "https://www.mercadolivre.com.br/suporte-movel-pedestal-para-tv-de-32-a-65-com-2-bandejas-cor-preto/up/MLBU3787511778",
        "confidence": 0.86,
    },
    {
        "sheet": "AUDIO",
        "name": "Pioneer FLX4",
        "subcategory": "Cdj/Xdj/Ddj",
        "brand": "Pioneer",
        "value": 3258.00,
        "method": "manual_web_curado",
        "source_title": "Controlador Pioneer DJ DDJ FLX4 preto",
        "source_url": "https://lista.mercadolivre.com.br/ddj-flx4",
        "confidence": 0.93,
    },
    {
        "sheet": "AUDIO",
        "name": "RCF Art 712-A",
        "subcategory": "Caixa De Som",
        "brand": "RCF",
        "value": 6349.00,
        "method": "manual_web_curado",
        "source_title": "Alto falante RCF ART 712-A MK4 black 127V",
        "source_url": "https://www.mercadolivre.com.br/alto-falante-rcf-art-712-a-mk4-black-127v/p/MLB15541412",
        "confidence": 0.94,
    },
    {
        "sheet": "FORA DE OPERACAO",
        "name": "Mark Audio CA-1000",
        "subcategory": "Caixa De Som",
        "brand": "Mark Audio",
        "value": 850.00,
        "method": "manual_web_curado",
        "source_title": "Caixa Ativa Mark Audio CA 1000",
        "source_url": "https://www.chinasom.com.br/caixa-ativa-mark-audio-ca-1000-ps-12815-430-p17975",
        "confidence": 0.95,
    },
    {
        "sheet": "AUDIO",
        "name": "Mackie PROFX6V3",
        "subcategory": "Effects",
        "brand": "Mackie",
        "value": 1790.00,
        "method": "manual_web_curado",
        "source_title": "Mesa de som analogica Mackie ProFX6v3",
        "source_url": "https://www.audiovideoecia.com.br/MLB-3984812867-mesa-de-som-analogica-mackie-profx6v3-com-efeitos-e-usb-_JM",
        "confidence": 0.91,
    },
    {
        "sheet": "AUDIO",
        "name": "Yamaha YAMAHA MG10XU",
        "subcategory": "Mesa De Som",
        "brand": "Yamaha",
        "value": 1849.00,
        "method": "manual_web_curado",
        "source_title": "Mesa Yamaha Mixer MG10XU 10 canais com efeitos",
        "source_url": "https://lista.mercadolivre.com.br/yamaha-mg10xu",
        "confidence": 0.90,
    },
    {
        "sheet": "AUDIO",
        "name": "RCF Evox J8",
        "subcategory": "Line Vertical",
        "brand": "RCF",
        "value": 10990.00,
        "method": "manual_web_curado",
        "source_title": "Caixa de som ativa RCF Evox J8 1400W com bag",
        "source_url": "https://www.mossmusic.com.br/produtos/caixa-de-som-ativa-rcf-evox-j8-1400w-1x12-1x8-bag",
        "confidence": 0.90,
    },
    {
        "sheet": "AUDIO",
        "name": "Shure SM58",
        "subcategory": "Microfone Com Fio",
        "brand": "Shure",
        "value": 1119.00,
        "method": "manual_web_curado",
        "source_title": "Microfone Shure SM58 LC preto",
        "source_url": "https://lista.mercadolivre.com.br/shure-sm58",
        "confidence": 0.92,
    },
    {
        "sheet": "AUDIO",
        "name": "Shure GLXD4R",
        "subcategory": "Base Wireless",
        "brand": "Shure",
        "value": 4199.00,
        "method": "manual_web_curado",
        "source_title": "Receptor sem fio GLXD4R Plus dual band rackmount Shure",
        "source_url": "https://www.ciadosom.com.br/produto/microfone-shure-glxd4r-wireless-receiver-banda-z2",
        "confidence": 0.90,
    },
    {
        "sheet": "AUDIO",
        "name": "Phenyx Pro PHENYX PRO PTU-7000",
        "subcategory": "Microfone Sem Fio",
        "brand": "Phenyx Pro",
        "value": 2233.00,
        "method": "manual_web_curado",
        "source_title": "PTU-7000-4H | Sistema de microfone sem fio UHF quadruplo com varredura automatica",
        "source_url": "https://phenyxpro.com/pt-br/products/ptu-7000",
        "confidence": 0.82,
    },
    {
        "sheet": "ACESSORIO",
        "name": "8 Port Nway Switch modelo ENH908-NWY",
        "subcategory": "Roteador",
        "brand": "Encore Eletronics",
        "value": 70.20,
        "method": "manual_web_curado",
        "source_title": "Encore 8P 10/100Mbps Switch ENH908-NWY",
        "source_url": "https://www.hardstore.com.br/shop/products/encore-8p-10-100mbps-switch-enh908-nwy",
        "confidence": 0.94,
    },
    {
        "sheet": "ACESSORIO",
        "name": "Dlink DES-1008C",
        "subcategory": "Roteador",
        "brand": "Dlink",
        "value": 79.33,
        "method": "manual_web_curado",
        "source_title": "Switch D-Link 8 Portas 10/100Mbps Fast Ethernet DES-1008C",
        "source_url": "https://www.premium.com.br/produto/switch-8-portas-10100mbps-fast-ethernet-des-1008a-dlink/",
        "confidence": 0.92,
    },
    {
        "sheet": "ACESSORIO",
        "name": "Apple APPLE ROUTER BRANCO",
        "subcategory": "Roteador",
        "brand": "Apple",
        "value": 249.00,
        "method": "manual_web_curado",
        "source_title": "Roteador Apple Airport Express (2nd geracao) em anuncios OLX",
        "source_url": "https://www.olx.com.br/anuncios/roteador-apple-airport-express",
        "confidence": 0.78,
    },
    {
        "sheet": "ACESSORIO",
        "name": "Poloclima PLI-45M",
        "subcategory": "Climatizador",
        "brand": "Poloclima",
        "value": 1499.00,
        "method": "manual_web_curado",
        "source_title": "Climatizador 045 Lt 240w 220v PLI-45M Poloclima",
        "source_url": "https://www.jlmeurer.com.br/climatizador-045-lt-240w-220v-pli-45m-poloclima",
        "confidence": 0.94,
    },
    {
        "sheet": "VIDEO",
        "name": "Book 550XDA",
        "subcategory": "Notebook",
        "brand": "Samsung",
        "value": 3719.35,
        "method": "manual_web_curado",
        "source_title": "Samsung Book NP550XDA com menor preco no Zoom",
        "source_url": "https://www.zoom.com.br/busca/samsung%2Bbook%2Bnp%2B550xda",
        "confidence": 0.88,
    },
    {
        "sheet": "VIDEO",
        "name": "MACBOOK PRO A1398",
        "subcategory": "Notebook",
        "brand": "Apple",
        "value": 1250.00,
        "method": "manual_web_curado",
        "source_title": "MacBook Pro A1398 em anuncios OLX Brasil",
        "source_url": "https://www.olx.com.br/anuncios/macbook-pro-a1398",
        "confidence": 0.78,
    },
    {
        "sheet": "VIDEO",
        "name": "MP515 DIGITAL PROJECTOR",
        "subcategory": "Projetor",
        "brand": "Benq",
        "value": 370.00,
        "method": "manual_web_curado",
        "source_title": "Projetores BenQ MP515 em anuncios OLX Brasil",
        "source_url": "https://www.olx.com.br/anuncios/projetores-benq-mp515",
        "confidence": 0.82,
    },
    {
        "sheet": "VIDEO",
        "name": "Sendcard",
        "subcategory": "Sendbox",
        "brand": "Linsn",
        "value": 943.40,
        "method": "manual_web_curado",
        "source_title": "Send Card Linsn para painel de LED",
        "source_url": "https://lista.mercadolivre.com.br/send-card-linsn",
        "confidence": 0.86,
    },
    {
        "sheet": "VIDEO",
        "name": "Suporte Preto fixo para notebook",
        "subcategory": "Suporte",
        "brand": "",
        "value": 62.99,
        "method": "manual_web_curado",
        "source_title": "Base para notebook Multi com Cooler AC166",
        "source_url": "https://www.buscape.com.br/suporte-para-notebook/laptop-stand",
        "confidence": 0.78,
    },
    {
        "sheet": "VIDEO",
        "name": "A16 - A3354",
        "subcategory": "Tablet",
        "brand": "Ipad",
        "value": 2849.00,
        "method": "manual_web_curado",
        "source_title": "iPad A3354 com menor preco no Buscape",
        "source_url": "https://www.buscape.com.br/busca/ipad%2Ba3354",
        "confidence": 0.90,
    },
    {
        "sheet": "ILUMINACAO",
        "name": "Showtech ST-OP1806CS",
        "subcategory": "Outdoor",
        "brand": "Showtech",
        "value": 640.00,
        "method": "manual_web_curado",
        "source_title": "Par Led 18 RGBWA+UV IP65 ST-OP1806",
        "source_url": "https://skylinesomluzled.com.br/produto/par-led-18-rgbwauv-ip-65/",
        "confidence": 0.91,
    },
    {
        "sheet": "ACESSORIO",
        "name": "Showtech ST-OP1806CS",
        "subcategory": "Par Leds",
        "brand": "Showtech",
        "value": 640.00,
        "method": "manual_web_curado",
        "source_title": "Par Led 18 RGBWA+UV IP65 ST-OP1806",
        "source_url": "https://skylinesomluzled.com.br/produto/par-led-18-rgbwauv-ip-65/",
        "confidence": 0.91,
    },
    {
        "sheet": "AUDIO",
        "name": "PARA MICROFONE SM58",
        "subcategory": "Capsula Mic",
        "brand": "",
        "value": 1119.00,
        "method": "manual_web_curado",
        "source_title": "Microfone Shure SM58 LC preto",
        "source_url": "https://lista.mercadolivre.com.br/shure-sm58",
        "confidence": 0.92,
    },
    {
        "sheet": "AUDIO",
        "name": "Shure Beta 58A + GLXD2 Z2",
        "subcategory": "Microfone Wireless",
        "brand": "Shure",
        "value": 3290.00,
        "method": "manual_web_curado",
        "source_title": "Microfone transmissor Shure GLXD2 Beta58A",
        "source_url": "https://www.cariocainstrumentos.com.br/microfone-transmissor-shure-glxd2-beta58a-z2",
        "confidence": 0.88,
    },
    {
        "sheet": "FORA DE OPERACAO",
        "name": "Globo espelhado com motor",
        "subcategory": "Globo",
        "brand": "",
        "value": 239.90,
        "method": "manual_web_curado",
        "source_title": "Globo Espelhado 30cm + Motor Bivolt + Correntinha",
        "source_url": "https://lista.mercadolivre.com.br/globo-espelhado-30-cm",
        "confidence": 0.78,
    },
    {
        "sheet": "ILUMINACAO",
        "name": "SOG-512 DMX512",
        "subcategory": "Mesa De Luz",
        "brand": "SOG",
        "value": 247.99,
        "method": "manual_web_curado",
        "source_title": "Mesa controladora DMX 512 192 canais no Mercado Livre",
        "source_url": "https://lista.mercadolivre.com.br/mesa-dmx-512",
        "confidence": 0.78,
    },
    {
        "sheet": "AUDIO",
        "name": "ANTK-CBL-BNC",
        "subcategory": "Acessorio",
        "brand": "Phenyx Pro",
        "value": 131.61,
        "method": "manual_web_curado",
        "source_title": "Phenyx Pro ANTK-CBL cabos BNC/TNC 50cm, US$25.00",
        "source_url": "https://phenyxpro.com/pt-br/products/antk-cbl",
        "confidence": 0.88,
    },
    {
        "sheet": "AUDIO",
        "name": "PHENIX PRO DIRECTIONAL ANTENNA PAS-82",
        "subcategory": "Amp Rf",
        "brand": "Phenyx Pro",
        "value": 710.53,
        "method": "manual_web_curado",
        "source_title": "Phenyx Pro PAS-82 directional antenna paddles, US$134.99",
        "source_url": "https://phenyxpro.com/es/products/pas-82",
        "confidence": 0.86,
    },
    {
        "sheet": "AUDIO",
        "name": "PAS-225 ANTENNA DISTRIBUTION SYSTEM",
        "subcategory": "Amp Rf",
        "brand": "Phenyx Pro",
        "value": 1848.27,
        "method": "manual_web_curado",
        "source_title": "Phenyx Pro PAS-225X antenna distribution bundle, US$350.99",
        "source_url": "https://phenyxpro.com/es/products/pas-225x",
        "confidence": 0.84,
    },
    {
        "sheet": "AUDIO",
        "name": "POWERPLAY P16-I na caixa",
        "subcategory": "Amp Fones",
        "brand": "Behringer",
        "value": 550.00,
        "method": "manual_web_curado",
        "source_title": "Powerplay Behringer P16-i usado no Mercado Livre",
        "source_url": "https://lista.mercadolivre.com.br/powerplay-usado",
        "confidence": 0.84,
    },
    {
        "sheet": "AUDIO",
        "name": "UHF WIRELESS MONO IN-EAR MONITOR SYSTEM PTM-33 SERIES",
        "subcategory": "In Ear S / Fio",
        "brand": "Phenyx Pro",
        "value": 2632.15,
        "method": "manual_web_curado",
        "source_title": "Phenyx Pro PTM-33-4B mono wireless in-ear monitor system, US$499.99",
        "source_url": "https://phenyxpro.com/en-br/products/2hm-blk",
        "confidence": 0.84,
    },
    {
        "sheet": "AUDIO",
        "name": "EURORACK UB802",
        "subcategory": "Mesa De Som",
        "brand": "Behringer",
        "value": 300.00,
        "method": "manual_web_curado",
        "source_title": "Mesa Behringer Eurorack UB802 usada em anuncios OLX",
        "source_url": "https://www.olx.com.br/anuncios/mesa-som-behringer-ub802",
        "confidence": 0.82,
    },
    {
        "sheet": "AUDIO",
        "name": "X32 RACK",
        "subcategory": "Mesa De Som",
        "brand": "Behringer",
        "value": 12900.00,
        "method": "manual_web_curado",
        "source_title": "Mesa de som digital Behringer X32 Rack 40 canais",
        "source_url": "https://produto.mercadolivre.com.br/MLB-5206525400-mesa-de-som-digital-behringer-x32-rack-40-canais-bivolt-_JM",
        "confidence": 0.92,
    },
    {
        "sheet": "AUDIO",
        "name": "58A",
        "subcategory": "Mic C/ Fio",
        "brand": "Beta",
        "value": 1228.00,
        "method": "manual_web_curado",
        "source_title": "Microfone bastao Beta 58A Shure no Mercado Livre",
        "source_url": "https://lista.mercadolivre.com.br/shure-beta-58a-sem-fio",
        "confidence": 0.84,
    },
    {
        "sheet": "AUDIO",
        "name": "SM58",
        "subcategory": "Mic C/ Fio",
        "brand": "Shure",
        "value": 1119.00,
        "method": "manual_web_curado",
        "source_title": "Microfone Shure SM58 LC preto",
        "source_url": "https://www.mercadolivre.com.br/microfone-shure-sm58-lc-dinamico-sm-58-series-cardioide/up/MLBU1273211836",
        "confidence": 0.92,
    },
    {
        "sheet": "AUDIO",
        "name": "Super 55",
        "subcategory": "Mic C/ Fio",
        "brand": "Shure",
        "value": 2712.00,
        "method": "manual_web_curado",
        "source_title": "Microfone Vocal Shure Super55 no Mercado Livre",
        "source_url": "https://lista.mercadolivre.com.br/shure-super-55-deluxe",
        "confidence": 0.88,
    },
    {
        "sheet": "AUDIO",
        "name": "PTU-7000",
        "subcategory": "Mic S/ Fio",
        "brand": "Phenyx Pro",
        "value": 2233.00,
        "method": "manual_web_curado",
        "source_title": "Phenyx Pro PTU-7000-4H quad UHF wireless microphone system",
        "source_url": "https://phenyxpro.com/en-br/products/ptu-7000",
        "confidence": 0.90,
    },
    {
        "sheet": "AUDIO",
        "name": "PTU-7000B MADONNA",
        "subcategory": "Mic S/ Fio",
        "brand": "Phenyx Pro",
        "value": 263.17,
        "method": "manual_web_curado",
        "source_title": "Phenyx Pro PWB-7 wireless bodypack transmitter, US$49.99",
        "source_url": "https://phenyxpro.com/pt/products/pwb-7-bodypack",
        "confidence": 0.82,
    },
    {
        "sheet": "AUDIO",
        "name": "MICROFONE SEM FIO PHENYX PRO",
        "subcategory": "Mic S/ Fio",
        "brand": "Phenyx Pro",
        "value": 357.00,
        "method": "manual_web_curado",
        "source_title": "Phenyx Pro PWH-7 handheld transmitter",
        "source_url": "https://phenyxpro.com/es-br/collections/handheld",
        "confidence": 0.80,
    },
    {
        "sheet": "AUDIO",
        "name": "Microfone bastao Freq range 502-603Mhz",
        "subcategory": "Mic S/ Fio",
        "brand": "Phenyx Pro",
        "value": 357.00,
        "method": "manual_web_curado",
        "source_title": "Phenyx Pro PWH-7 handheld transmitter",
        "source_url": "https://phenyxpro.com/es-br/collections/handheld",
        "confidence": 0.80,
    },
    {
        "sheet": "AUDIO",
        "name": "SKM100 G3 - Freq E",
        "subcategory": "Mic S/ Fio",
        "brand": "Sennheiser",
        "value": 2750.00,
        "method": "manual_web_curado",
        "source_title": "Sennheiser SKM100 G3 bastao em anuncios OLX Fortaleza",
        "source_url": "https://www.olx.com.br/audio/microfones-e-gravadores/microfone-sennheiser/estado-ce/fortaleza-e-regiao/fortaleza",
        "confidence": 0.78,
    },
    {
        "sheet": "AUDIO",
        "name": "SKM100 G3 - Freq G",
        "subcategory": "Mic S/ Fio",
        "brand": "Sennheiser",
        "value": 2750.00,
        "method": "manual_web_curado",
        "source_title": "Sennheiser SKM100 G3 bastao em anuncios OLX Fortaleza",
        "source_url": "https://www.olx.com.br/audio/microfones-e-gravadores/microfone-sennheiser/estado-ce/fortaleza-e-regiao/fortaleza",
        "confidence": 0.78,
    },
    {
        "sheet": "AUDIO",
        "name": "SKM100 G4 - Freq G",
        "subcategory": "Mic S/ Fio",
        "brand": "Sennheiser",
        "value": 2250.00,
        "method": "manual_web_curado",
        "source_title": "Sennheiser SKM100 G4 bastao nova na caixa em anuncios OLX Fortaleza",
        "source_url": "https://www.olx.com.br/audio/microfones-e-gravadores/microfone-sennheiser/estado-ce/fortaleza-e-regiao/fortaleza",
        "confidence": 0.80,
    },
    {
        "sheet": "AUDIO",
        "name": "Receiver - EM100 G3 - Freq A",
        "subcategory": "Rx Mic",
        "brand": "Sennheiser",
        "value": 790.00,
        "method": "manual_web_curado",
        "source_title": "Sennheiser EM100 G3 receiver usado no eBay, US$150",
        "source_url": "https://www.ebay.com/itm/406084396086",
        "confidence": 0.80,
    },
    {
        "sheet": "AUDIO",
        "name": "Receiver - EM100 G3 - Freq E",
        "subcategory": "Rx Mic",
        "brand": "Sennheiser",
        "value": 790.00,
        "method": "manual_web_curado",
        "source_title": "Sennheiser EM100 G3 receiver usado no eBay, US$150",
        "source_url": "https://www.ebay.com/itm/406084396086",
        "confidence": 0.80,
    },
    {
        "sheet": "AUDIO",
        "name": "Receiver - EM100 G3 - Freq G",
        "subcategory": "Rx Mic",
        "brand": "Sennheiser",
        "value": 790.00,
        "method": "manual_web_curado",
        "source_title": "Sennheiser EM100 G3 receiver usado no eBay, US$150",
        "source_url": "https://www.ebay.com/itm/406084396086",
        "confidence": 0.80,
    },
    {
        "sheet": "AUDIO",
        "name": "Receiver Wireless EK100 G3",
        "subcategory": "Rx Mic",
        "brand": "Sennheiser",
        "value": 1185.00,
        "method": "manual_web_curado",
        "source_title": "Sennheiser EK100 G3 bodypack receiver usado no eBay, US$224.97",
        "source_url": "https://www.ebay.com/itm/205490867327",
        "confidence": 0.80,
    },
    {
        "sheet": "AUDIO",
        "name": "T4V Diversity Receiver",
        "subcategory": "Rx Mic",
        "brand": "Shure",
        "value": 237.00,
        "method": "manual_web_curado",
        "source_title": "Shure T4V diversity receiver usado no eBay, US$44.99",
        "source_url": "https://www.ebay.com/itm/276706705604",
        "confidence": 0.78,
    },
    {
        "sheet": "AUDIO",
        "name": "XS Wireless EM 10 com duas antenas",
        "subcategory": "Rx Mic",
        "brand": "Sennheiser",
        "value": 252.65,
        "method": "manual_web_curado",
        "source_title": "Sennheiser XS Wireless EM-10 true diversity receiver usado no eBay, US$47.99",
        "source_url": "https://www.ebay.com/p/16011632112",
        "confidence": 0.78,
    },
    {
        "sheet": "ACESSORIO",
        "name": "ewd skm-s freq.q1-6",
        "subcategory": "Bodypack",
        "brand": "Sennheiser",
        "value": 2205.00,
        "method": "manual_web_curado",
        "source_title": "Sennheiser EW-D SKM-S (R4-9) no Mercado Livre",
        "source_url": "https://lista.mercadolivre.com.br/microfone-sennheiser-skm-9100",
        "confidence": 0.80,
    },
    {
        "sheet": "ACESSORIO",
        "name": "SKM100s G4 - Freq g",
        "subcategory": "Mic",
        "brand": "Sennheiser",
        "value": 2250.00,
        "method": "manual_web_curado",
        "source_title": "Sennheiser SKM100 G4 bastao nova na caixa em anuncios OLX Fortaleza",
        "source_url": "https://www.olx.com.br/audio/microfones-e-gravadores/microfone-sennheiser/estado-ce/fortaleza-e-regiao/fortaleza",
        "confidence": 0.80,
    },
    {
        "sheet": "FORA DE OPERACAO",
        "name": "PHENIX PRO PTU-7000 RECEIVER",
        "subcategory": "Mic S/ Fio",
        "brand": "Phenyx Pro",
        "value": 1117.00,
        "method": "manual_web_curado",
        "source_title": "Phenyx Pro PWR-7000 wireless receiver",
        "source_url": "https://phenyxpro.com/en-br/products/ptu-7000",
        "confidence": 0.90,
    },
    {
        "sheet": "AUDIO",
        "name": "Transmitter SK100 G3",
        "subcategory": "Mic S/ Fio",
        "brand": "Sennheiser",
        "value": 995.00,
        "method": "manual_web_curado",
        "source_title": "Sennheiser SK100 G3 bodypack transmitter usado no eBay, US$189",
        "source_url": "https://www.ebay.com/p/2255394973",
        "confidence": 0.80,
    },
    {
        "sheet": "AUDIO",
        "name": "Sennheiser",
        "subcategory": "Mic S/ Fio",
        "brand": "Sennheiser",
        "value": 1500.00,
        "method": "manual_web_curado",
        "source_title": "Microfones sem fio Sennheiser em anuncios OLX Fortaleza",
        "source_url": "https://www.olx.com.br/audio/microfones-e-gravadores/microfone-sennheiser/estado-ce/fortaleza-e-regiao/fortaleza",
        "confidence": 0.74,
    },
    {
        "sheet": "AUDIO",
        "name": "Cinza",
        "subcategory": "Caixa Som",
        "brand": "CSR",
        "value": 2470.00,
        "method": "manual_web_curado",
        "source_title": "Caixa ativa CSR 12 polegadas 200W 5512A",
        "source_url": "https://www.mercadolivre.com.br/caixa-ativa-12-polegadas-200w-2-vias-xlr-rca-p10-csr-5512a/p/MLB41788658",
        "confidence": 0.88,
    },
    {
        "sheet": "AUDIO",
        "name": "12'' PASSIVO",
        "subcategory": "Caixa Som",
        "brand": "NHL",
        "value": 1699.00,
        "method": "manual_web_curado",
        "source_title": "Caixa acustica retorno 12 polegadas passiva no Mercado Livre",
        "source_url": "https://lista.mercadolivre.com.br/kit-som-ativo-profissional-nhl",
        "confidence": 0.82,
    },
    {
        "sheet": "AUDIO",
        "name": "12''ATIVO",
        "subcategory": "Caixa Som",
        "brand": "NHL",
        "value": 2499.00,
        "method": "manual_web_curado",
        "source_title": "Monitor de palco PA ativo RP12.400A 12 polegadas NHL",
        "source_url": "https://lista.mercadolivre.com.br/caixa-ativa-12-nhl",
        "confidence": 0.84,
    },
    {
        "sheet": "AUDIO",
        "name": "Leacs",
        "subcategory": "Caixa Som",
        "brand": "Leacs",
        "value": 2299.00,
        "method": "manual_web_curado",
        "source_title": "Caixa ativa Leacs Fit320A Bluetooth 250W 12 polegadas",
        "source_url": "https://www.mercadolivre.com.br/caixa-ativa-leacs-fit320a-bluetooth-250w-12-monitor-de-palco/p/MLB36854322",
        "confidence": 0.86,
    },
    {
        "sheet": "AUDIO",
        "name": "Microfone",
        "subcategory": "Mic",
        "brand": "Skp-Pro Audio",
        "value": 301.67,
        "method": "manual_web_curado",
        "source_title": "Microfone SKP Pro Audio com fio em resultados Mercado Livre",
        "source_url": "https://lista.mercadolivre.com.br/microfone-skp-pro-audio",
        "confidence": 0.78,
    },
    {
        "sheet": "ACESSORIO",
        "name": "ewd skm-s freq.q1-6",
        "subcategory": "Mic",
        "brand": "Sennheiser",
        "value": 2205.00,
        "method": "manual_web_curado",
        "source_title": "Sennheiser EW-D SKM-S (R4-9) no Mercado Livre",
        "source_url": "https://lista.mercadolivre.com.br/microfone-sennheiser-skm-9100",
        "confidence": 0.80,
    },
    {
        "sheet": "FORA DE OPERACAO",
        "name": "SWA1501",
        "subcategory": "Sub",
        "brand": "Mackie",
        "value": 3000.00,
        "method": "manual_web_curado",
        "source_title": "Sub Woofer Grave Mackie SWA1501 usado em anuncio OLX",
        "source_url": "https://www.olx.com.br/anuncios/sub-antera-15",
        "confidence": 0.82,
    },
]
CURRENT_MARKET_OVERRIDES = [
    {
        "sheet": "ILUMINACAO",
        "name": "ST-X251LAY",
        "subcategory": "9R",
        "brand": "Showtech",
        "value": 2500.00,
        "method": "web_atual_curado",
        "source_title": "2 Moving Head Beam 9R 250W Borda Led ST-X251LAY com case",
        "source_url": "https://lista.mercadolivre.com.br/moving-head-9r",
        "confidence": 0.90,
    },
    {
        "sheet": "ACESSORIO",
        "name": "ST-X251LAY",
        "subcategory": "Moving Beam",
        "brand": "Showtech",
        "value": 2500.00,
        "method": "web_atual_curado",
        "source_title": "2 Moving Head Beam 9R 250W Borda Led ST-X251LAY com case",
        "source_url": "https://lista.mercadolivre.com.br/moving-head-9r",
        "confidence": 0.90,
    },
    {
        "sheet": "ILUMINACAO",
        "name": "ST-XQDFS24 - BLINDADA",
        "subcategory": "Fixa",
        "brand": "Showtech",
        "value": 819.99,
        "method": "web_atual_curado",
        "source_title": "Ribalta Led RGBW 24x12W Pixel Quadriled 4n1 Outdoor IP67",
        "source_url": "https://www.mercadolivre.com.br/ribalta-led-rgbw-24x12w-pixel-quadriled-4n1-outdoor-ip67/p/MLB50404325",
        "confidence": 0.92,
    },
    {
        "sheet": "AUDIO",
        "name": "K12.2",
        "subcategory": "Caixa Som",
        "brand": "QSC",
        "value": 8990.00,
        "method": "web_atual_curado",
        "source_title": "QSC K12.2 no Mercado Livre",
        "source_url": "https://lista.mercadolivre.com.br/qsc-k12.2",
        "confidence": 0.90,
    },
    {
        "sheet": "AUDIO",
        "name": "Evox J8",
        "subcategory": "Line Vertical",
        "brand": "RCF",
        "value": 9998.00,
        "method": "web_atual_curado",
        "source_title": "RCF Evox J8 no Mercado Livre",
        "source_url": "https://lista.mercadolivre.com.br/rcf-evox-j8",
        "confidence": 0.90,
    },
    {
        "sheet": "AUDIO",
        "name": "DXS12",
        "subcategory": "Sub",
        "brand": "Yamaha",
        "value": 6373.43,
        "method": "web_atual_curado",
        "source_title": "Subwoofer Yamaha DXS12 Dxs Series Un",
        "source_url": "https://www.lanworks.com.br/subwoofer-yamaha-dxs12-dxs-series",
        "confidence": 0.73,
    },
    {
        "sheet": "AUDIO",
        "name": "K10.2",
        "subcategory": "Caixa Som",
        "brand": "QSC",
        "value": 8354.00,
        "method": "web_atual_curado",
        "source_title": "QSC K10.2 no Mercado Livre",
        "source_url": "https://lista.mercadolivre.com.br/qsc-k10.2",
        "confidence": 0.90,
    },
    {
        "sheet": "AUDIO",
        "name": "X32 RACK",
        "subcategory": "Mesa Som",
        "brand": "Behringer",
        "value": 10490.00,
        "method": "web_atual_curado",
        "source_title": "Behringer X32 Rack no Mercado Livre",
        "source_url": "https://lista.mercadolivre.com.br/behringer-x32-rack",
        "confidence": 0.90,
    },
    {
        "sheet": "AUDIO",
        "name": "Art 712-A",
        "subcategory": "Caixa De Som",
        "brand": "RCF",
        "value": 6149.00,
        "method": "web_atual_curado",
        "source_title": "Caixa De Som RCF ART 712-A MK5 129 dB",
        "source_url": "https://lista.mercadolivre.com.br/rcf-art-712",
        "confidence": 0.76,
    },
    {
        "sheet": "AUDIO",
        "name": "DXR15",
        "subcategory": "Caixa Som",
        "brand": "Yamaha",
        "value": 8499.00,
        "method": "web_atual_curado",
        "source_title": "Yamaha DXR15 no Mercado Livre",
        "source_url": "https://lista.mercadolivre.com.br/yamaha-dxr15",
        "confidence": 0.88,
    },
    {
        "sheet": "AUDIO",
        "name": "HANDSONIC HPD-20",
        "subcategory": "Bat Mao",
        "brand": "Roland",
        "value": 7839.00,
        "method": "web_atual_curado",
        "source_title": "Roland HPD-20 HandSonic",
        "source_url": "https://store.roland.com.br/products/pad-de-percussao-hpd-20",
        "confidence": 0.98,
    },
    {
        "sheet": "AUDIO",
        "name": "SPD-SX SAMPLING PAD",
        "subcategory": "Bat Mao",
        "brand": "Roland",
        "value": 6619.00,
        "method": "web_atual_curado",
        "source_title": "Roland SPD-SX",
        "source_url": "https://store.roland.com.br/sample-pad-spd-sx/p",
        "confidence": 0.98,
    },
    {
        "sheet": "AUDIO",
        "name": "ew digital freq.q1-6",
        "subcategory": "Base Mic",
        "brand": "Sennheiser",
        "value": 4699.90,
        "method": "web_atual_curado",
        "source_title": "EW-D EM",
        "source_url": "https://pt-br.shop.sennheiser.com/products/ew-d-em",
        "confidence": 0.90,
    },
    {
        "sheet": "AUDIO",
        "name": "ew digital freq.r1-6",
        "subcategory": "Base Mic",
        "brand": "Sennheiser",
        "value": 4699.90,
        "method": "web_atual_curado",
        "source_title": "EW-D EM",
        "source_url": "https://pt-br.shop.sennheiser.com/products/ew-d-em",
        "confidence": 0.84,
    },
    {
        "sheet": "AUDIO",
        "name": "EWD1",
        "subcategory": "Mic S/ Fio",
        "brand": "Microfone SEM FIO",
        "value": 3581.65,
        "method": "web_atual_curado",
        "source_title": "EW-D SKM-S",
        "source_url": "https://pt-br.shop.sennheiser.com/products/ew-d-skm-s",
        "confidence": 0.82,
    },
    {
        "sheet": "AUDIO",
        "name": "T2 Wireless Transmitter - Freq. CF",
        "subcategory": "Mic S/ Fio",
        "brand": "Shure",
        "value": 334.01,
        "method": "web_atual_curado",
        "source_title": "Shure 14A T2 Vocal Artist Transmitter, convertido de USD em 03/04/2026",
        "source_url": "https://reverb.com/item/75937694-shure-14a-t2-vocal-artist-transmitter",
        "confidence": 0.74,
    },
    {
        "sheet": "AUDIO",
        "name": "T2 Wireless Transmitter - Freq. CV",
        "subcategory": "Mic S/ Fio",
        "brand": "Shure",
        "value": 334.01,
        "method": "web_atual_curado",
        "source_title": "Shure 14A T2 Vocal Artist Transmitter, convertido de USD em 03/04/2026",
        "source_url": "https://reverb.com/item/75937694-shure-14a-t2-vocal-artist-transmitter",
        "confidence": 0.74,
    },
    {
        "sheet": "AUDIO",
        "name": "Sennheiser",
        "subcategory": "Mic S/ Fio",
        "brand": "Sennheiser",
        "value": 1500.00,
        "method": "web_atual_curado",
        "source_title": "Microfones sem fio Sennheiser em anuncios OLX Fortaleza",
        "source_url": "https://www.olx.com.br/audio/microfones-e-gravadores/microfone-sennheiser/estado-ce/fortaleza-e-regiao/fortaleza",
        "confidence": 0.74,
    },
    {
        "sheet": "ACESSORIO",
        "name": "ewd skm-s freq.q1-6",
        "subcategory": "Bodypack",
        "brand": "Sennheiser",
        "value": 3581.65,
        "method": "web_atual_curado",
        "source_title": "EW-D SKM-S",
        "source_url": "https://pt-br.shop.sennheiser.com/products/ew-d-skm-s",
        "confidence": 0.95,
    },
    {
        "sheet": "ACESSORIO",
        "name": "ewd skm-s freq.q1-6",
        "subcategory": "Mic",
        "brand": "Sennheiser",
        "value": 3581.65,
        "method": "web_atual_curado",
        "source_title": "EW-D SKM-S",
        "source_url": "https://pt-br.shop.sennheiser.com/products/ew-d-skm-s",
        "confidence": 0.95,
    },
    {
        "sheet": "ACESSORIO",
        "name": "SKM100s G4 - Freq g",
        "subcategory": "Mic",
        "brand": "Sennheiser",
        "value": 2250.00,
        "method": "web_atual_curado",
        "source_title": "Sennheiser SKM100 G4 bastao nova na caixa em anuncios OLX Fortaleza",
        "source_url": "https://www.olx.com.br/audio/microfones-e-gravadores/microfone-sennheiser/estado-ce/fortaleza-e-regiao/fortaleza",
        "confidence": 0.80,
    },
    {
        "sheet": "AUDIO",
        "name": "XDJ-XZ PIONEER",
        "subcategory": "Cdj / Xdj / Ddj",
        "brand": "Pioneer",
        "value": 15900.00,
        "method": "web_atual_curado",
        "source_title": "Controlador DJ Pioneer XDJ-XZ preto 4 canais",
        "source_url": "https://lista.mercadolivre.com.br/pioneer-xdj-xz",
        "confidence": 0.92,
    },
    {
        "sheet": "AUDIO",
        "name": "XDJ-RR PIONEER",
        "subcategory": "Cdj / Xdj / Ddj",
        "brand": "Pioneer",
        "value": 11319.00,
        "method": "web_atual_curado",
        "source_title": "Controlador DJ Pioneer XDJ-RR preto 2 canais",
        "source_url": "https://lista.mercadolivre.com.br/controlador-dj-pioneer-xdj-rr",
        "confidence": 0.93,
    },
    {
        "sheet": "AUDIO",
        "name": "DDJ-400 PIONEER",
        "subcategory": "Cdj / Xdj / Ddj",
        "brand": "Pioneer",
        "value": 2500.00,
        "method": "web_atual_curado",
        "source_title": "Controlador Pioneer DJ DDJ-400",
        "source_url": "https://lista.mercadolivre.com.br/controladora-ddj-400-pioneer-usada",
        "confidence": 0.91,
    },
    {
        "sheet": "AUDIO",
        "name": "8003-AS II 2200W",
        "subcategory": "Sub",
        "brand": "RCF",
        "value": 26246.00,
        "method": "web_atual_curado",
        "source_title": "Subwoofer Ativo RCF 8003AS II 2200W Som Natural Preto",
        "source_url": "https://lista.mercadolivre.com.br/subwoofer-rcf",
        "confidence": 0.97,
    },
    {
        "sheet": "AUDIO",
        "name": "PARA MICROFONE SM58",
        "subcategory": "Capsula Mic",
        "brand": "",
        "value": 1119.00,
        "method": "web_atual_curado",
        "source_title": "Microfone Shure SM58 LC preto",
        "source_url": "https://lista.mercadolivre.com.br/shure-sm58",
        "confidence": 0.92,
    },
]
LIVE_WEB_METHODS = {"manual_web_curado", "mercadolivre_busca"}
SET_SOURCE_MARKERS = (" BASE SET", " SET ", " KIT ", " BUNDLE ", " SYSTEM ")
COMPONENT_NAME_MARKERS = (
    "TRANSMITTER",
    "RECEIVER",
    "BODYPACK",
    "GLXD2",
    "SKM",
    "EWD",
    "CAPSULA",
    "SAMPLING PAD",
    "HANDSONIC",
    "SPD SX",
)
RISKY_INFERENCE_SUBCATEGORIES = {"WIRELESS", "BAT MAO"}


@dataclass
class SerialRecord:
    sheet: str
    row_idx: int
    codigo: str
    nome: str
    subcategoria: str | None
    marca: str | None
    estado: str | None
    desgaste: float | int | None
    valor_original: float | None

    @property
    def key(self) -> tuple[str, str, str | None, str | None]:
        return (self.sheet, self.nome, self.subcategoria, self.marca)


@dataclass
class ResolvedPricing:
    original_value: float
    original_match: object
    current_value: float
    current_match: object


def load_helpers():
    spec = importlib.util.spec_from_file_location("fill_inventory_values", HELPER_PATH)
    module = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def find_header_row(ws) -> int | None:
    for row_idx in range(1, 8):
        values = [cell.value for cell in ws[row_idx]]
        if "Codigo" in values and "Nome" in values and "Valor Original" in values:
            return row_idx
    return None


def parse_serial_rows(path: Path, helpers) -> list[SerialRecord]:
    wb = load_workbook(path, data_only=True)
    serials: list[SerialRecord] = []
    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        header_row = find_header_row(ws)
        if header_row is None:
            continue
        headers = [cell.value for cell in ws[header_row]]
        if "Codigo" not in headers or "Nome" not in headers:
            continue

        idx = {name: headers.index(name) for name in headers if name}
        data_start = header_row + 1
        for row_idx, row in enumerate(ws.iter_rows(min_row=data_start, values_only=True), start=data_start):
            codigo = row[idx["Codigo"]] if idx.get("Codigo") is not None else None
            nome = row[idx["Nome"]] if idx.get("Nome") is not None else None
            if not codigo:
                continue
            codigo_text = str(codigo).strip()
            if not codigo_text.startswith("MMD-"):
                continue
            subcategoria = row[idx["Subcategoria"]] if idx.get("Subcategoria") is not None else None
            nome_text = str(nome).strip() if nome not in (None, "") else str(subcategoria).strip() if subcategoria not in (None, "") else "SEM NOME"
            serials.append(
                SerialRecord(
                    sheet=sheet_name,
                    row_idx=row_idx,
                    codigo=codigo_text,
                    nome=nome_text,
                    subcategoria=subcategoria,
                    marca=row[idx["Marca"]] if idx.get("Marca") is not None else None,
                    estado=row[idx["Estado"]] if idx.get("Estado") is not None else None,
                    desgaste=row[idx["Desgaste"]] if idx.get("Desgaste") is not None else None,
                    valor_original=helpers.safe_float(row[idx["Valor Original"]]) if idx.get("Valor Original") is not None else None,
                )
            )
    return serials


def build_unique_items(serials: list[SerialRecord], helpers):
    grouped: dict[tuple[str, str, str | None, str | None], list[SerialRecord]] = defaultdict(list)
    for record in serials:
        grouped[record.key].append(record)

    items = []
    for item_id, (key, rows) in enumerate(grouped.items(), start=1):
        sheet, nome, subcategoria, marca = key
        existing_values = [row.valor_original for row in rows if row.valor_original is not None]
        item = helpers.ItemRow(
            row_idx=rows[0].row_idx,
            item_id=item_id,
            nome=nome,
            categoria=sheet,
            subcategoria=subcategoria,
            marca=marca,
            modelo=nome,
            rastreamento="INDIVIDUAL",
            quantidade=len(rows),
            valor=statistics.median(existing_values) if existing_values else None,
        )
        items.append((key, item, rows))
    return items


def normalize_identity(sheet: str | None, name: str | None, subcategory: str | None, brand: str | None, helpers) -> tuple[str, str, str, str]:
    subcategory_key = (
        helpers.canonical_subcategory(subcategory)
        if hasattr(helpers, "canonical_subcategory")
        else helpers.fold_text(subcategory)
    )
    return (
        helpers.fold_text(sheet),
        helpers.fold_text(name),
        subcategory_key,
        helpers.fold_text(brand),
    )


def build_manual_override_map(helpers) -> dict[tuple[str, str, str, str], object]:
    overrides = {}
    for entry in MANUAL_SOURCE_OVERRIDES:
        match = helpers.SourceMatch(
            value=entry["value"],
            method=entry["method"],
            source_title=entry["source_title"],
            source_url=entry["source_url"],
            confidence=entry["confidence"],
            notes="override manual curado",
        )
        key_variants = {
            normalize_identity(entry["sheet"], entry["name"], entry["subcategory"], entry["brand"], helpers),
        }
        stripped_name = helpers.strip_brand_preserving_tokens(entry["name"], entry["brand"])
        if stripped_name:
            key_variants.add(normalize_identity(entry["sheet"], stripped_name, entry["subcategory"], entry["brand"], helpers))
        for candidate in helpers.extract_model_candidates(entry["name"]):
            key_variants.add(normalize_identity(entry["sheet"], candidate, entry["subcategory"], entry["brand"], helpers))

        for key in key_variants:
            overrides[key] = match
    return overrides


def build_current_market_override_map(helpers) -> dict[tuple[str, str, str, str], object]:
    overrides = {}
    entries = []
    if CURRENT_MARKET_OVERRIDES_PATH.exists():
        with CURRENT_MARKET_OVERRIDES_PATH.open(encoding="utf-8", newline="") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                if not row.get("sheet") or not row.get("name"):
                    continue
                try:
                    value = float(row["value"])
                    confidence = float(row["confidence"])
                except (KeyError, TypeError, ValueError):
                    continue
                entries.append(
                    {
                        "sheet": row["sheet"],
                        "name": row["name"],
                        "subcategory": row.get("subcategory", ""),
                        "brand": row.get("brand", ""),
                        "value": value,
                        "method": row.get("method", "web_atual_curado"),
                        "source_title": row.get("source_title", ""),
                        "source_url": row.get("source_url", ""),
                        "confidence": confidence,
                    }
                )
    entries.extend(CURRENT_MARKET_OVERRIDES)

    for entry in entries:
        match = helpers.SourceMatch(
            value=entry["value"],
            method=entry["method"],
            source_title=entry["source_title"],
            source_url=entry["source_url"],
            confidence=entry["confidence"],
            notes="override manual curado para valor atual",
        )
        key_variants = {
            normalize_identity(entry["sheet"], entry["name"], entry["subcategory"], entry["brand"], helpers),
        }
        stripped_name = helpers.strip_brand_preserving_tokens(entry["name"], entry["brand"])
        if stripped_name:
            key_variants.add(normalize_identity(entry["sheet"], stripped_name, entry["subcategory"], entry["brand"], helpers))
        for candidate in helpers.extract_model_candidates(entry["name"]):
            key_variants.add(normalize_identity(entry["sheet"], candidate, entry["subcategory"], entry["brand"], helpers))

        for key in key_variants:
            overrides[key] = match
    return overrides


def clone_match(
    helpers,
    base_match,
    *,
    value: float | None = None,
    method: str | None = None,
    source_title: str | None = None,
    source_url: str | None = None,
    confidence: float | None = None,
    notes: str | None = None,
):
    return helpers.SourceMatch(
        value=value if value is not None else base_match.value,
        method=method or base_match.method,
        source_title=source_title or base_match.source_title,
        source_url=source_url or base_match.source_url,
        confidence=base_match.confidence if confidence is None else confidence,
        notes=base_match.notes if notes is None else notes,
    )


def extract_model_keys(value: str | None, helpers) -> set[str]:
    keys = set()
    for candidate in helpers.extract_model_candidates(value):
        compact = helpers.compact_text(candidate)
        if compact:
            keys.add(compact)
    return keys


def current_value_is_outlier(item, value: float, helpers) -> bool:
    original_value = helpers.safe_float(item.valor) or 0.0
    if original_value <= 0:
        return False
    canonical_subcat = helpers.canonical_subcategory(item.subcategoria)
    threshold = 2.0 if canonical_subcat == "BAT MAO" else 2.4 if canonical_subcat == "WIRELESS" else 3.0
    return float(value) > (original_value * threshold)


def source_looks_like_full_system(title: str, helpers) -> bool:
    folded_title = helpers.fold_text(title)
    if not folded_title:
        return False
    if any(marker in folded_title for marker in SET_SOURCE_MARKERS):
        return True
    has_receiver = "RECEPTOR" in folded_title or "RECEIVER" in folded_title
    has_transmitter = any(marker in folded_title for marker in ("TRANSMISSOR", "TRANSMITTER", "HANDHELD", "BASTAO", "MICROFONE"))
    if has_receiver and has_transmitter:
        return True
    return any(marker in folded_title for marker in ("2 BODYPACK", "4 BODYPACK", "4 CANAIS", "4H"))


def should_reject_current_market_match(item, source_match, helpers) -> bool:
    if source_match is None or source_match.value is None:
        return False

    folded_name = helpers.fold_text(item.nome)
    folded_title = helpers.fold_text(source_match.source_title)
    canonical_subcat = helpers.canonical_subcategory(item.subcategoria)
    component_like = canonical_subcat in RISKY_INFERENCE_SUBCATEGORIES or any(
        marker in folded_name for marker in COMPONENT_NAME_MARKERS
    )
    if component_like and source_looks_like_full_system(source_match.source_title, helpers):
        return True

    if not current_value_is_outlier(item, float(source_match.value), helpers):
        return False

    item_models = extract_model_keys(item.nome, helpers) | extract_model_keys(item.subcategoria, helpers)
    title_models = extract_model_keys(source_match.source_title, helpers)
    if item_models and title_models and item_models.isdisjoint(title_models):
        return True

    sparse_identity = len([token for token in helpers.tokenize(item.nome) if len(token) > 1]) <= 1 and not item_models
    if sparse_identity and title_models:
        return True

    return False


def should_use_conservative_current_fallback(item, fallback_match, helpers) -> bool:
    if fallback_match is None or fallback_match.value is None:
        return False
    if not str(fallback_match.method).startswith("estimativa_"):
        return False
    if not current_value_is_outlier(item, float(fallback_match.value), helpers):
        return False

    canonical_subcat = helpers.canonical_subcategory(item.subcategoria)
    folded_name = helpers.fold_text(item.nome)
    return canonical_subcat in RISKY_INFERENCE_SUBCATEGORIES or any(
        marker in folded_name for marker in COMPONENT_NAME_MARKERS
    )


def search_market_match(item, helpers, search_cache: dict[str, list[dict[str, object]]], max_queries: int = 4):
    if item.categoria == "CABO":
        return None

    for query in helpers.build_query_variants(item)[:max_queries]:
        cache_key = helpers.slugify_query(query)
        if cache_key not in search_cache:
            url = f"https://lista.mercadolivre.com.br/{helpers.quote(cache_key, safe='-')}"
            try:
                html = helpers.fetch_url(url)
            except RuntimeError:
                search_cache[cache_key] = []
            else:
                search_cache[cache_key] = helpers.parse_mercadolivre_results(html) if html else []

        market_match = helpers.choose_market_match(item, query, search_cache[cache_key])
        if market_match is not None:
            return market_match

    return None


def estimate_cable_value(name: str, subcategory: str | None, helpers) -> float:
    folded_name = helpers.fold_text(name)
    folded_subcat = helpers.fold_text(subcategory)

    match = helpers.re.search(r"(\d+)\s*[.,]?\s*(\d*)\s*(M|CM)", folded_name)
    meters = 1.0
    if match:
        whole = match.group(1)
        decimal = match.group(2)
        unit = match.group(3)
        value = float(f"{whole}.{decimal}") if decimal else float(whole)
        meters = value / 100 if unit == "CM" else value
    elif "10M" in helpers.compact_text(name):
        meters = 10

    base_by_subcat = {
        "AC": 12,
        "AC 4X": 28,
        "AC 8": 45,
        "ADAPTADOR DIVISOR P10": 20,
        "DISPLAY PORT": 26,
        "DMX": 35,
        "DVI": 24,
        "DVI HDMI": 28,
        "EXTENSOR AC": 18,
        "EXTENSOR HDMI": 32,
        "EXTENSOR USB": 28,
        "EXTENSOR USB A": 28,
        "GERAL": 25,
        "HDMI": 20,
        "HDMI DISPLAY PORT": 30,
        "MIDI DIN": 20,
        "MINI DISPLAYPORT": 26,
        "MINI DPPORT P DPPORT": 26,
        "P10": 20,
        "P10 P10": 24,
        "P10 P2": 22,
        "P10 XLRF": 30,
        "P10 XLRM": 30,
        "P2 P2": 18,
        "P2 USB A": 16,
        "PARALELO": 12,
        "POWERCON AC": 36,
        "POWERCON PIAL": 42,
        "RCA P10": 20,
        "RCA P2": 18,
        "RCA RCA": 18,
        "SPEAK ON": 42,
        "USB A MICRO USB B": 18,
        "USB A USB A": 18,
        "USB A USB B": 18,
        "USB A USB C": 18,
        "USB B USB A": 18,
        "USB B USB C": 20,
        "USB USB": 18,
        "VGA": 22,
        "XLR": 36,
        "XLR P10": 30,
        "XLR P2": 24,
        "XLR USB C": 26,
        "XLRM P10": 30,
        "XLRM RCA": 24,
        "XLRM USB C": 26,
    }
    per_meter_by_subcat = {
        "AC": 7,
        "AC 4X": 14,
        "AC 8": 18,
        "DISPLAY PORT": 9,
        "DMX": 10,
        "DVI": 8,
        "DVI HDMI": 9,
        "EXTENSOR AC": 8,
        "EXTENSOR HDMI": 10,
        "EXTENSOR USB": 10,
        "EXTENSOR USB A": 10,
        "HDMI": 8,
        "HDMI DISPLAY PORT": 10,
        "MIDI DIN": 8,
        "MINI DISPLAYPORT": 9,
        "MINI DPPORT P DPPORT": 9,
        "P10": 8,
        "P10 P10": 8,
        "P10 P2": 7,
        "P10 XLRF": 10,
        "P10 XLRM": 10,
        "P2 P2": 6,
        "P2 USB A": 6,
        "PARALELO": 5,
        "POWERCON AC": 12,
        "POWERCON PIAL": 14,
        "RCA P10": 7,
        "RCA P2": 6,
        "RCA RCA": 6,
        "SPEAK ON": 12,
        "USB A MICRO USB B": 6,
        "USB A USB A": 6,
        "USB A USB B": 6,
        "USB A USB C": 6,
        "USB B USB A": 6,
        "USB B USB C": 6,
        "USB USB": 6,
        "VGA": 8,
        "XLR": 11,
        "XLR P10": 10,
        "XLR P2": 8,
        "XLR USB C": 8,
        "XLRM P10": 10,
        "XLRM RCA": 8,
        "XLRM USB C": 8,
    }

    base = base_by_subcat.get(folded_subcat, 25)
    per_meter = per_meter_by_subcat.get(folded_subcat, 7)
    value = round(base + (meters * per_meter), 2)
    if "BRANCO" in folded_name or "CINZA" in folded_name:
        value += 2
    return value


def fallback_estimate(item, helpers, known_values, global_values) -> helpers.SourceMatch:
    inferred = helpers.infer_from_known_items(item, known_values)
    if inferred and inferred.value is not None:
        return inferred
    value = round(statistics.median(global_values), 2) if global_values else 100.0
    return helpers.SourceMatch(
        value=value,
        method="estimativa_global",
        source_title=item.subcategoria or item.categoria,
        source_url="estimativa_interna",
        confidence=0.2,
        notes="fallback global para nao deixar campo vazio",
    )


def resolve_current_market_value(
    item,
    fallback_match,
    helpers,
    search_cache: dict[str, list[dict[str, object]]],
    current_market_overrides: dict[tuple[str, str, str, str], object],
):
    override_key = normalize_identity(item.categoria, item.nome, item.subcategoria, item.marca, helpers)
    override_match = current_market_overrides.get(override_key)
    if override_match is not None and override_match.value is not None:
        if should_reject_current_market_match(item, override_match, helpers):
            override_match = None
    if override_match is not None and override_match.value is not None:
        return clone_match(
            helpers,
            override_match,
            value=round(float(override_match.value), 2),
            notes="valor_atual via curadoria web de hoje",
        )

    live_match = search_market_match(item, helpers, search_cache)
    if live_match is not None and live_match.value is not None:
        if should_reject_current_market_match(item, live_match, helpers):
            live_match = None
    if live_match is not None and live_match.value is not None:
        return clone_match(
            helpers,
            live_match,
            value=round(float(live_match.value), 2),
            method="mercadolivre_busca_valor_atual",
            notes="valor_atual via busca web de hoje",
        )

    if fallback_match is None or fallback_match.value is None:
        return None

    fallback_method = str(fallback_match.method)
    proxy_confidence = round(max(0.2, min(float(fallback_match.confidence), 0.89)), 3)
    return clone_match(
        helpers,
        fallback_match,
        value=round(float(fallback_match.value), 2),
        method=f"{fallback_method}_fallback_valor_atual",
        confidence=proxy_confidence,
        notes="sem match web atual confiavel; valor_atual mantido por fallback",
    )


def parse_wear_value(desgaste: object) -> float:
    if desgaste is None:
        return 3.0
    if isinstance(desgaste, (int, float)):
        return float(desgaste)

    text = str(desgaste).strip()
    if not text:
        return 3.0
    if "★" in text or "☆" in text:
        stars = text.count("★")
        if stars:
            return float(stars)

    match = re.search(r"\d+(?:[.,]\d+)?", text)
    if match:
        return float(match.group(0).replace(",", "."))
    return 3.0


def compute_current_value(valor_original: float, estado: str | None, desgaste: float | int | None) -> tuple[float, float]:
    factor = STATE_FACTORS.get((estado or "USADO").strip().upper(), 0.65)
    wear = parse_wear_value(desgaste)
    wear = max(1.0, min(wear, 5.0))
    remaining_pct = (wear / 5.0) * factor
    valor_atual = round(valor_original * remaining_pct, 2)
    dep_pct = round((1.0 - remaining_pct) * 100.0, 2)
    return valor_atual, dep_pct


def compute_depreciation_from_values(valor_original: float, valor_atual: float) -> float:
    if valor_original <= 0:
        return 0.0
    return round(((valor_original - valor_atual) / valor_original) * 100.0, 2)


def format_brl(value: float) -> str:
    formatted = f"{value:,.2f}"
    formatted = formatted.replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {formatted}"


def export_markdown(output_path: Path, rows: list[dict[str, object]], summary: dict[str, object], review_rows: list[dict[str, object]]) -> None:
    csv_buffer = io.StringIO()
    writer = csv.DictWriter(
        csv_buffer,
        fieldnames=[
            "Aba",
            "Codigo",
            "Nome",
            "Subcategoria",
            "Marca",
            "Valor Original (R$)",
            "Valor Atual (R$)",
            "Deprec. (%)",
            "Metodo",
            "Confianca",
            "Fonte Titulo",
            "Fonte URL",
            "Metodo Valor Atual",
            "Confianca Valor Atual",
            "Fonte Titulo Valor Atual",
            "Fonte URL Valor Atual",
        ],
    )
    writer.writeheader()
    for row in rows:
        writer.writerow(row)

    review_buffer = io.StringIO()
    review_writer = csv.DictWriter(
        review_buffer,
        fieldnames=[
            "Aba",
            "Nome",
            "Subcategoria",
            "Marca",
            "Valor Original (R$)",
            "Valor Atual (R$)",
            "Metodo Valor Original",
            "Metodo Valor Atual",
            "Confianca Valor Atual",
            "Fonte Titulo Valor Atual",
            "Fonte URL Valor Atual",
        ],
    )
    review_writer.writeheader()
    for row in review_rows:
        review_writer.writerow(row)

    lines = [
        "# Valores Para Importacao",
        "",
        f"Gerado em: {summary['generated_at']}",
        f"Arquivo-base: `{summary['workbook']}`",
        "",
        "## Resumo",
        f"- Seriais exportados: {summary['serial_count']}",
        f"- Itens unicos avaliados: {summary['unique_items']}",
        f"- Itens que tinham valor previo na planilha: {summary['items_with_prior_value']}",
        f"- Valor Original preservado da planilha: {summary['preserved_original_values']}",
        f"- Valor Atual resolvido via web atual ou proxy de mercado: {summary['resolved_current_market']}",
        f"- Distribuicao por metodo valor original: {summary['original_methods']}",
        f"- Distribuicao por metodo valor atual: {summary['current_methods']}",
        "",
        "## CSV Para Importacao",
        "",
        "```csv",
        csv_buffer.getvalue().rstrip(),
        "```",
        "",
        "## Revisao Prioritaria",
        "",
        "Linhas abaixo sao as que ficaram com proxy, estimativa ou confianca abaixo de 0.70 no valor atual.",
        "",
        "```csv",
        review_buffer.getvalue().rstrip(),
        "```",
        "",
    ]
    output_path.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    helpers = load_helpers()
    serials = parse_serial_rows(WORKBOOK_PATH, helpers)
    unique_items = build_unique_items(serials, helpers)
    exact_map, fuzzy_pool = helpers.load_local_reference_sources(ORIGINAL_PATH)
    manual_overrides = build_manual_override_map(helpers)
    current_market_overrides = build_current_market_override_map(helpers)
    search_cache: dict[str, list[dict[str, object]]] = {}

    source_backed_values = []
    source_backed_global_values = []
    resolved_values: dict[tuple[str, str, str | None, str | None], ResolvedPricing] = {}
    unresolved_items = []

    for index, (_, item, _) in enumerate(unique_items, start=1):
        key = (item.categoria, item.nome, item.subcategoria, item.marca)
        if item.valor is None:
            continue

        original_value = round(float(item.valor), 2)
        original_match = helpers.SourceMatch(
            value=original_value,
            method="valor_original_planilha",
            source_title="coluna Valor Original da planilha",
            source_url=f"arquivo_local:{WORKBOOK_PATH.name}#{item.categoria}",
            confidence=1.0,
            notes="valor original preservado conforme planilha",
        )

        match = manual_overrides.get(normalize_identity(item.categoria, item.nome, item.subcategoria, item.marca, helpers))

        if match is None:
            match = helpers.pick_best_local_match(item, exact_map, fuzzy_pool)

        if match is None and item.categoria == "CABO":
            value = estimate_cable_value(item.nome or "", item.subcategoria, helpers)
            match = helpers.SourceMatch(
                value=value,
                method="estimativa_cabo_regra",
                source_title=item.subcategoria or "CABO",
                source_url="estimativa_interna",
                confidence=0.35,
                notes="regra por tipo de cabo e metragem",
            )

        if index % 25 == 0:
            print(f"[precificacao] {index}/{len(unique_items)} itens processados", flush=True)

        current_match = resolve_current_market_value(item, match, helpers, search_cache, current_market_overrides)
        if current_match is None or current_match.value is None:
            unresolved_items.append((key, item, original_value, original_match))
            continue

        resolved_values[key] = ResolvedPricing(
            original_value=original_value,
            original_match=original_match,
            current_value=round(float(current_match.value), 2),
            current_match=current_match,
        )
        if not str(current_match.method).startswith("estimativa"):
            source_backed_values.append((item, round(float(current_match.value), 2)))
            source_backed_global_values.append(round(float(current_match.value), 2))

    for key, item, original_value, original_match in unresolved_items:
        match = fallback_estimate(item, helpers, source_backed_values, source_backed_global_values)
        if should_use_conservative_current_fallback(item, match, helpers):
            match = helpers.SourceMatch(
                value=original_value,
                method="fallback_original_conservador",
                source_title="valor original preservado por salvaguarda contra inferencia inflada",
                source_url=f"arquivo_local:{WORKBOOK_PATH.name}#{item.categoria}",
                confidence=0.25,
                notes="fallback por inferencia descartado por risco de contaminar componente com preco de kit/set",
            )
        current_match = resolve_current_market_value(item, match, helpers, search_cache, current_market_overrides)
        assert current_match.value is not None
        resolved_values[key] = ResolvedPricing(
            original_value=original_value,
            original_match=original_match,
            current_value=round(float(current_match.value), 2),
            current_match=current_match,
        )

    export_rows = []
    review_rows = []
    seen_review = set()
    original_method_counter = Counter()
    current_method_counter = Counter()

    for record in serials:
        key = (record.sheet, record.nome, record.subcategoria, record.marca)
        pricing = resolved_values[key]
        valor_original = pricing.original_value
        valor_atual = pricing.current_value
        dep_pct = compute_depreciation_from_values(valor_original, valor_atual)
        row = {
            "Aba": record.sheet,
            "Codigo": record.codigo,
            "Nome": record.nome,
            "Subcategoria": record.subcategoria or "",
            "Marca": record.marca or "",
            "Valor Original (R$)": format_brl(valor_original),
            "Valor Atual (R$)": format_brl(valor_atual),
            "Deprec. (%)": f"{dep_pct:.2f}",
            "Metodo": pricing.original_match.method,
            "Confianca": f"{pricing.original_match.confidence:.3f}",
            "Fonte Titulo": pricing.original_match.source_title,
            "Fonte URL": pricing.original_match.source_url,
            "Metodo Valor Atual": pricing.current_match.method,
            "Confianca Valor Atual": f"{pricing.current_match.confidence:.3f}",
            "Fonte Titulo Valor Atual": pricing.current_match.source_title,
            "Fonte URL Valor Atual": pricing.current_match.source_url,
        }
        export_rows.append(row)
        original_method_counter[pricing.original_match.method] += 1
        current_method_counter[pricing.current_match.method] += 1
        current_method = str(pricing.current_match.method)
        needs_review = (
            pricing.current_match.confidence < 0.70
            or current_method != "mercadolivre_busca_valor_atual"
        )
        if needs_review:
            review_key = (
                record.sheet,
                record.nome,
                record.subcategoria or "",
                record.marca or "",
                f"{valor_original:.2f}",
                f"{valor_atual:.2f}",
                pricing.original_match.method,
                pricing.current_match.method,
                f"{pricing.current_match.confidence:.3f}",
                pricing.current_match.source_title,
                pricing.current_match.source_url,
            )
            if review_key not in seen_review:
                seen_review.add(review_key)
                review_rows.append(
                    {
                        "Aba": record.sheet,
                        "Nome": record.nome,
                        "Subcategoria": record.subcategoria or "",
                        "Marca": record.marca or "",
                        "Valor Original (R$)": f"{valor_original:.2f}",
                        "Valor Atual (R$)": f"{valor_atual:.2f}",
                        "Metodo Valor Original": pricing.original_match.method,
                        "Metodo Valor Atual": pricing.current_match.method,
                        "Confianca Valor Atual": f"{pricing.current_match.confidence:.3f}",
                        "Fonte Titulo Valor Atual": pricing.current_match.source_title,
                        "Fonte URL Valor Atual": pricing.current_match.source_url,
                    }
                )

    summary = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "workbook": str(WORKBOOK_PATH),
        "serial_count": len(serials),
        "unique_items": len(unique_items),
        "items_with_prior_value": sum(1 for _, item, _ in unique_items if item.valor is not None),
        "preserved_original_values": sum(1 for _, item, _ in unique_items if item.valor is not None),
        "resolved_current_market": len(resolved_values),
        "original_methods": ", ".join(f"{name}={count}" for name, count in original_method_counter.most_common()),
        "current_methods": ", ".join(f"{name}={count}" for name, count in current_method_counter.most_common()),
    }
    export_markdown(OUTPUT_PATH, export_rows, summary, review_rows)
    print(f"Markdown gerado em: {OUTPUT_PATH}")
    print(f"Seriais exportados: {len(serials)}")
    print(f"Itens unicos: {len(unique_items)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
