"""
MMD Estoque Inteligente v2 - Planilha Profissional
13 abas: MANUAL, DASHBOARD, 8 categorias, LOTES, FORA DE OPERACAO, REF CATEGORIAS
Sistema de condicao: Estado + Desgaste + Depreciacao
"""
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.utils import get_column_letter
from collections import defaultdict
import os, datetime

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT = os.path.join(SCRIPT_DIR, '..', 'data', 'inventario-original.xlsx')
OUTPUT = os.path.join(SCRIPT_DIR, '..', 'data', 'inventario-limpo.xlsx')

# ── DESIGN TOKENS ─────────────────────────────────────────────
DARK = '1A1F36'
WHITE = 'FFFFFF'
ALT_BG = 'F8F9FC'
BORDER_CLR = 'E2E8F0'
MUTED = '94A3B8'

CAT_DESIGN = {
    'ILUMINACAO': {'tab': '818CF8', 'bg': 'EEF2FF', 'tx': '4F46E5', 'prefix': 'ILU', 'desc': 'Equipamentos de iluminacao para eventos'},
    'AUDIO':      {'tab': '34D399', 'bg': 'ECFDF5', 'tx': '059669', 'prefix': 'AUD', 'desc': 'Equipamentos de audio e som'},
    'CABO':       {'tab': 'FBBF24', 'bg': 'FFFBEB', 'tx': 'B45309', 'prefix': 'CAB', 'desc': 'Cabos e conexoes'},
    'ENERGIA':    {'tab': 'F87171', 'bg': 'FEF2F2', 'tx': 'DC2626', 'prefix': 'ENE', 'desc': 'Distribuicao de energia'},
    'ESTRUTURA':  {'tab': '60A5FA', 'bg': 'EFF6FF', 'tx': '2563EB', 'prefix': 'EST', 'desc': 'Estruturas e suportes'},
    'EFEITO':     {'tab': 'A78BFA', 'bg': 'F5F3FF', 'tx': '7C3AED', 'prefix': 'EFE', 'desc': 'Efeitos especiais'},
    'VIDEO':      {'tab': '2DD4BF', 'bg': 'F0FDFA', 'tx': '0D9488', 'prefix': 'VID', 'desc': 'Video e projecao'},
    'ACESSORIO':  {'tab': 'FB923C', 'bg': 'FFF7ED', 'tx': 'C2410C', 'prefix': 'ACE', 'desc': 'Acessorios e ferramentas'},
}
CAT_ORDER = ['ILUMINACAO','AUDIO','CABO','ENERGIA','ESTRUTURA','EFEITO','VIDEO','ACESSORIO']

STATUS_DESIGN = {
    'DISPONIVEL': {'bg':'D1FAE5','tx':'065F46'},
    'PACKED':     {'bg':'DBEAFE','tx':'1E40AF'},
    'EM_CAMPO':   {'bg':'FEF3C7','tx':'92400E'},
    'MANUTENCAO': {'bg':'FEE2E2','tx':'991B1B'},
    'EMPRESTADO': {'bg':'EDE9FE','tx':'5B21B6'},
    'VENDIDO':    {'bg':'F1F5F9','tx':'475569'},
    'BAIXA':      {'bg':'F1F5F9','tx':'475569'},
}

ESTADO_DESIGN = {
    'NOVO':           {'bg':'D1FAE5','tx':'065F46'},
    'SEMI_NOVO':      {'bg':'DBEAFE','tx':'1E40AF'},
    'USADO':          {'bg':'FEF3C7','tx':'92400E'},
    'RECONDICIONADO': {'bg':'FED7AA','tx':'9A3412'},
}

DESGASTE_BG = {5:'D1FAE5', 4:'DBEAFE', 3:'FEF3C7', 2:'FED7AA', 1:'FEE2E2'}

CATEGORY_MAP = {
    'PAR LEDS': ('ILUMINACAO','Par Led'), 'PAR LEDS ': ('ILUMINACAO','Par Led'),
    'RIBALTA': ('ILUMINACAO','Ribalta'), 'Ribalta': ('ILUMINACAO','Ribalta'),
    'MINI MOVING': ('ILUMINACAO','Mini Moving'), 'Mini Moving': ('ILUMINACAO','Mini Moving'),
    'MOVING BEAM': ('ILUMINACAO','Moving Beam'), 'Moving Beam': ('ILUMINACAO','Moving Beam'),
    'MINI BRUTE': ('ILUMINACAO','Mini Brute'), 'OUTRAS LUZES': ('ILUMINACAO','Outras Luzes'),
    'LUZ': ('ILUMINACAO','Outras Luzes'), 'LUZ NEGRA': ('ILUMINACAO','Luz Negra'),
    'LASER': ('ILUMINACAO','Laser'), 'STROBO': ('ILUMINACAO','Strobo'),
    'Iluminação': ('ILUMINACAO','Geral'),
    'AUDIO': ('AUDIO','Geral'), 'Audio': ('AUDIO','Geral'), 'audio': ('AUDIO','Geral'),
    'SOM': ('AUDIO','Geral'), 'Som': ('AUDIO','Geral'), 'som': ('AUDIO','Geral'),
    'INSTRUMENTO': ('AUDIO','Instrumento'), 'CAPSULAS MIC': ('AUDIO','Capsula Mic'),
    'MIDI': ('AUDIO','Midi'),
    'CABO': ('CABO','Geral'), 'EXTENSOR': ('CABO','Extensor'),
    'REGUA': ('ENERGIA','Regua'), 'ELETRICA': ('ENERGIA','Eletrica'),
    'Eletrica': ('ENERGIA','Eletrica'), 'Energia': ('ENERGIA','Geral'),
    'FONTE': ('ENERGIA','Fonte'),
    'TRIPÉ': ('ESTRUTURA','Tripe'), 'SUPORTE': ('ESTRUTURA','Suporte'),
    'Estrutura': ('ESTRUTURA','Geral'), 'PRATICÁVEL': ('ESTRUTURA','Praticavel'),
    'Praticável': ('ESTRUTURA','Praticavel'),
    'EFEITOS': ('EFEITO','Geral'), 'Efeitos': ('EFEITO','Geral'),
    'FLUIDO': ('EFEITO','Fluido'), 'Fluido': ('EFEITO','Fluido'),
    'GLOBO': ('EFEITO','Globo'), 'Globo': ('EFEITO','Globo'),
    'VIDEO': ('VIDEO','Geral'), 'COMPUTADORES': ('VIDEO','Computador'),
    'NOTEBOOK': ('VIDEO','Notebook'), 'Notebook': ('VIDEO','Notebook'),
    'TABLET': ('VIDEO','Tablet'), 'Tablet': ('VIDEO','Tablet'),
    'CASE': ('ACESSORIO','Case'), 'CASES': ('ACESSORIO','Case'),
    'FERRAMENTAS': ('ACESSORIO','Ferramenta'),
    'ROTEADOR': ('ACESSORIO','Roteador'), 'Roteador': ('ACESSORIO','Roteador'),
    'REDE': ('ACESSORIO','Rede'), 'RADIO': ('ACESSORIO','Radio'), 'Radio': ('ACESSORIO','Radio'),
    'Refrigeracao': ('ACESSORIO','Refrigeracao'), 'REFRIGERACAO': ('ACESSORIO','Refrigeracao'),
}

ESTADO_FATOR = {'NOVO':1.0, 'SEMI_NOVO':0.85, 'USADO':0.65, 'RECONDICIONADO':0.50}

# ── HELPERS ───────────────────────────────────────────────────
def F(sz=10, bold=False, color=DARK):
    return Font(name='Arial', size=sz, bold=bold, color=color)

def BG(color):
    return PatternFill(start_color=color, end_color=color, fill_type='solid')

def BD():
    s = Side(style='thin', color=BORDER_CLR)
    return Border(left=s, right=s, top=s, bottom=s)

def title_case_brand(brand):
    if not brand or brand.upper() in ('SEM MARCA','S/ MARCA','NONE','S/MARCA','CHINES',''):
        return ''
    words = brand.strip().split()
    result = []
    for w in words:
        if len(w) <= 3 and w.upper() == w:
            result.append(w.upper())
        else:
            result.append(w.title())
    return ' '.join(result)

def hdr_style(ws, row, ncols):
    for c in range(1, ncols+1):
        cl = ws.cell(row=row, column=c)
        cl.font = F(10, True, WHITE)
        cl.fill = BG(DARK)
        cl.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cl.border = BD()

def data_row_style(ws, row, ncols, alt=False):
    bg = BG(ALT_BG) if alt else BG(WHITE)
    for c in range(1, ncols+1):
        cl = ws.cell(row=row, column=c)
        cl.fill = bg
        cl.border = BD()
        cl.alignment = Alignment(vertical='center')
        cl.font = F(10)

def widths(ws, ws_list):
    for i, w in enumerate(ws_list, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ── READ DATA ─────────────────────────────────────────────────
def read_data():
    wb = load_workbook(INPUT, data_only=True)
    ws = wb['EQUIPAMENTOS']
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=False):
        ok = row[1].value
        cat_raw = str(row[2].value).strip() if row[2].value else ''
        modelo = str(row[5].value).strip() if row[5].value else ''
        if not cat_raw and not modelo:
            continue
        subcat_raw = str(row[3].value).strip() if row[3].value else ''
        marca_raw = str(row[4].value).strip() if row[4].value else ''
        serial_fab = str(row[9].value).strip() if row[9].value else ''
        valor = row[10].value
        notas = str(row[13].value).strip() if len(row) > 13 and row[13].value else ''

        mapped = CATEGORY_MAP.get(cat_raw, ('ACESSORIO','Outro'))
        cat, subcat = mapped
        if subcat_raw and subcat_raw.upper() not in ('SEM SUB','NONE',''):
            subcat = subcat_raw.title()

        has_serial = serial_fab and 'S/ SERIAL' not in serial_fab.upper() and 'SEM' not in serial_fab.upper() and serial_fab.upper() != 'NONE'
        valor_num = None
        if valor:
            try: valor_num = float(valor)
            except: pass

        marca = title_case_brand(marca_raw)

        ok_str = str(ok).strip().upper() if ok else ''
        notas_str = notas.upper() if notas else ''
        if 'VENDID' in ok_str or 'VENDID' in notas_str:
            status = 'VENDIDO'
        elif 'BIGODE' in ok_str:
            status = 'EMPRESTADO'
        elif 'QUEBRAD' in ok_str:
            status = 'BAIXA'
        else:
            status = 'DISPONIVEL'

        motivo = ''
        if status == 'VENDIDO':
            motivo = notas if notas and notas != 'None' else 'Vendido'
        elif status == 'EMPRESTADO':
            motivo = 'Emprestado - Seu Bigode'
        elif status == 'BAIXA':
            motivo = 'Quebrado'

        rows.append({
            'categoria': cat, 'subcategoria': subcat, 'marca': marca,
            'modelo': modelo, 'serial_fabrica': serial_fab if has_serial else '',
            'valor_mercado': valor_num, 'status': status, 'motivo': motivo,
            'notas': notas if notas and notas != 'None' else '',
        })
    return rows

def assign_codes(rows):
    counters = defaultdict(int)
    for r in rows:
        counters[r['categoria']] += 1
        pfx = CAT_DESIGN[r['categoria']]['prefix']
        r['codigo'] = f"MMD-{pfx}-{counters[r['categoria']]:04d}"
        r['estado'] = 'USADO'
        r['desgaste'] = 3
    return rows

def make_lots(rows):
    cable_groups = defaultdict(list)
    for r in rows:
        if r['categoria'] == 'CABO' and r['status'] == 'DISPONIVEL':
            key = (r['marca'], r['modelo'], r['subcategoria'])
            cable_groups[key].append(r)
    lots = []
    for i, ((marca, modelo, subcat), items) in enumerate(sorted(cable_groups.items()), 1):
        nome = f"{marca} {modelo}".strip() if marca else modelo
        lots.append({
            'codigo': f"MMD-LOT-{i:03d}",
            'descricao': f"Lote {len(items)}x {nome}",
            'subcategoria': subcat, 'quantidade': len(items),
            'status': 'DISPONIVEL', 'tag_rfid': '', 'qr_code': f"https://mmd.app/l/MMD-LOT-{i:03d}",
        })
    return lots

# ── TAB: MANUAL ───────────────────────────────────────────────
def build_manual(wb):
    ws = wb.active
    ws.title = 'MANUAL'
    ws.sheet_properties.tabColor = '808080'
    widths(ws, [3, 80])
    ws.column_dimensions['A'].width = 3

    sections = [
        ('MANUAL DE USO — INVENTARIO MMD EVENTOS', 18, True, DARK),
        ('', 10, False, MUTED),
        ('1. COMO NAVEGAR', 13, True, DARK),
        ('Cada aba na parte inferior corresponde a uma categoria de equipamento.', 10, False, '374151'),
        ('Use as abas: ILUMINACAO, AUDIO, CABO, ENERGIA, ESTRUTURA, EFEITO, VIDEO, ACESSORIO.', 10, False, '374151'),
        ('O DASHBOARD mostra o resumo geral com graficos.', 10, False, '374151'),
        ('Use os filtros no header de cada coluna para buscar itens especificos.', 10, False, '374151'),
        ('', 8, False, MUTED),
        ('2. COLUNAS', 13, True, DARK),
        ('Codigo — Identificador unico do item (ex: MMD-ILU-0001). Nunca muda.', 10, False, '374151'),
        ('Nome — Descricao do equipamento (marca + modelo).', 10, False, '374151'),
        ('Subcategoria — Tipo especifico (Par Led, Moving, Caixa de Som, etc).', 10, False, '374151'),
        ('Marca — Fabricante do equipamento.', 10, False, '374151'),
        ('Serial Fab. — Numero de serie do fabricante (quando disponivel).', 10, False, '374151'),
        ('Tag RFID — ID da tag RFID colada no equipamento (preenchido apos instalacao).', 10, False, '374151'),
        ('Status — Onde o item esta agora (ver secao 3).', 10, False, '374151'),
        ('Estado — Classificacao do ciclo de vida (ver secao 4).', 10, False, '374151'),
        ('Desgaste — Nota de 1 a 5 da condicao fisica (ver secao 4).', 10, False, '374151'),
        ('Valor Original — Preco de mercado quando novo (R$).', 10, False, '374151'),
        ('Valor Atual — Valor depreciado, calculado automaticamente.', 10, False, '374151'),
        ('Deprec.% — Percentual do valor original ja depreciado.', 10, False, '374151'),
        ('Notas — Observacoes livres.', 10, False, '374151'),
        ('', 8, False, MUTED),
        ('3. STATUS (onde o item esta)', 13, True, DARK),
        ('DISPONIVEL — No galpao, pronto para uso.', 10, False, '065F46'),
        ('EM_CAMPO — Saiu para evento, em uso.', 10, False, '92400E'),
        ('PACKED — Separado para evento, ainda no galpao.', 10, False, '1E40AF'),
        ('MANUTENCAO — Em reparo.', 10, False, '991B1B'),
        ('EMPRESTADO — Com terceiro.', 10, False, '5B21B6'),
        ('VENDIDO — Vendido, nao faz mais parte do estoque.', 10, False, '475569'),
        ('BAIXA — Descartado ou perdido.', 10, False, '475569'),
        ('', 8, False, MUTED),
        ('4. CONDICAO DO EQUIPAMENTO', 13, True, DARK),
        ('', 6, False, MUTED),
        ('ESTADO (historia do item):', 11, True, '374151'),
        ('NOVO — Nunca foi para evento. Embalagem original.', 10, False, '065F46'),
        ('SEMI_NOVO — Usado 1 a 5 vezes. Sem marcas visiveis.', 10, False, '1E40AF'),
        ('USADO — Uso regular. Marcas normais de operacao.', 10, False, '92400E'),
        ('RECONDICIONADO — Foi reparado ou restaurado.', 10, False, '9A3412'),
        ('', 6, False, MUTED),
        ('DESGASTE (condicao fisica atual, nota 1 a 5):', 11, True, '374151'),
        ('5 = Excelente (como novo, sem marcas)', 10, False, '065F46'),
        ('4 = Bom (marcas leves de uso, 100% funcional)', 10, False, '1E40AF'),
        ('3 = Regular (uso visivel, funcional, cosmetico ok)', 10, False, '92400E'),
        ('2 = Desgastado (funcional mas com problemas cosmeticos)', 10, False, '9A3412'),
        ('1 = Critico (precisa reparo ou substituicao)', 10, False, '991B1B'),
        ('', 6, False, MUTED),
        ('DEPRECIACAO (calculada automaticamente):', 11, True, '374151'),
        ('Valor Atual = Valor Original x (Desgaste / 5) x Fator do Estado', 10, False, '374151'),
        ('Deprec.% = 100% - (Valor Atual / Valor Original)', 10, False, '374151'),
        ('Fatores: NOVO=100%, SEMI_NOVO=85%, USADO=65%, RECONDICIONADO=50%', 10, False, '6B7280'),
        ('', 8, False, MUTED),
        ('5. COMO ATUALIZAR', 13, True, DARK),
        ('Item novo: adicionar na aba da categoria. Estado=NOVO, Desgaste=5.', 10, False, '374151'),
        ('Apos evento: revisar desgaste dos itens que foram.', 10, False, '374151'),
        ('Item com defeito: mudar status para MANUTENCAO, ajustar desgaste.', 10, False, '374151'),
        ('Item vendido: mover para aba FORA DE OPERACAO, status=VENDIDO.', 10, False, '374151'),
        ('', 8, False, MUTED),
        ('6. CORES', 13, True, DARK),
        ('Verde = disponivel / excelente / novo', 10, False, '065F46'),
        ('Azul = packed / bom / semi-novo', 10, False, '1E40AF'),
        ('Amarelo = em campo / regular / usado', 10, False, '92400E'),
        ('Laranja = desgastado / recondicionado', 10, False, '9A3412'),
        ('Vermelho = manutencao / critico', 10, False, '991B1B'),
        ('Roxo = emprestado', 10, False, '5B21B6'),
        ('Cinza = vendido / baixa', 10, False, '475569'),
    ]
    for i, (text, size, bold, color) in enumerate(sections, 1):
        cell = ws.cell(row=i, column=2)
        cell.value = text
        cell.font = F(size, bold, color)
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        ws.row_dimensions[i].height = 28 if size >= 13 else (12 if not text else 20)

# ── TAB: DASHBOARD ────────────────────────────────────────────
def build_dashboard(wb, serials, lots):
    ws = wb.create_sheet('DASHBOARD')
    ws.sheet_properties.tabColor = '6366F1'
    widths(ws, [3, 20, 14, 16, 14, 14, 14, 14, 3, 20, 14])

    activos = [s for s in serials if s['status'] not in ('VENDIDO','EMPRESTADO','BAIXA')]
    total = len(activos)
    disponivel = sum(1 for s in activos if s['status']=='DISPONIVEL')
    em_campo = sum(1 for s in activos if s['status']=='EM_CAMPO')
    manut = sum(1 for s in activos if s['status']=='MANUTENCAO')
    valor_orig = sum(s['valor_mercado'] or 0 for s in activos)
    desgastes = [s['desgaste'] for s in activos]
    avg_desg = sum(desgastes)/len(desgastes) if desgastes else 0
    criticos = sum(1 for d in desgastes if d <= 2)

    # Title
    ws.merge_cells('B2:H2')
    ws['B2'].value = 'MMD EVENTOS — Inventario Inteligente'
    ws['B2'].font = F(20, True, '1A1F36')
    ws.row_dimensions[2].height = 40

    ws.merge_cells('B3:H3')
    ws['B3'].value = f'Atualizado em: {datetime.date.today().strftime("%d/%m/%Y")}'
    ws['B3'].font = F(11, False, MUTED)

    # KPIs row
    kpis = [
        ('Total Itens', total, '6366F1'),
        ('Disponivel', disponivel, '10B981'),
        ('Em Campo', em_campo, 'F59E0B'),
        ('Manutencao', manut, 'EF4444'),
        ('Valor Original', valor_orig, '6366F1'),
    ]
    for i, (label, val, color) in enumerate(kpis):
        col = 2 + i
        c_val = ws.cell(row=5, column=col)
        c_val.value = val
        if isinstance(val, float):
            c_val.number_format = 'R$ #,##0'
        c_val.font = F(20, True, color)
        c_val.alignment = Alignment(horizontal='center')
        c_lbl = ws.cell(row=6, column=col)
        c_lbl.value = label
        c_lbl.font = F(9, False, MUTED)
        c_lbl.alignment = Alignment(horizontal='center')
    ws.row_dimensions[5].height = 38
    ws.row_dimensions[6].height = 18

    # Saude
    ws.merge_cells('B8:D8')
    ws['B8'].value = 'SAUDE DO INVENTARIO'
    ws['B8'].font = F(10, True, MUTED)
    health = [
        ('Media Desgaste', f'{avg_desg:.1f}/5', '6366F1'),
        ('Itens Criticos', criticos, 'EF4444'),
    ]
    for i, (lbl, val, clr) in enumerate(health):
        ws.cell(row=9, column=2+i).value = val
        ws.cell(row=9, column=2+i).font = F(16, True, clr)
        ws.cell(row=9, column=2+i).alignment = Alignment(horizontal='center')
        ws.cell(row=10, column=2+i).value = lbl
        ws.cell(row=10, column=2+i).font = F(9, False, MUTED)
        ws.cell(row=10, column=2+i).alignment = Alignment(horizontal='center')

    # Por Categoria
    ws.merge_cells('B12:E12')
    ws['B12'].value = 'DISTRIBUICAO POR CATEGORIA'
    ws['B12'].font = F(10, True, MUTED)

    cat_headers = ['Categoria', 'Qtd', 'Valor (R$)', '%']
    for i, h in enumerate(cat_headers):
        cl = ws.cell(row=13, column=2+i)
        cl.value = h
        cl.font = F(10, True, WHITE)
        cl.fill = BG(DARK)
        cl.alignment = Alignment(horizontal='center', vertical='center')
        cl.border = BD()

    cat_data = defaultdict(lambda: {'count':0, 'value':0})
    for s in activos:
        cat_data[s['categoria']]['count'] += 1
        cat_data[s['categoria']]['value'] += s['valor_mercado'] or 0

    r = 14
    for cat in CAT_ORDER:
        d = cat_data.get(cat, {'count':0,'value':0})
        ws.cell(row=r, column=2).value = cat
        ws.cell(row=r, column=2).font = F(10, True, CAT_DESIGN[cat]['tx'])
        ws.cell(row=r, column=3).value = d['count']
        ws.cell(row=r, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=4).value = d['value']
        ws.cell(row=r, column=4).number_format = '#,##0'
        pct = d['count']/total if total else 0
        ws.cell(row=r, column=5).value = pct
        ws.cell(row=r, column=5).number_format = '0.0%'
        ws.cell(row=r, column=5).alignment = Alignment(horizontal='center')
        data_row_style(ws, r, 5, alt=(r%2==0))
        ws.cell(row=r, column=2).font = F(10, True, CAT_DESIGN[cat]['tx'])
        r += 1

    # Pie chart
    pie = PieChart()
    pie.title = None
    pie.style = 10
    pie.width = 14
    pie.height = 10
    cats_ref = Reference(ws, min_col=2, min_row=14, max_row=14+len(CAT_ORDER)-1)
    vals_ref = Reference(ws, min_col=3, min_row=14, max_row=14+len(CAT_ORDER)-1)
    pie.add_data(vals_ref)
    pie.set_categories(cats_ref)
    ws.add_chart(pie, 'G12')

    # Por Status
    r_status = 14 + len(CAT_ORDER) + 2
    ws.merge_cells(f'B{r_status}:D{r_status}')
    ws.cell(row=r_status, column=2).value = 'DISTRIBUICAO POR STATUS'
    ws.cell(row=r_status, column=2).font = F(10, True, MUTED)

    r_status += 1
    for i, h in enumerate(['Status', 'Qtd', '%']):
        cl = ws.cell(row=r_status, column=2+i)
        cl.value = h
        cl.font = F(10, True, WHITE)
        cl.fill = BG(DARK)
        cl.alignment = Alignment(horizontal='center', vertical='center')
        cl.border = BD()

    status_counts = defaultdict(int)
    for s in activos:
        status_counts[s['status']] += 1

    r_s = r_status + 1
    status_start = r_s
    for st in ['DISPONIVEL','EM_CAMPO','PACKED','MANUTENCAO']:
        cnt = status_counts.get(st, 0)
        if cnt == 0: continue
        ws.cell(row=r_s, column=2).value = st
        sd = STATUS_DESIGN.get(st, {})
        ws.cell(row=r_s, column=2).font = F(10, True, sd.get('tx', DARK))
        ws.cell(row=r_s, column=3).value = cnt
        ws.cell(row=r_s, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=r_s, column=4).value = cnt/total if total else 0
        ws.cell(row=r_s, column=4).number_format = '0.0%'
        ws.cell(row=r_s, column=4).alignment = Alignment(horizontal='center')
        data_row_style(ws, r_s, 4, alt=(r_s%2==0))
        ws.cell(row=r_s, column=2).font = F(10, True, sd.get('tx', DARK))
        r_s += 1

    # Bar chart
    if r_s > status_start:
        bar = BarChart()
        bar.type = 'bar'
        bar.style = 10
        bar.title = None
        bar.width = 14
        bar.height = 8
        bar.y_axis.delete = True
        vals_ref = Reference(ws, min_col=3, min_row=status_start, max_row=r_s-1)
        cats_ref = Reference(ws, min_col=2, min_row=status_start, max_row=r_s-1)
        bar.add_data(vals_ref)
        bar.set_categories(cats_ref)
        bar.legend = None
        ws.add_chart(bar, f'F{r_status}')

    # Top 10
    r_top = r_s + 2
    ws.merge_cells(f'B{r_top}:F{r_top}')
    ws.cell(row=r_top, column=2).value = 'TOP 10 ITENS MAIS VALIOSOS'
    ws.cell(row=r_top, column=2).font = F(10, True, MUTED)
    r_top += 1
    for i, h in enumerate(['Nome','Categoria','Valor Original','Desgaste']):
        cl = ws.cell(row=r_top, column=2+i)
        cl.value = h
        cl.font = F(10, True, WHITE)
        cl.fill = BG(DARK)
        cl.alignment = Alignment(horizontal='center', vertical='center')
        cl.border = BD()
    r_top += 1

    top10 = sorted([s for s in activos if s['valor_mercado']], key=lambda x: -(x['valor_mercado'] or 0))[:10]
    for idx, s in enumerate(top10):
        nome = f"{s['marca']} {s['modelo']}".strip() if s['marca'] else s['modelo']
        ws.cell(row=r_top, column=2).value = nome
        ws.cell(row=r_top, column=2).font = F(10, True)
        ws.cell(row=r_top, column=3).value = s['categoria']
        ws.cell(row=r_top, column=3).font = F(10, False, CAT_DESIGN[s['categoria']]['tx'])
        ws.cell(row=r_top, column=4).value = s['valor_mercado']
        ws.cell(row=r_top, column=4).number_format = 'R$ #,##0'
        ws.cell(row=r_top, column=5).value = s['desgaste']
        ws.cell(row=r_top, column=5).alignment = Alignment(horizontal='center')
        dbg = DESGASTE_BG.get(s['desgaste'], 'FEF3C7')
        ws.cell(row=r_top, column=5).fill = BG(dbg)
        data_row_style(ws, r_top, 5, alt=(idx%2==0))
        ws.cell(row=r_top, column=2).font = F(10, True)
        ws.cell(row=r_top, column=3).font = F(10, False, CAT_DESIGN[s['categoria']]['tx'])
        ws.cell(row=r_top, column=5).fill = BG(dbg)
        r_top += 1

# ── TAB: CATEGORY ─────────────────────────────────────────────
COLS = ['Codigo','Nome','Subcategoria','Marca','Serial Fab.','Tag RFID','Status','Estado','Desgaste','Valor Original','Valor Atual','Deprec.%','Notas']
COL_W = [18, 38, 18, 16, 20, 16, 14, 16, 12, 14, 14, 10, 28]
NCOLS = len(COLS)

def build_category_tab(wb, cat, items):
    design = CAT_DESIGN[cat]
    ws = wb.create_sheet(cat)
    ws.sheet_properties.tabColor = design['tab']
    widths(ws, COL_W)

    # Title
    ws.merge_cells(f'A1:{get_column_letter(NCOLS)}1')
    ws['A1'].value = cat
    ws['A1'].font = F(16, True, design['tx'])
    ws['A1'].alignment = Alignment(vertical='center')
    ws.row_dimensions[1].height = 38

    vals = [i['valor_mercado'] or 0 for i in items]
    degs = [i['desgaste'] for i in items]
    avg_d = sum(degs)/len(degs) if degs else 0
    ws.merge_cells(f'A2:{get_column_letter(NCOLS)}2')
    ws['A2'].value = f'{len(items)} itens  |  Valor total: R$ {sum(vals):,.0f}  |  Media desgaste: {avg_d:.1f}/5'
    ws['A2'].font = F(10, False, MUTED)
    ws.row_dimensions[3].height = 6

    # Header
    for i, h in enumerate(COLS):
        ws.cell(row=4, column=i+1).value = h
    hdr_style(ws, 4, NCOLS)
    ws.row_dimensions[4].height = 30
    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f'A4:{get_column_letter(NCOLS)}4'

    # Group by subcategory
    by_sub = defaultdict(list)
    for item in items:
        by_sub[item['subcategoria']].append(item)

    r = 5
    data_idx = 0
    for subcat in sorted(by_sub.keys()):
        # Separator
        ws.merge_cells(f'A{r}:{get_column_letter(NCOLS)}{r}')
        ws.cell(row=r, column=1).value = f'  {subcat} ({len(by_sub[subcat])})'
        ws.cell(row=r, column=1).font = F(11, True, design['tx'])
        ws.cell(row=r, column=1).fill = BG(design['bg'])
        ws.cell(row=r, column=1).alignment = Alignment(vertical='center')
        ws.row_dimensions[r].height = 26
        for c in range(1, NCOLS+1):
            ws.cell(row=r, column=c).fill = BG(design['bg'])
            ws.cell(row=r, column=c).border = BD()
        r += 1

        for item in sorted(by_sub[subcat], key=lambda x: x['codigo']):
            nome = f"{item['marca']} {item['modelo']}".strip() if item['marca'] else item['modelo']
            ws.cell(row=r, column=1).value = item['codigo']
            ws.cell(row=r, column=2).value = nome
            ws.cell(row=r, column=3).value = item['subcategoria']
            ws.cell(row=r, column=4).value = item['marca']
            ws.cell(row=r, column=5).value = item['serial_fabrica']
            ws.cell(row=r, column=6).value = ''  # Tag RFID
            ws.cell(row=r, column=7).value = item['status']
            ws.cell(row=r, column=8).value = item['estado']
            ws.cell(row=r, column=9).value = item['desgaste']
            ws.cell(row=r, column=10).value = item['valor_mercado']
            if item['valor_mercado']:
                ws.cell(row=r, column=10).number_format = '#,##0'
                # Valor Atual formula
                ws.cell(row=r, column=11).value = f'=J{r}*(I{r}/5)*IF(H{r}="NOVO",1,IF(H{r}="SEMI_NOVO",0.85,IF(H{r}="USADO",0.65,0.5)))'
                ws.cell(row=r, column=11).number_format = '#,##0'
                # Deprec % formula
                ws.cell(row=r, column=12).value = f'=IF(J{r}>0,(J{r}-K{r})/J{r},"")'
                ws.cell(row=r, column=12).number_format = '0%'
            ws.cell(row=r, column=13).value = item['notas']

            data_row_style(ws, r, NCOLS, alt=(data_idx%2==0))
            ws.row_dimensions[r].height = 22

            # Color codigo
            ws.cell(row=r, column=1).font = F(10, True, design['tx'])
            # Color status
            sd = STATUS_DESIGN.get(item['status'], {})
            if sd:
                ws.cell(row=r, column=7).font = F(10, True, sd['tx'])
                ws.cell(row=r, column=7).fill = BG(sd['bg'])
            # Color estado
            ed = ESTADO_DESIGN.get(item['estado'], {})
            if ed:
                ws.cell(row=r, column=8).font = F(10, False, ed['tx'])
                ws.cell(row=r, column=8).fill = BG(ed['bg'])
            # Color desgaste
            dbg = DESGASTE_BG.get(item['desgaste'], 'FEF3C7')
            ws.cell(row=r, column=9).fill = BG(dbg)
            ws.cell(row=r, column=9).alignment = Alignment(horizontal='center')
            ws.cell(row=r, column=9).font = F(10, True)

            data_idx += 1
            r += 1

# ── TAB: LOTES ────────────────────────────────────────────────
def build_lotes(wb, lots):
    ws = wb.create_sheet('LOTES')
    ws.sheet_properties.tabColor = 'F59E0B'
    headers = ['Codigo Lote','Descricao','Subcategoria','Quantidade','Status','Tag RFID','QR Code']
    widths(ws, [18, 40, 18, 12, 14, 16, 32])
    for i, h in enumerate(headers):
        ws.cell(row=1, column=i+1).value = h
    hdr_style(ws, 1, len(headers))
    ws.row_dimensions[1].height = 30
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'
    ws.freeze_panes = 'A2'

    for idx, lot in enumerate(lots):
        r = idx + 2
        ws.cell(row=r, column=1).value = lot['codigo']
        ws.cell(row=r, column=2).value = lot['descricao']
        ws.cell(row=r, column=3).value = lot['subcategoria']
        ws.cell(row=r, column=4).value = lot['quantidade']
        ws.cell(row=r, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=5).value = lot['status']
        ws.cell(row=r, column=6).value = lot['tag_rfid']
        ws.cell(row=r, column=7).value = lot['qr_code']
        data_row_style(ws, r, len(headers), alt=(idx%2==0))
        ws.cell(row=r, column=1).font = F(10, True, 'B45309')
        sd = STATUS_DESIGN.get(lot['status'], {})
        if sd:
            ws.cell(row=r, column=5).font = F(10, True, sd['tx'])
            ws.cell(row=r, column=5).fill = BG(sd['bg'])
        ws.row_dimensions[r].height = 22

# ── TAB: FORA DE OPERACAO ─────────────────────────────────────
def build_fora(wb, items):
    ws = wb.create_sheet('FORA DE OPERACAO')
    ws.sheet_properties.tabColor = '6B7280'
    headers = COLS + ['Motivo']
    w = COL_W + [30]
    widths(ws, w)
    for i, h in enumerate(headers):
        ws.cell(row=1, column=i+1).value = h
    hdr_style(ws, 1, len(headers))
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'

    for idx, item in enumerate(items):
        r = idx + 2
        nome = f"{item['marca']} {item['modelo']}".strip() if item['marca'] else item['modelo']
        ws.cell(row=r, column=1).value = item['codigo']
        ws.cell(row=r, column=2).value = nome
        ws.cell(row=r, column=3).value = item['subcategoria']
        ws.cell(row=r, column=4).value = item['marca']
        ws.cell(row=r, column=5).value = item['serial_fabrica']
        ws.cell(row=r, column=6).value = ''
        ws.cell(row=r, column=7).value = item['status']
        ws.cell(row=r, column=8).value = item['estado']
        ws.cell(row=r, column=9).value = item['desgaste']
        ws.cell(row=r, column=10).value = item['valor_mercado']
        if item['valor_mercado']:
            ws.cell(row=r, column=10).number_format = '#,##0'
        ws.cell(row=r, column=13).value = item['notas']
        ws.cell(row=r, column=14).value = item['motivo']
        data_row_style(ws, r, len(headers), alt=(idx%2==0))
        ws.cell(row=r, column=1).font = F(10, True, MUTED)
        sd = STATUS_DESIGN.get(item['status'], {})
        if sd:
            ws.cell(row=r, column=7).font = F(10, True, sd['tx'])
            ws.cell(row=r, column=7).fill = BG(sd['bg'])
        ws.row_dimensions[r].height = 22

# ── TAB: REF CATEGORIAS ──────────────────────────────────────
def build_ref(wb, serials):
    ws = wb.create_sheet('REF CATEGORIAS')
    ws.sheet_properties.tabColor = '94A3B8'
    widths(ws, [3, 22, 30, 35, 10])

    subcat_data = defaultdict(lambda: defaultdict(lambda: {'count':0, 'examples': set()}))
    for s in serials:
        sc = subcat_data[s['categoria']][s['subcategoria']]
        sc['count'] += 1
        nome = f"{s['marca']} {s['modelo']}".strip() if s['marca'] else s['modelo']
        if len(sc['examples']) < 2:
            sc['examples'].add(nome)

    r = 1
    for cat in CAT_ORDER:
        design = CAT_DESIGN[cat]
        # Header block
        ws.merge_cells(f'B{r}:E{r}')
        ws.cell(row=r, column=2).value = f'{cat} ({design["prefix"]})'
        ws.cell(row=r, column=2).font = F(14, True, design['tx'])
        ws.cell(row=r, column=2).fill = BG(design['bg'])
        for c in range(2, 6):
            ws.cell(row=r, column=c).fill = BG(design['bg'])
            ws.cell(row=r, column=c).border = BD()
        ws.row_dimensions[r].height = 32
        r += 1

        ws.cell(row=r, column=2).value = design['desc']
        ws.cell(row=r, column=2).font = F(10, False, MUTED)
        r += 1

        # Sub header
        for i, h in enumerate(['Subcategoria','Exemplo','Qtd']):
            cl = ws.cell(row=r, column=2+i)
            cl.value = h
            cl.font = F(9, True, '6B7280')
            cl.border = BD()
        r += 1

        subs = subcat_data.get(cat, {})
        for subcat in sorted(subs.keys()):
            sd = subs[subcat]
            ws.cell(row=r, column=2).value = subcat
            ws.cell(row=r, column=2).font = F(10)
            ws.cell(row=r, column=3).value = ', '.join(sorted(sd['examples']))
            ws.cell(row=r, column=3).font = F(9, False, '6B7280')
            ws.cell(row=r, column=4).value = sd['count']
            ws.cell(row=r, column=4).alignment = Alignment(horizontal='center')
            ws.cell(row=r, column=4).font = F(10, True)
            for c in range(2, 5):
                ws.cell(row=r, column=c).border = BD()
            r += 1

        r += 1  # spacer

# ── MAIN ──────────────────────────────────────────────────────
if __name__ == '__main__':
    print('Lendo inventario original...')
    rows = read_data()
    print(f'  {len(rows)} itens lidos')

    print('Atribuindo codigos e defaults...')
    rows = assign_codes(rows)

    activos = [r for r in rows if r['status'] not in ('VENDIDO','EMPRESTADO','BAIXA')]
    fora = [r for r in rows if r['status'] in ('VENDIDO','EMPRESTADO','BAIXA')]
    print(f'  {len(activos)} ativos, {len(fora)} fora de operacao')

    lots = make_lots(rows)
    print(f'  {len(lots)} lotes de cabos')

    # Stats
    cats = defaultdict(int)
    for r in activos: cats[r['categoria']] += 1
    print('\nPor categoria (ativos):')
    for c in CAT_ORDER:
        print(f'  {c}: {cats.get(c,0)}')

    print('\nConstruindo planilha v2...')
    wb = Workbook()

    build_manual(wb)
    print('  [1/13] MANUAL')

    build_dashboard(wb, activos, lots)
    print('  [2/13] DASHBOARD')

    for i, cat in enumerate(CAT_ORDER):
        cat_items = [r for r in activos if r['categoria'] == cat]
        build_category_tab(wb, cat, cat_items)
        print(f'  [{i+3}/13] {cat} ({len(cat_items)} itens)')

    build_lotes(wb, lots)
    print('  [11/13] LOTES')

    build_fora(wb, fora)
    print(f'  [12/13] FORA DE OPERACAO ({len(fora)} itens)')

    build_ref(wb, activos)
    print('  [13/13] REF CATEGORIAS')

    wb.save(OUTPUT)
    print(f'\nSalvo em: {OUTPUT}')
    print('Pronto!')
