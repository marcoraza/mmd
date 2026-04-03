#!/usr/bin/env python3
from __future__ import annotations

import csv
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


PROJECT_DIR = Path(__file__).resolve().parent.parent
WORKBOOK_PATH = PROJECT_DIR / "data" / "inventario-limpo.xlsx"
MARKDOWN_PATH = PROJECT_DIR / "data" / "valores-para-importacao.md"
ACTIVE_SHEETS = ["ILUMINACAO", "AUDIO", "ENERGIA", "ESTRUTURA", "EFEITO", "VIDEO", "ACESSORIO"]


def extract_csv_block(markdown_text: str) -> str:
    marker = "## CSV Para Importacao"
    start = markdown_text.find(marker)
    if start == -1:
        raise RuntimeError("Secao 'CSV Para Importacao' nao encontrada no markdown.")

    fence_start = markdown_text.find("```csv", start)
    if fence_start == -1:
        raise RuntimeError("Bloco CSV nao encontrado no markdown.")

    content_start = fence_start + len("```csv")
    fence_end = markdown_text.find("```", content_start)
    if fence_end == -1:
        raise RuntimeError("Fim do bloco CSV nao encontrado.")

    return markdown_text[content_start:fence_end].strip()


def load_price_map(markdown_path: Path) -> dict[str, dict[str, str]]:
    csv_text = extract_csv_block(markdown_path.read_text(encoding="utf-8"))
    reader = csv.DictReader(csv_text.splitlines())
    price_map: dict[str, dict[str, str]] = {}
    for row in reader:
        code = (row.get("Codigo") or "").strip()
        if not code:
            continue
        price_map[code] = row
    return price_map


def header_row_index(ws) -> int | None:
    for row_idx in range(1, 8):
        values = [cell.value for cell in ws[row_idx]]
        if "Codigo" in values and "Valor Original" in values:
            return row_idx
    return None


def safe_float(value) -> float:
    if value in (None, ""):
        return 0.0
    return float(value)


def format_brl(value: float) -> str:
    return f"{value:,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")


def star_rating(value: float) -> str:
    rating = int(max(0, min(5, round(value))))
    return ("★" * rating) + ("☆" * (5 - rating))


def collect_sheet_records(ws) -> list[dict[str, object]]:
    row_idx = header_row_index(ws)
    if row_idx is None:
        return []

    headers = [cell.value for cell in ws[row_idx]]
    idx = {name: headers.index(name) for name in headers if name}
    records = []
    for row in ws.iter_rows(min_row=row_idx + 1, values_only=True):
        codigo = row[idx["Codigo"]] if idx.get("Codigo") is not None else None
        nome = row[idx["Nome"]] if idx.get("Nome") is not None else None
        if not codigo or not nome:
            continue
        records.append(
            {
                "codigo": str(codigo).strip(),
                "status": (row[idx["Status"]] or "").strip() if idx.get("Status") is not None and row[idx["Status"]] else "",
                "desgaste": safe_float(row[idx["Desgaste"]]) if idx.get("Desgaste") is not None else 0.0,
                "valor_original": safe_float(row[idx["Valor Original"]]) if idx.get("Valor Original") is not None else 0.0,
                "valor_atual": safe_float(row[idx["Valor Atual"]]) if idx.get("Valor Atual") is not None else 0.0,
            }
        )
    return records


def refresh_category_summary(ws, records: list[dict[str, object]]) -> None:
    if not records or ws.max_row < 2:
        return
    avg_wear = sum(record["desgaste"] for record in records) / len(records)
    total_value = sum(record["valor_original"] for record in records)
    ws["A2"] = f"{len(records)} itens   |   Valor: R$ {format_brl(total_value)}   |   Desgaste medio: {avg_wear:.1f}/5"


def refresh_dashboard(wb) -> None:
    if "DASHBOARD" not in wb.sheetnames:
        return

    dashboard = wb["DASHBOARD"]
    active_data = {sheet: collect_sheet_records(wb[sheet]) for sheet in ACTIVE_SHEETS if sheet in wb.sheetnames}

    for sheet_name, records in active_data.items():
        refresh_category_summary(wb[sheet_name], records)

    active_records = [record for records in active_data.values() for record in records]
    lotes = []
    if "LOTES" in wb.sheetnames:
        lotes_ws = wb["LOTES"]
        for row in lotes_ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                lotes.append(row[0])

    total = len(active_records)
    disponivel = sum(1 for record in active_records if record["status"] == "DISPONIVEL")
    em_campo = sum(1 for record in active_records if record["status"] == "EM_CAMPO")
    manutencao = sum(1 for record in active_records if record["status"] == "MANUTENCAO")
    total_original = sum(record["valor_original"] for record in active_records)
    total_atual = sum(record["valor_atual"] for record in active_records)
    avg_wear = (sum(record["desgaste"] for record in active_records) / total) if total else 0.0
    criticos = sum(1 for record in active_records if record["desgaste"] <= 2)
    deprec_pct = round((total_atual / total_original) * 100, 2) if total_original else 0.0

    dashboard["B3"] = f"Inventario Inteligente  ·  {datetime.now().strftime('%d/%m/%Y')}"
    dashboard["B5"] = total
    dashboard["D5"] = disponivel
    dashboard["F5"] = em_campo
    dashboard["B8"] = manutencao
    dashboard["D8"] = round(total_original, 2)
    dashboard["F8"] = round(total_atual, 2)
    dashboard["B12"] = star_rating(avg_wear)
    dashboard["C12"] = criticos
    dashboard["D12"] = f"{deprec_pct:.0f}%"
    dashboard["E12"] = len(lotes)

    row = 17
    for sheet_name in ACTIVE_SHEETS:
        records = active_data.get(sheet_name, [])
        category_total = sum(record["valor_original"] for record in records)
        dashboard.cell(row=row, column=2).value = sheet_name
        dashboard.cell(row=row, column=3).value = len(records)
        dashboard.cell(row=row, column=4).value = round(category_total, 2)
        dashboard.cell(row=row, column=5).value = (category_total / total_original) if total_original else 0
        row += 1

    while row <= 24:
        for col in range(2, 6):
            dashboard.cell(row=row, column=col).value = None
        row += 1


def main() -> int:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Planilha nao encontrada: {WORKBOOK_PATH}")
    if not MARKDOWN_PATH.exists():
        raise FileNotFoundError(f"Markdown nao encontrado: {MARKDOWN_PATH}")

    price_map = load_price_map(MARKDOWN_PATH)

    wb = load_workbook(WORKBOOK_PATH)
    applied = 0
    unmatched_markdown = set(price_map.keys())

    for ws in wb.worksheets:
        row_idx = header_row_index(ws)
        if row_idx is None:
            continue

        headers = [cell.value for cell in ws[row_idx]]
        idx = {name: headers.index(name) for name in headers if name}
        orig_col = idx["Valor Original"] + 1
        curr_col = idx["Valor Atual"] + 1
        dep_col = idx["Deprec.%"] + 1
        code_col = idx["Codigo"]

        for data_row in ws.iter_rows(min_row=row_idx + 1, values_only=False):
            code = data_row[code_col].value
            if not code:
                continue
            code = str(code).strip()
            payload = price_map.get(code)
            if not payload:
                continue

            try:
                valor_original = float(payload["Valor Original (R$)"])
                valor_atual = float(payload["Valor Atual (R$)"])
                dep_pct = float(payload["Deprec. (%)"])
            except (TypeError, ValueError) as exc:
                raise RuntimeError(f"Valores invalidos no markdown para {code}: {payload}") from exc

            ws.cell(row=data_row[0].row, column=orig_col).value = valor_original
            ws.cell(row=data_row[0].row, column=curr_col).value = valor_atual
            ws.cell(row=data_row[0].row, column=dep_col).value = dep_pct
            applied += 1
            unmatched_markdown.discard(code)

    backup_path = WORKBOOK_PATH.with_name(f"{WORKBOOK_PATH.stem}.backup-apply-{datetime.now().strftime('%Y%m%d-%H%M%S')}{WORKBOOK_PATH.suffix}")
    shutil.copy2(WORKBOOK_PATH, backup_path)
    refresh_dashboard(wb)
    wb.save(WORKBOOK_PATH)

    print(f"Backup criado em: {backup_path}")
    print(f"Linhas aplicadas: {applied}")
    print(f"Codigos no markdown sem correspondencia na planilha atual: {len(unmatched_markdown)}")
    if unmatched_markdown:
        for code in sorted(unmatched_markdown)[:25]:
            print(f" - {code}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
