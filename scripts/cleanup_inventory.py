from __future__ import annotations

import json
import re
import unicodedata
import uuid
from collections import Counter, defaultdict
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parents[1]
INPUT_WORKBOOK = ROOT_DIR / "data" / "inventario-original.xlsx"
ITEMS_OUTPUT = ROOT_DIR / "data" / "items.json"
SERIALS_OUTPUT = ROOT_DIR / "data" / "serial_numbers.json"
LOTES_OUTPUT = ROOT_DIR / "data" / "lotes.json"
MIGRATION_OUTPUT = ROOT_DIR / "data" / "migration.sql"
REPORT_OUTPUT = ROOT_DIR / "data" / "cleanup_report.txt"

UUID_NAMESPACE = uuid.UUID("4f83f8d2-4335-44d6-9080-75487cbf3c99")

CATEGORY_ORDER = [
    "ILUMINACAO",
    "AUDIO",
    "CABO",
    "ENERGIA",
    "ESTRUTURA",
    "EFEITO",
    "VIDEO",
    "ACESSORIO",
]

CATEGORY_PREFIX = {
    "ILUMINACAO": "ILU",
    "AUDIO": "AUD",
    "CABO": "CAB",
    "ENERGIA": "ENE",
    "ESTRUTURA": "EST",
    "EFEITO": "EFE",
    "VIDEO": "VID",
    "ACESSORIO": "ACE",
}

STATE_FACTOR = {
    "NOVO": 1.00,
    "SEMI_NOVO": 0.85,
    "USADO": 0.65,
    "RECONDICIONADO": 0.50,
}

DEFAULT_STATE = "USADO"
DEFAULT_WEAR = 3
DEFAULT_STATUS = "DISPONIVEL"
LOT_DEFAULT_STATUS = "DISPONIVEL"

LOT_TRACKED_RAW_CATEGORIES = {"CABO", "EXTENSOR"}

RAW_CATEGORY_TO_ENUM = {
    "PAR LEDS": "ILUMINACAO",
    "RIBALTA": "ILUMINACAO",
    "Ribalta": "ILUMINACAO",
    "Mini Moving": "ILUMINACAO",
    "MINI BRUTE": "ILUMINACAO",
    "Moving Beam": "ILUMINACAO",
    "OUTRAS LUZES": "ILUMINACAO",
    "LUZ": "ILUMINACAO",
    "Iluminação": "ILUMINACAO",
    "LASER": "ILUMINACAO",
    "LUZ NEGRA": "ILUMINACAO",
    "Audio": "AUDIO",
    "AUDIO": "AUDIO",
    "audio": "AUDIO",
    "Som": "AUDIO",
    "som": "AUDIO",
    "INSTRUMENTO": "AUDIO",
    "CAPSULAS MIC": "AUDIO",
    "CABO": "CABO",
    "EXTENSOR": "CABO",
    "REGUA": "ENERGIA",
    "Energia": "ENERGIA",
    "Eletrica": "ENERGIA",
    "ELETRICA": "ENERGIA",
    "FONTE": "ENERGIA",
    "TRIPÉ": "ESTRUTURA",
    "SUPORTE": "ESTRUTURA",
    "Estrutura": "ESTRUTURA",
    "Praticável": "ESTRUTURA",
    "EFEITOS": "EFEITO",
    "Efeitos": "EFEITO",
    "Fluido": "EFEITO",
    "Globo": "EFEITO",
    "67.0": "EFEITO",
    "COMPUTADORES": "VIDEO",
    "VIDEO": "VIDEO",
    "Notebook": "VIDEO",
    "Tablet": "VIDEO",
    "CASE": "ACESSORIO",
    "CASES": "ACESSORIO",
    "MIDI": "ACESSORIO",
    "Refrigeracao": "ACESSORIO",
    "Roteador": "ACESSORIO",
    "REDE": "ACESSORIO",
    "FERRAMENTAS": "ACESSORIO",
    "Radio": "ACESSORIO",
    "": "ACESSORIO",
}

EXTENSOR_SUBCATEGORY_TO_ENUM = {
    "EXTENSOR AC": "ENERGIA",
    "EXTENSAO 1 PONTA": "ENERGIA",
    "EXTENSOR P10": "CABO",
    "P10EXT - P10F": "CABO",
    "EXTENSOR P2": "CABO",
    "extensor hdmi": "CABO",
    "EXTENSOR USB A": "CABO",
    "": "CABO",
}

BRAND_MAP = {
    "QSC": "QSC",
    "SHOWTECH": "Showtech",
    "BRIWAX": "Briwax",
    "SENNHEISER": "Sennheiser",
    "BEHRINGER": "Behringer",
    "PIONEER": "Pioneer",
    "SHURE": "Shure",
    "JBL": "JBL",
    "YAMAHA": "Yamaha",
    "HOT MACHINE": "Hot Machine",
    "PHENYX PRO": "Phenyx Pro",
    "PHENIX PRO": "Phenyx Pro",
    "WIRECONNEX": "Wireconnex",
    "WIRECONEX": "Wireconnex",
    "STANER AUDIO AMERICA": "Staner",
    "RMV": "RMV",
    "BDL": "BDL",
    "LEACS": "Leacs",
    "SOG": "SOG",
    "TOMATE": "Tomate",
    "MEGA LIGHT": "Mega Light",
    "STAR LIGHT DIVISION": "Star Light Division",
    "STAR LIGHTING DIVISION": "Star Light Division",
    "SKYPIX": "Skypix",
    "AURA": "Aura",
    "CSR": "CSR",
    "MACKIE": "Mackie",
    "RCF": "RCF",
    "SOUNDBOX": "Soundbox",
    "MARK AUDIO": "Mark Audio",
    "NHL": "NHL",
    "HOWS": "Hows",
    "ROXON": "Roxon",
    "NEO": "Neo",
    "AMERICAN PRO": "American Pro",
    "FLOLUX": "Fiolux",
    "POLOCLIMA": "Poloclima",
    "DYLAM": "Dylan",
    "WATTSOM": "Wattsom",
    "ROLLAND": "Roland",
    "ROLAND": "Roland",
    "OK": "OK",
    "S/ MARCA": "",
    "SEM MARCA": "",
    "S/MARCA": "",
    "CHINES": "",
    "NONE": "",
}

SUBCATEGORY_ALIASES = {
    "MICROFONE SEM FIO": "Microfone sem fio",
    "MICROFONE WIRELESS": "Microfone sem fio",
    "MICROFONE COM FIO": "Microfone com fio",
    "MICROFONE": "Microfone",
    "MIFROFONE": "Microfone",
    "MIROFONE": "Microfone",
    "MIFROFONE SEM FIO": "Microfone sem fio",
    "MIROFONE SEM FIO": "Microfone sem fio",
    "CAPSULAS MIC": "Capsula microfone",
    "CAIXA DE SOM": "Caixa de som",
    "SUBWOOFER": "Subwoofer",
    "RECEPTOR MICROFONE": "Receptor microfone",
    "DIRECT BOX": "Direct box",
    "MESA DE SOM": "Mesa de som",
    "MESA DE LUZ": "Mesa de luz",
    "CONTROLADOR DMX": "Controlador DMX",
    "SPLITTER DMX": "Splitter DMX",
    "TRIPÉ DE MICROFONE": "Tripe microfone",
    "TRIPE DE MICROFONE": "Tripe microfone",
    "TRIPÉ DE CAIXA": "Tripe caixa",
    "TRIPE DE CAIXA": "Tripe caixa",
    "SUPORTE DE TV": "Suporte TV",
    "MÁQUINA DE FUMAÇA": "Maquina de fumaca",
    "MAQUINA DE FUMAÇA": "Maquina de fumaca",
    "MAQUINA DE FUMACA": "Maquina de fumaca",
    "BICO CO2": "CO2",
    "RIBALTA LASER": "Ribalta laser",
    "FIXA": "Fixa",
    "OUTDOOR": "Outdoor",
    "INDOOR": "Indoor",
    "5R": "Moving",
    "9R": "Moving",
    "COB": "COB",
    "RÉGUA": "Regua",
    "REGUA": "Regua",
    "P10EXT - P10F": "P10 extensor",
    "EXTENSAO 1 PONTA": "Extensao 1 ponta",
    "EXTENSOR AC": "Extensor AC",
    "EXTENSOR P10": "Extensor P10",
    "EXTENSOR P2": "Extensor P2",
    "EXTENSOR USB A": "Extensor USB",
    "EXTENSOR HDMI": "Extensor HDMI",
    "EFFECTS": "Effects",
}

MAIO_SECTION_MAP: dict[str, tuple[str | None, str | None]] = {
    "CAIXAS": ("AUDIO", "Caixa de som"),
    "SUBWOOFERS": ("AUDIO", "Subwoofer"),
    "MÁQUINA DE FUMAÇA": ("EFEITO", "Maquina de fumaca"),
    "TRANSFORMADOR 110-220V": ("ENERGIA", "Transformador"),
    "POTÊNCIA": ("AUDIO", "Amplificador"),
    "AMPLIFICADORES?": ("AUDIO", "Amplificador"),
    "EQUALIZADOR": ("AUDIO", "Equalizador"),
    "DIRECT BOX": ("AUDIO", "Direct box"),
    "CROSSOVER ?": ("AUDIO", "Processador"),
    "MICROFONES E AFINS": ("AUDIO", "Microfone"),
    "XDJS": ("AUDIO", "CDJ/XDJ"),
    "DDJ": ("AUDIO", "CDJ/XDJ"),
    "MESA DE SOM": ("AUDIO", "Mesa de som"),
    "PROCESSADOR": ("AUDIO", "Processador"),
    "MESA DE LUZ": ("ILUMINACAO", "Mesa de luz"),
    "NOTEBOOK": ("VIDEO", "Notebook"),
    "PROJETOR": ("VIDEO", "Projetor"),
    "ITENS SOLTOS": (None, None),
    "MOVING 9R": ("ILUMINACAO", "Moving"),
    "RIBALTAS BEAM": ("ILUMINACAO", "Ribalta"),
    "RIBALTAS LASER": ("ILUMINACAO", "Laser"),
    "MINI BRUTE": ("ILUMINACAO", "Mini Brute"),
    "OUTRAS LUZES": ("ILUMINACAO", "Iluminacao diversa"),
    "LUZ NEGRA": ("ILUMINACAO", "Luz Negra"),
    "LASER": ("ILUMINACAO", "Laser"),
    "BOX TRUSS": ("ESTRUTURA", "Box Truss"),
    "PAR LEDS": ("ILUMINACAO", "Par LED"),
}

KNOWN_MAIO_HEADERS = set(MAIO_SECTION_MAP) | {"Documentação/relação equipamentos MMD"}

MATCH_KEY_ALIASES = {
    "qsc k8 k10": ["qsc k10", "qsc k10 mkm cx2"],
    "k8 k10": ["qsc k10", "qsc k10 mkm cx2"],
    "sender box sb 8": ["senderbox"],
    "fishman wireless midi controller guitar": ["fishman wireless midi controller guitar"],
    "teil 2": ["teil 2"],
    "lp 354": ["lp354"],
    "p354": ["lp354", "p354"],
}

LOT_FAMILY_LABELS = {
    "AC": "Cabo AC",
    "DMX": "Cabo DMX",
    "XLR": "Cabo XLR",
    "XLR-P10": "Cabo XLR/P10",
    "P10-XLR": "Cabo P10/XLR",
    "P10": "Cabo P10",
    "P10-P2": "Cabo P10/P2",
    "P2": "Cabo P2",
    "RCA": "Cabo RCA",
    "RCA-P10": "Cabo RCA/P10",
    "SPEAK ON": "Cabo Speakon",
    "HDMI": "Cabo HDMI",
    "VIDEO": "Cabo video",
    "USB": "Cabo USB",
    "MIDI": "Cabo MIDI",
    "ENERGIA": "Fonte",
    "OUTRO": "Cabo diverso",
    "XLR-P2": "Cabo XLR/P2",
    "MINI DPPORT P DPPORT": "Cabo Mini DisplayPort",
    "XLRM P10": "Cabo XLR/P10",
}

LENGTH_BUCKET_LABELS = {
    "CURTO": "curto",
    "MEDIO": "medio",
    "LONGO": "longo",
    "SEM_MEDIDA": "sem medida",
}


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").replace("\r", " ").strip()
    return re.sub(r"\s+", " ", text)


def normalize_lookup_text(value: str) -> str:
    text = unicodedata.normalize("NFKD", clean_text(value))
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_brand(raw_brand: Any) -> str:
    brand = clean_text(raw_brand)
    if not brand:
        return ""
    key = brand.upper()
    if key in BRAND_MAP:
        return BRAND_MAP[key]
    words = brand.split()
    return " ".join(word.upper() if len(word) <= 3 else word.title() for word in words)


def normalize_subcategory(raw_subcategory: Any, fallback: str = "") -> str:
    value = clean_text(raw_subcategory) or clean_text(fallback)
    if not value:
        return ""
    key = value.upper()
    alias = SUBCATEGORY_ALIASES.get(key)
    return alias or value.title()


def parse_numeric(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    text = clean_text(value)
    if not text:
        return None
    normalized = text.upper().replace("R$", "").replace(".", "").replace(",", ".")
    if any(marker in normalized for marker in ("N ACHEI", "NAO ACHEI", "NONE", "S/ VALOR")):
        return None
    try:
        return round(float(normalized), 2)
    except ValueError:
        return None


def parse_quantity(value: Any) -> int:
    if value is None or clean_text(value) == "":
        return 1
    if isinstance(value, (int, float)):
        return max(1, int(value))
    match = re.search(r"\d+", clean_text(value))
    return max(1, int(match.group())) if match else 1


def parse_legacy_internal_code(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return clean_text(value)


def parse_serial_fabrica(raw_serial: Any) -> str:
    serial = clean_text(raw_serial)
    if not serial:
        return ""
    invalid_markers = ("S/ SERIAL", "SEM SERIAL", "SEM SERIAL APARENTE", "SEM ID", "NONE")
    serial_upper = serial.upper()
    if any(marker in serial_upper for marker in invalid_markers):
        return ""

    label_match = re.match(r"^(BODYPACK|BASE|NA CAIXA)\s*:\s*(.+)$", serial, flags=re.IGNORECASE)
    if label_match:
        serial = label_match.group(2)
        serial_upper = serial.upper()

    if " - " in serial:
        serial = serial.split(" - ", 1)[0]

    if "*" in serial:
        serial = serial.split("*", 1)[0]

    zero_prefix = re.match(r"^\((?:\d+\s+ZEROS?|ZERO(?: ZERO)?)\)\s*(.+)$", serial, flags=re.IGNORECASE)
    if zero_prefix:
        serial = zero_prefix.group(1)

    trailing_note = re.match(r"^(.+?)\s*\(([^)]+)\)$", serial)
    if trailing_note and any(token in trailing_note.group(2).upper() for token in ("CAIXA", "CODIGO", "RISCAD", "ETIQUETA")):
        serial = trailing_note.group(1)

    serial = serial.strip(" -:;")
    serial = re.sub(r"\s*/\s*", " / ", serial)
    serial = re.sub(r"\s+", " ", serial)
    if re.fullmatch(r"\d+\.0+", serial):
        serial = serial.split(".", 1)[0]
    return serial.upper()


def infer_status(ok_value: Any, model: str, notes: str) -> str:
    haystack = " ".join(filter(None, [clean_text(ok_value), model, notes])).upper()
    if "BIGODE" in haystack:
        return "EMPRESTADO"
    if "VENDID" in haystack:
        return "VENDIDO"
    if "QUEBRAD" in haystack:
        return "BAIXA"
    return DEFAULT_STATUS


def compute_condition(valor_mercado_unitario: float | None) -> tuple[str, int, float | None, float | None]:
    if valor_mercado_unitario is None:
        return DEFAULT_STATE, DEFAULT_WEAR, None, None
    fator = STATE_FACTOR[DEFAULT_STATE]
    valor_atual = round(valor_mercado_unitario * (DEFAULT_WEAR / 5) * fator, 2)
    depreciacao = round((DEFAULT_WEAR / 5) * fator * 100, 2)
    return DEFAULT_STATE, DEFAULT_WEAR, depreciacao, valor_atual


def build_display_name(brand: str, model: str, subcategory: str, category: str) -> str:
    if model:
        if brand and normalize_lookup_text(model).startswith(normalize_lookup_text(brand)):
            return model
        if brand:
            return f"{brand} {model}".strip()
        return model
    if brand and subcategory:
        return f"{subcategory} {brand}".strip()
    return subcategory or category.title()


def infer_category_from_text(raw_category: str, raw_subcategory: str, brand: str, model: str, notes: str) -> str:
    lookup = normalize_lookup_text(" ".join(filter(None, [raw_category, raw_subcategory, brand, model, notes])))
    if any(token in lookup for token in ("microfone", "receiver", "bodypack", "pedaleira", "mixer", "caixa", "subwoofer", "line array", "amplifier", "amplificador")):
        return "AUDIO"
    if any(token in lookup for token in ("roteador", "router", "senderbox", "wireless midi", "kaoss", "traktor", "cable tester", "midi")):
        return "ACESSORIO"
    if any(token in lookup for token in ("projetor", "notebook", "macbook", "computador", "tablet")):
        return "VIDEO"
    if any(token in lookup for token in ("fumaca", "smoke", "globo", "co2", "effects")):
        return "EFEITO"
    if any(token in lookup for token in ("box truss", "truss", "praticavel", "suporte", "tripe")):
        return "ESTRUTURA"
    if any(token in lookup for token in ("dmx", "laser", "par led", "moving", "ribalta", "pilot 2000", "command wing", "lumikit")):
        return "ILUMINACAO"
    if any(token in lookup for token in ("transformador", "regua", "filtro de linha", "energia")):
        return "ENERGIA"
    return "ACESSORIO"


def resolve_category(raw_category: str, raw_subcategory: str, brand: str, model: str, notes: str, warnings: list[str], source_ref: str) -> str:
    base_category = RAW_CATEGORY_TO_ENUM.get(raw_category)
    if base_category is None:
        inferred = infer_category_from_text(raw_category, raw_subcategory, brand, model, notes)
        warnings.append(f"[{source_ref}] Categoria bruta sem mapeamento direto: {raw_category!r}. Inferida como {inferred}.")
        return inferred

    if raw_category == "EXTENSOR":
        override = EXTENSOR_SUBCATEGORY_TO_ENUM.get(raw_subcategory, "CABO")
        warnings.append(f"[{source_ref}] Categoria ambigua EXTENSOR resolvida como {override} com base na subcategoria {raw_subcategory!r}.")
        return override

    if raw_category in {"", "67.0"}:
        inferred = infer_category_from_text(raw_category, raw_subcategory, brand, model, notes)
        warnings.append(f"[{source_ref}] Categoria ambigua {raw_category!r} resolvida como {inferred} com base no contexto da linha.")
        return inferred

    return base_category


def is_lot_tracked(raw_category: str) -> bool:
    return raw_category in LOT_TRACKED_RAW_CATEGORIES


def normalize_item_key(name: str, brand: str, model: str) -> tuple[str, str, str]:
    return (
        normalize_lookup_text(name),
        normalize_lookup_text(brand),
        normalize_lookup_text(model),
    )


def make_uuid(*parts: str) -> str:
    return str(uuid.uuid5(UUID_NAMESPACE, "::".join(parts)))


def assign_legacy_internal_uniques(rows: list[dict[str, Any]], duplicate_log: list[str]) -> None:
    occurrences: defaultdict[str, int] = defaultdict(int)
    first_row_by_code: dict[str, int] = {}
    for row in rows:
        legacy_code = row["legacy_internal_code"]
        if not legacy_code:
            row["legacy_internal_unique"] = f"ROW-{row['source_row']}"
            continue
        occurrences[legacy_code] += 1
        suffix = occurrences[legacy_code]
        row["legacy_internal_unique"] = legacy_code if suffix == 1 else f"{legacy_code}-{suffix}"
        if suffix == 1:
            first_row_by_code[legacy_code] = row["source_row"]
        else:
            duplicate_log.append(
                f"Codigo legado duplicado {legacy_code}: linha {first_row_by_code[legacy_code]} e linha {row['source_row']} -> renomeado para {row['legacy_internal_unique']}."
            )


def resolve_maio_category(section: str, item_name: str, warnings: list[str], row_idx: int) -> tuple[str, str]:
    mapped = MAIO_SECTION_MAP.get(section)
    if mapped and mapped[0]:
        return mapped[0], mapped[1] or ""

    inferred = infer_category_from_text(section, "", "", item_name, "")
    if inferred == "ILUMINACAO":
        subcategory = "Par LED" if "PAR" in item_name.upper() else "Iluminacao diversa"
    elif inferred == "AUDIO":
        subcategory = "Audio"
    elif inferred == "VIDEO":
        subcategory = "Video"
    elif inferred == "ENERGIA":
        subcategory = "Energia"
    else:
        subcategory = "Acessorio"
    warnings.append(f"[EQUIPAMENTOS - MAIO:{row_idx}] Secao {section!r} inferida como {inferred} para {item_name!r}.")
    return inferred, subcategory


def parse_old_rows(warnings: list[str], duplicate_log: list[str]) -> tuple[list[dict[str, Any]], set[str]]:
    workbook = load_workbook(INPUT_WORKBOOK, data_only=True)
    worksheet = workbook["EQUIPAMENTOS"]
    rows: list[dict[str, Any]] = []
    raw_categories: set[str] = set()

    for row_idx, row in enumerate(worksheet.iter_rows(min_row=3, values_only=True), start=3):
        meaningful_fields = [row[2], row[3], row[4], row[5], row[9], row[10], row[13]]
        if not any(value not in (None, "") for value in meaningful_fields):
            continue

        raw_category = clean_text(row[2])
        raw_subcategory = clean_text(row[3])
        brand = normalize_brand(row[4])
        model = clean_text(row[5])
        notes = clean_text(row[13])
        final_category = resolve_category(raw_category, raw_subcategory, brand, model, notes, warnings, f"EQUIPAMENTOS:{row_idx}")
        subcategory = normalize_subcategory(raw_subcategory, raw_category or final_category)
        status = infer_status(row[1], model, notes)
        serial_fabrica = parse_serial_fabrica(row[9])
        valor_mercado = parse_numeric(row[10])
        raw_categories.add(raw_category)

        row_payload = {
            "source_sheet": "EQUIPAMENTOS",
            "source_row": row_idx,
            "legacy_internal_code": parse_legacy_internal_code(row[0]),
            "raw_category": raw_category,
            "raw_subcategory": raw_subcategory,
            "categoria": final_category,
            "subcategoria": subcategory,
            "marca": brand,
            "modelo": model,
            "nome": build_display_name(brand, model, subcategory, final_category),
            "serial_fabrica": serial_fabrica,
            "valor_mercado_unitario": valor_mercado,
            "status": status,
            "notas": notes,
            "tracking_type": "LOTE" if is_lot_tracked(raw_category) else "INDIVIDUAL",
        }
        rows.append(row_payload)

    assign_legacy_internal_uniques(rows, duplicate_log)
    return rows, raw_categories


def is_maio_header_row(values: tuple[Any, ...], current_section: str | None) -> str | None:
    first = clean_text(values[0] if len(values) > 0 else None)
    second = values[1] if len(values) > 1 else None
    third = values[2] if len(values) > 2 else None
    fourth = values[3] if len(values) > 3 else None

    if not first:
        return None
    if first in {"QTD", "numero serie / id"}:
        return None
    if first in KNOWN_MAIO_HEADERS and all(value in (None, "") for value in values[1:7]):
        return first
    if first in KNOWN_MAIO_HEADERS and isinstance(second, str) and not clean_text(third) and fourth is None:
        return first
    if first == "CAIXAS" and clean_text(second).upper() == "QTD":
        return "CAIXAS"
    if current_section == "ITENS SOLTOS" and first == "PAR LEDS" and isinstance(second, str):
        return "PAR LEDS"
    return None


def parse_maio_groups(warnings: list[str]) -> dict[str, dict[str, Any]]:
    workbook = load_workbook(INPUT_WORKBOOK, data_only=True)
    worksheet = workbook["EQUIPAMENTOS - MAIO"]

    records: list[dict[str, Any]] = []
    current_section = "CAIXAS"

    for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        header = is_maio_header_row(row, current_section)
        if header:
            current_section = header
            continue

        item_name = clean_text(row[0] if len(row) > 0 else None)
        if not item_name:
            continue
        if item_name.upper() == "CAIXAS" and clean_text(row[1] if len(row) > 1 else None).upper() == "QTD":
            current_section = "CAIXAS"
            continue

        qty = parse_quantity(row[1] if len(row) > 1 else None)
        serial = parse_serial_fabrica(row[2] if len(row) > 2 else None)
        valor = parse_numeric(row[3] if len(row) > 3 else None)
        site_ref = clean_text(row[4] if len(row) > 4 else None)
        note = clean_text(row[6] if len(row) > 6 else None)
        categoria, subcategoria = resolve_maio_category(current_section, item_name, warnings, row_idx)
        lookup_key = normalize_lookup_text(item_name)

        records.append(
            {
                "source_row": row_idx,
                "section": current_section,
                "item_name": item_name,
                "lookup_key": lookup_key,
                "categoria": categoria,
                "subcategoria": subcategoria,
                "quantidade": qty,
                "serial_fabrica": serial,
                "valor_mercado_unitario": valor,
                "site_ref": site_ref,
                "notas": note,
            }
        )

    groups: dict[str, dict[str, Any]] = {}
    for record in records:
        lookup_key = record["lookup_key"]
        group = groups.setdefault(
            lookup_key,
            {
                "lookup_key": lookup_key,
                "item_name": record["item_name"],
                "categoria": record["categoria"],
                "subcategoria": record["subcategoria"],
                "quantidade_total": 0,
                "serials": [],
                "values": [],
                "notes": [],
                "site_refs": [],
                "source_rows": [],
            },
        )
        group["quantidade_total"] += record["quantidade"]
        if record["serial_fabrica"]:
            group["serials"].append(record["serial_fabrica"])
        if record["valor_mercado_unitario"] is not None:
            group["values"].append(record["valor_mercado_unitario"])
        if record["notas"]:
            group["notes"].append(record["notas"])
        if record["site_ref"]:
            group["site_refs"].append(record["site_ref"])
        group["source_rows"].append(record["source_row"])

    for group in groups.values():
        group["serials"] = dedupe_keep_order(group["serials"])
        group["notes"] = dedupe_keep_order(group["notes"])
        group["site_refs"] = dedupe_keep_order(group["site_refs"])
        group["valor_preferencial"] = choose_numeric_value(group["values"])

    return groups


def dedupe_keep_order(values: list[str]) -> list[str]:
    seen: set[str] = set()
    output: list[str] = []
    for value in values:
        if value in seen:
            continue
        seen.add(value)
        output.append(value)
    return output


def choose_numeric_value(values: list[float]) -> float | None:
    if not values:
        return None
    normalized = [round(value, 2) for value in values]
    most_common = Counter(normalized).most_common(1)[0][0]
    return round(float(most_common), 2)


def build_group_match_keys(group: dict[str, Any]) -> list[str]:
    keys = [
        normalize_lookup_text(group["nome"]),
        normalize_lookup_text(group["modelo"]),
        normalize_lookup_text(f"{group['marca']} {group['modelo']}"),
        normalize_lookup_text(f"{group['marca']} {group['nome']}"),
    ]
    aliases = MATCH_KEY_ALIASES.get(keys[0], []) + MATCH_KEY_ALIASES.get(keys[1], [])
    keys.extend(normalize_lookup_text(alias) for alias in aliases)
    return [key for key in dedupe_keep_order([key for key in keys if key])]


def categories_are_compatible(item_category: str, maio_category: str) -> bool:
    if item_category == maio_category:
        return True
    compatible_sets = [
        {"ACESSORIO", "AUDIO"},
        {"ACESSORIO", "ILUMINACAO"},
        {"CABO", "ENERGIA"},
    ]
    return any({item_category, maio_category}.issubset(group) for group in compatible_sets)


def digit_token_compatible(left: str, right: str) -> bool:
    left_tokens = [token for token in re.findall(r"[a-z0-9]+", left) if any(char.isdigit() for char in token)]
    right_tokens = [token for token in re.findall(r"[a-z0-9]+", right) if any(char.isdigit() for char in token)]
    if not left_tokens or not right_tokens:
        return True
    for left_token in left_tokens:
        for right_token in right_tokens:
            if left_token == right_token or left_token in right_token or right_token in left_token:
                return True
    return False


def match_maio_groups(item_groups: list[dict[str, Any]], maio_groups: dict[str, dict[str, Any]], warnings: list[str]) -> dict[str, int]:
    stats = {"matched_exact": 0, "matched_fuzzy": 0, "unmatched": 0}
    by_category: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
    for maio_group in maio_groups.values():
        by_category[maio_group["categoria"]].append(maio_group)

    for item_group in item_groups:
        candidate_keys = build_group_match_keys(item_group)
        matched_group = None
        match_mode = None

        for key in candidate_keys:
            candidate = maio_groups.get(key)
            if candidate and categories_are_compatible(item_group["categoria"], candidate["categoria"]):
                matched_group = candidate
                match_mode = "exact"
                break

        if matched_group is None:
            candidates = [
                candidate
                for category, groups in by_category.items()
                if categories_are_compatible(item_group["categoria"], category)
                for candidate in groups
            ]
            best_score = 0.0
            best_candidate = None
            for candidate in candidates:
                if not any(digit_token_compatible(key, candidate["lookup_key"]) for key in candidate_keys):
                    continue
                score = max(
                    SequenceMatcher(None, key, candidate["lookup_key"]).ratio()
                    for key in candidate_keys
                )
                if score > best_score:
                    best_score = score
                    best_candidate = candidate
            if best_candidate and best_score >= 0.86:
                matched_group = best_candidate
                match_mode = "fuzzy"
                warnings.append(
                    f"[MATCH FUZZY] {item_group['nome']!r} -> {best_candidate['item_name']!r} (score {best_score:.2f})."
                )

        if matched_group is None:
            item_group["maio_match"] = None
            stats["unmatched"] += 1
            continue

        item_group["maio_match"] = matched_group
        item_group["maio_match_mode"] = match_mode
        if match_mode == "exact":
            stats["matched_exact"] += 1
        else:
            stats["matched_fuzzy"] += 1

    return stats


def apply_maio_enrichment(item_groups: list[dict[str, Any]]) -> dict[str, int]:
    stats = {
        "value_overrides": 0,
        "value_backfills": 0,
        "serial_backfills": 0,
        "serial_unmatched_slots": 0,
        "unused_maio_serials": 0,
    }

    for item_group in item_groups:
        match = item_group.get("maio_match")
        if not match:
            continue

        preferred_value = match.get("valor_preferencial")
        if preferred_value is not None:
            for row in item_group["rows"]:
                if row["valor_mercado_unitario"] is None:
                    stats["value_backfills"] += 1
                elif row["valor_mercado_unitario"] != preferred_value:
                    stats["value_overrides"] += 1
                row["valor_mercado_unitario"] = preferred_value

        used_serials = {row["serial_fabrica"] for row in item_group["rows"] if row["serial_fabrica"]}
        available_serials = [serial for serial in match["serials"] if serial not in used_serials]
        for row in item_group["rows"]:
            if row["serial_fabrica"]:
                continue
            if available_serials:
                row["serial_fabrica"] = available_serials.pop(0)
                stats["serial_backfills"] += 1
            else:
                stats["serial_unmatched_slots"] += 1
        stats["unused_maio_serials"] += len(available_serials)

    return stats


def build_individual_item_groups(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str, str], dict[str, Any]] = {}
    for row in rows:
        if row["tracking_type"] != "INDIVIDUAL":
            continue
        key = normalize_item_key(row["nome"], row["marca"], row["modelo"])
        group = grouped.setdefault(
            key,
            {
                "key": key,
                "nome": row["nome"],
                "categoria": row["categoria"],
                "marca": row["marca"],
                "modelo": row["modelo"],
                "rows": [],
            },
        )
        group["rows"].append(row)
    return sorted(
        grouped.values(),
        key=lambda group: (
            CATEGORY_ORDER.index(group["categoria"]),
            normalize_lookup_text(group["nome"]),
            normalize_lookup_text(group["marca"]),
            normalize_lookup_text(group["modelo"]),
        ),
    )


def pick_mode(values: list[str]) -> str:
    filtered = [value for value in values if value]
    if not filtered:
        return ""
    return Counter(filtered).most_common(1)[0][0]


def build_notes_from_group(group: dict[str, Any]) -> str:
    notes: list[str] = []
    statuses = Counter(row["status"] for row in group["rows"])
    if len(statuses) > 1:
        notes.append(
            "Status mistos no grupo: "
            + ", ".join(f"{status}={count}" for status, count in sorted(statuses.items()))
        )
    match = group.get("maio_match")
    if match:
        match_mode = group.get("maio_match_mode", "exact")
        notes.append(
            f"Enriquecido pela aba EQUIPAMENTOS - MAIO ({match_mode}) nas linhas {', '.join(str(value) for value in match['source_rows'])}."
        )
        if match["site_refs"]:
            notes.append("Refs MAIO: " + ", ".join(match["site_refs"]))
    return " ".join(notes)


def build_serial_items_and_records(item_groups: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    items: list[dict[str, Any]] = []
    serials: list[dict[str, Any]] = []

    for group in item_groups:
        item_id = make_uuid("item", group["categoria"], group["nome"], group["marca"], group["modelo"], "INDIVIDUAL")
        subcategoria = pick_mode([row["subcategoria"] for row in group["rows"]])
        valor_mercado = choose_numeric_value([row["valor_mercado_unitario"] for row in group["rows"] if row["valor_mercado_unitario"] is not None])
        item_notes = build_notes_from_group(group)

        items.append(
            {
                "id": item_id,
                "nome": group["nome"],
                "categoria": group["categoria"],
                "subcategoria": subcategoria or None,
                "marca": group["marca"] or None,
                "modelo": group["modelo"] or None,
                "tipo_rastreamento": "INDIVIDUAL",
                "quantidade_total": len(group["rows"]),
                "valor_mercado_unitario": valor_mercado,
                "foto_url": None,
                "notas": item_notes or None,
                "source_kind": "individual",
            }
        )

        for row in sorted(group["rows"], key=lambda value: value["source_row"]):
            estado, desgaste, depreciacao_pct, valor_atual = compute_condition(row["valor_mercado_unitario"])
            serial_id = make_uuid("serial", item_id, row["legacy_internal_unique"])
            serial_notes = []
            if row["legacy_internal_code"]:
                serial_notes.append(f"Serial interno legado: {row['legacy_internal_unique']}")
            if row["notas"]:
                serial_notes.append(row["notas"])

            serials.append(
                {
                    "id": serial_id,
                    "item_id": item_id,
                    "codigo_interno": None,
                    "serial_fabrica": row["serial_fabrica"] or None,
                    "tag_rfid": None,
                    "qr_code": None,
                    "status": row["status"],
                    "estado": estado,
                    "desgaste": desgaste,
                    "depreciacao_pct": depreciacao_pct,
                    "valor_atual": valor_atual,
                    "localizacao": None,
                    "notas": " | ".join(serial_notes) if serial_notes else None,
                    "categoria": group["categoria"],
                    "item_nome": group["nome"],
                    "source_row": row["source_row"],
                    "legacy_internal_unique": row["legacy_internal_unique"],
                }
            )

    return items, serials


def lot_family_from_row(row: dict[str, Any]) -> str:
    raw_category = row["raw_category"]
    raw_subcategory = row["raw_subcategory"].upper()
    modelo = row["modelo"].upper()

    if raw_category == "EXTENSOR":
        if "AC" in raw_subcategory or "1 PONTA" in raw_subcategory:
            return "AC"
        if "P10" in raw_subcategory:
            return "P10"
        if "P2" in raw_subcategory:
            return "P2"
        if "HDMI" in raw_subcategory:
            return "HDMI"
        if "USB" in raw_subcategory:
            return "USB"
        return "OUTRO"

    mapping = {
        "POWERCON PIAL": "AC",
        "POWERCON POWERCON": "AC",
        "POWERCON AC": "AC",
        "DMX PIAL": "DMX",
        "XLR P10": "XLR-P10",
        "P10 XLRM": "P10-XLR",
        "P10 XLRF": "P10-XLR",
        "RCA P10": "RCA-P10",
        "RCA RCA": "RCA",
        "P10 P2": "P10-P2",
        "USB B USB A": "USB",
        "USB B USB C": "USB",
        "USB A USB C": "USB",
        "USB A USB B": "USB",
        "USB A USB A": "USB",
        "EXTENSOR USB": "USB",
        "USB USB": "USB",
        "XLRM USB C": "USB",
        "XLR USB C": "USB",
        "DISPLAY PORT": "VIDEO",
        "MINI DISPLAYPORT": "VIDEO",
        "IPHONE 30 PIN": "USB",
        "MIDI DIN": "MIDI",
        "AC 4X": "AC",
        "PARALELO": "AC",
        "FONTE": "ENERGIA",
        "VGA": "VIDEO",
        "DVI": "VIDEO",
        "USB A MICRO USB B": "USB",
        "XLRM RCA": "RCA",
        "RCA P2": "RCA",
        "XLR P2": "XLR-P2",
        "P2 P2": "P2",
        "P2 USB A": "USB",
        "P10 P10": "P10",
        "HDMI DISPLAY PORT": "VIDEO",
        "DVI HDMI": "VIDEO",
        "ADAPTADOR DIVISOR P10": "P10",
        "AC 8": "AC",
    }
    if raw_subcategory in mapping:
        return mapping[raw_subcategory]
    if "HDMI" in raw_subcategory or "HDMI" in modelo:
        return "HDMI"
    if "USB" in raw_subcategory or "USB" in modelo:
        return "USB"
    return raw_subcategory or "OUTRO"


def lot_length_bucket(row: dict[str, Any]) -> str:
    text = clean_text(row["modelo"] or row["raw_subcategory"]).upper().replace(",", ".")
    numeric_lengths = [float(value) for value in re.findall(r"(\d+(?:\.\d+)?)\s*(?:M|METRO)", text)]

    if any(token in text for token in ("50CM", "20CM", "35CM", "MEIO METRO")) or "0.5" in text:
        value = 0.5
    elif numeric_lengths:
        value = max(numeric_lengths)
    elif any(token in text for token in ("70CM", "85CM", "90CM")):
        value = 1.0
    else:
        value = None

    if value is None:
        if "LONG" in text or "COMPRID" in text:
            return "LONGO"
        if "CURT" in text or "BEM CURT" in text:
            return "CURTO"
        if "MEDI" in text:
            return "MEDIO"
        return "SEM_MEDIDA"

    if value <= 1.2:
        return "CURTO"
    if value <= 3.5:
        return "MEDIO"
    return "LONGO"


def build_lot_items_and_lotes(rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    lot_rows = [row for row in rows if row["tracking_type"] == "LOTE"]
    item_groups: dict[tuple[str, str], dict[str, Any]] = {}
    lote_groups: dict[tuple[str, str, str, str], list[dict[str, Any]]] = defaultdict(list)

    for row in lot_rows:
        family = lot_family_from_row(row)
        length_bucket = lot_length_bucket(row)
        item_category = row["categoria"]
        item_key = (item_category, family)
        item_group = item_groups.setdefault(
            item_key,
            {
                "categoria": item_category,
                "family": family,
                "rows": [],
            },
        )
        item_group["rows"].append(row)
        lote_groups[(item_category, family, length_bucket, row["status"])].append(row)

    items: list[dict[str, Any]] = []
    item_ids: dict[tuple[str, str], str] = {}
    for item_key, item_group in sorted(
        item_groups.items(),
        key=lambda entry: (CATEGORY_ORDER.index(entry[0][0]), entry[0][1]),
    ):
        category, family = item_key
        item_id = make_uuid("item", category, family, "LOTE")
        item_ids[item_key] = item_id
        label = LOT_FAMILY_LABELS.get(family, f"Cabo {family.title()}")
        notes = f"Gerado por agrupamento de cabos/extensores do inventario original. Familia {family}."
        items.append(
            {
                "id": item_id,
                "nome": label,
                "categoria": category,
                "subcategoria": family,
                "marca": None,
                "modelo": None,
                "tipo_rastreamento": "LOTE",
                "quantidade_total": len(item_group["rows"]),
                "valor_mercado_unitario": None,
                "foto_url": None,
                "notas": notes,
                "source_kind": "lote",
            }
        )

    lotes: list[dict[str, Any]] = []
    for group_key, group_rows in sorted(
        lote_groups.items(),
        key=lambda entry: (
            CATEGORY_ORDER.index(entry[0][0]),
            entry[0][1],
            ["CURTO", "MEDIO", "LONGO", "SEM_MEDIDA"].index(entry[0][2]),
            entry[0][3],
        ),
    ):
        category, family, length_bucket, status = group_key
        item_id = item_ids[(category, family)]
        family_label = LOT_FAMILY_LABELS.get(family, family.title())
        length_label = LENGTH_BUCKET_LABELS.get(length_bucket, length_bucket.lower())
        sample_models = Counter(clean_text(row["modelo"]) for row in group_rows if clean_text(row["modelo"]))
        sample = ", ".join(model for model, _ in sample_models.most_common(3)) or "Sem detalhe"

        lotes.append(
            {
                "id": make_uuid("lote", item_id, family, length_bucket, status),
                "item_id": item_id,
                "codigo_lote": None,
                "descricao": f"{family_label} {length_label}",
                "quantidade": len(group_rows),
                "tag_rfid": None,
                "qr_code": None,
                "status": LOT_DEFAULT_STATUS if status == DEFAULT_STATUS else status,
                "notas": f"Amostra: {sample}",
                "categoria": category,
                "family": family,
                "length_bucket": length_bucket,
            }
        )

    return items, lotes


def assign_serial_codes(serials: list[dict[str, Any]]) -> None:
    counters: defaultdict[str, int] = defaultdict(int)
    serials.sort(
        key=lambda serial: (
            CATEGORY_ORDER.index(serial["categoria"]),
            normalize_lookup_text(serial["item_nome"]),
            serial["source_row"],
        )
    )
    for serial in serials:
        counters[serial["categoria"]] += 1
        prefix = CATEGORY_PREFIX[serial["categoria"]]
        serial["codigo_interno"] = f"MMD-{prefix}-{counters[serial['categoria']]:04d}"


def assign_lot_codes(lotes: list[dict[str, Any]]) -> None:
    for index, lote in enumerate(lotes, start=1):
        lote["codigo_lote"] = f"MMD-LOT-{index:03d}"


def serialize_json(records: list[dict[str, Any]], path: Path) -> None:
    path.write_text(json.dumps(records, indent=2, ensure_ascii=False), encoding="utf-8")


def build_sql_literal(value: Any) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float)):
        if isinstance(value, float):
            return f"{value:.2f}"
        return str(value)
    text = str(value).replace("'", "''")
    return f"'{text}'"


def build_enum_sql(name: str, values: list[str]) -> str:
    formatted_values = ", ".join(build_sql_literal(value) for value in values)
    return (
        f"DO $$ BEGIN\n"
        f"    CREATE TYPE {name} AS ENUM ({formatted_values});\n"
        f"EXCEPTION\n"
        f"    WHEN duplicate_object THEN NULL;\n"
        f"END $$;"
    )


def build_insert_sql(table_name: str, columns: list[str], rows: list[dict[str, Any]]) -> str:
    values_sql = []
    for row in rows:
        values_sql.append(
            "(" + ", ".join(build_sql_literal(row.get(column)) for column in columns) + ")"
        )
    joined_values = ",\n  ".join(values_sql)
    return f"INSERT INTO {table_name} ({', '.join(columns)}) VALUES\n  {joined_values}\nON CONFLICT DO NOTHING;"


def build_migration_sql(items: list[dict[str, Any]], serials: list[dict[str, Any]], lotes: list[dict[str, Any]]) -> str:
    sql_parts = [
        "-- Generated by scripts/cleanup_inventory.py",
        f"-- Source workbook: {INPUT_WORKBOOK.name}",
        "CREATE EXTENSION IF NOT EXISTS pgcrypto;",
        "",
        build_enum_sql("categoria_enum", CATEGORY_ORDER),
        build_enum_sql("tipo_rastreamento_enum", ["INDIVIDUAL", "LOTE", "BULK"]),
        build_enum_sql("status_serial_enum", ["DISPONIVEL", "PACKED", "EM_CAMPO", "RETORNANDO", "MANUTENCAO", "EMPRESTADO", "VENDIDO", "BAIXA"]),
        build_enum_sql("estado_enum", ["NOVO", "SEMI_NOVO", "USADO", "RECONDICIONADO"]),
        build_enum_sql("status_projeto_enum", ["PLANEJAMENTO", "CONFIRMADO", "EM_CAMPO", "FINALIZADO", "CANCELADO"]),
        build_enum_sql("tipo_movimentacao_enum", ["SAIDA", "RETORNO", "MANUTENCAO", "TRANSFERENCIA", "DANO"]),
        build_enum_sql("metodo_scan_enum", ["RFID", "QRCODE", "MANUAL"]),
        build_enum_sql("status_lote_enum", ["DISPONIVEL", "EM_CAMPO", "MANUTENCAO"]),
        "",
        """CREATE OR REPLACE FUNCTION set_updated_at()
RETURNS TRIGGER AS $$
BEGIN
    NEW.updated_at = now();
    RETURN NEW;
END;
$$ LANGUAGE plpgsql;""",
        "",
        """CREATE TABLE IF NOT EXISTS items (
  id uuid DEFAULT gen_random_uuid() PRIMARY KEY,
  nome text NOT NULL,
  categoria categoria_enum NOT NULL,
  subcategoria text,
  marca text,
  modelo text,
  tipo_rastreamento tipo_rastreamento_enum NOT NULL DEFAULT 'INDIVIDUAL',
  quantidade_total int NOT NULL DEFAULT 1,
  valor_mercado_unitario numeric(10,2),
  foto_url text,
  notas text,
  created_at timestamptz DEFAULT now(),
  updated_at timestamptz DEFAULT now()
);""",
        """CREATE TABLE IF NOT EXISTS serial_numbers (
  id uuid DEFAULT gen_random_uuid() PRIMARY KEY,
  item_id uuid NOT NULL REFERENCES items(id) ON DELETE CASCADE,
  codigo_interno text UNIQUE NOT NULL,
  serial_fabrica text,
  tag_rfid text UNIQUE,
  qr_code text UNIQUE,
  status status_serial_enum NOT NULL DEFAULT 'DISPONIVEL',
  estado estado_enum NOT NULL DEFAULT 'USADO',
  desgaste int NOT NULL DEFAULT 3 CHECK (desgaste BETWEEN 1 AND 5),
  depreciacao_pct numeric(5,2),
  valor_atual numeric(10,2),
  localizacao text,
  notas text,
  created_at timestamptz DEFAULT now(),
  updated_at timestamptz DEFAULT now()
);""",
        """CREATE TABLE IF NOT EXISTS projetos (
  id uuid DEFAULT gen_random_uuid() PRIMARY KEY,
  nome text NOT NULL,
  cliente text,
  data_inicio date,
  data_fim date,
  local text,
  status status_projeto_enum NOT NULL DEFAULT 'PLANEJAMENTO',
  notas text,
  created_at timestamptz DEFAULT now(),
  updated_at timestamptz DEFAULT now()
);""",
        """CREATE TABLE IF NOT EXISTS packing_list (
  id uuid DEFAULT gen_random_uuid() PRIMARY KEY,
  projeto_id uuid NOT NULL REFERENCES projetos(id) ON DELETE CASCADE,
  item_id uuid NOT NULL REFERENCES items(id) ON DELETE CASCADE,
  quantidade int NOT NULL DEFAULT 1,
  serial_numbers_designados uuid[],
  notas text
);""",
        """CREATE TABLE IF NOT EXISTS movimentacoes (
  id uuid DEFAULT gen_random_uuid() PRIMARY KEY,
  serial_number_id uuid NOT NULL REFERENCES serial_numbers(id) ON DELETE CASCADE,
  projeto_id uuid REFERENCES projetos(id) ON DELETE SET NULL,
  tipo tipo_movimentacao_enum NOT NULL,
  status_anterior text,
  status_novo text,
  registrado_por text,
  metodo_scan metodo_scan_enum,
  timestamp timestamptz DEFAULT now(),
  notas text
);""",
        """CREATE TABLE IF NOT EXISTS lotes (
  id uuid DEFAULT gen_random_uuid() PRIMARY KEY,
  item_id uuid NOT NULL REFERENCES items(id) ON DELETE CASCADE,
  codigo_lote text UNIQUE NOT NULL,
  descricao text,
  quantidade int NOT NULL DEFAULT 1,
  tag_rfid text UNIQUE,
  qr_code text UNIQUE,
  status status_lote_enum NOT NULL DEFAULT 'DISPONIVEL',
  created_at timestamptz DEFAULT now(),
  updated_at timestamptz DEFAULT now()
);""",
        "",
        "CREATE INDEX IF NOT EXISTS idx_items_categoria ON items(categoria);",
        "CREATE INDEX IF NOT EXISTS idx_items_nome ON items(nome);",
        "CREATE INDEX IF NOT EXISTS idx_serial_numbers_codigo_interno ON serial_numbers(codigo_interno);",
        "CREATE INDEX IF NOT EXISTS idx_serial_numbers_tag_rfid ON serial_numbers(tag_rfid);",
        "CREATE INDEX IF NOT EXISTS idx_serial_numbers_qr_code ON serial_numbers(qr_code);",
        "CREATE INDEX IF NOT EXISTS idx_serial_numbers_status ON serial_numbers(status);",
        "CREATE INDEX IF NOT EXISTS idx_serial_numbers_item_id ON serial_numbers(item_id);",
        "CREATE INDEX IF NOT EXISTS idx_lotes_codigo_lote ON lotes(codigo_lote);",
        "CREATE INDEX IF NOT EXISTS idx_lotes_tag_rfid ON lotes(tag_rfid);",
        "CREATE INDEX IF NOT EXISTS idx_lotes_qr_code ON lotes(qr_code);",
        "CREATE INDEX IF NOT EXISTS idx_lotes_status ON lotes(status);",
        "",
        "DROP TRIGGER IF EXISTS trg_items_updated_at ON items;",
        "DROP TRIGGER IF EXISTS trg_serial_numbers_updated_at ON serial_numbers;",
        "DROP TRIGGER IF EXISTS trg_projetos_updated_at ON projetos;",
        "DROP TRIGGER IF EXISTS trg_lotes_updated_at ON lotes;",
        """CREATE TRIGGER trg_items_updated_at
BEFORE UPDATE ON items
FOR EACH ROW EXECUTE FUNCTION set_updated_at();""",
        """CREATE TRIGGER trg_serial_numbers_updated_at
BEFORE UPDATE ON serial_numbers
FOR EACH ROW EXECUTE FUNCTION set_updated_at();""",
        """CREATE TRIGGER trg_projetos_updated_at
BEFORE UPDATE ON projetos
FOR EACH ROW EXECUTE FUNCTION set_updated_at();""",
        """CREATE TRIGGER trg_lotes_updated_at
BEFORE UPDATE ON lotes
FOR EACH ROW EXECUTE FUNCTION set_updated_at();""",
        "",
    ]

    for table_name in ("items", "serial_numbers", "projetos", "packing_list", "movimentacoes", "lotes"):
        sql_parts.extend(
            [
                f"ALTER TABLE {table_name} ENABLE ROW LEVEL SECURITY;",
                f"DROP POLICY IF EXISTS authenticated_all_{table_name} ON {table_name};",
                f"CREATE POLICY authenticated_all_{table_name} ON {table_name} FOR ALL TO authenticated USING (true) WITH CHECK (true);",
                "",
            ]
        )

    sql_parts.extend(
        [
            "-- Seed data",
            build_insert_sql(
                "items",
                [
                    "id",
                    "nome",
                    "categoria",
                    "subcategoria",
                    "marca",
                    "modelo",
                    "tipo_rastreamento",
                    "quantidade_total",
                    "valor_mercado_unitario",
                    "foto_url",
                    "notas",
                ],
                items,
            ),
            "",
            build_insert_sql(
                "serial_numbers",
                [
                    "id",
                    "item_id",
                    "codigo_interno",
                    "serial_fabrica",
                    "tag_rfid",
                    "qr_code",
                    "status",
                    "estado",
                    "desgaste",
                    "depreciacao_pct",
                    "valor_atual",
                    "localizacao",
                    "notas",
                ],
                serials,
            ),
            "",
            build_insert_sql(
                "lotes",
                [
                    "id",
                    "item_id",
                    "codigo_lote",
                    "descricao",
                    "quantidade",
                    "tag_rfid",
                    "qr_code",
                    "status",
                ],
                lotes,
            ),
        ]
    )

    return "\n".join(sql_parts) + "\n"


def summarize_category_counts(items: list[dict[str, Any]]) -> list[str]:
    counter = Counter(item["categoria"] for item in items)
    return [f"- {category}: {counter.get(category, 0)} items" for category in CATEGORY_ORDER]


def summarize_status_counts(serials: list[dict[str, Any]]) -> list[str]:
    counter = Counter(serial["status"] for serial in serials)
    ordered_statuses = ["DISPONIVEL", "EMPRESTADO", "VENDIDO", "BAIXA", "PACKED", "EM_CAMPO", "RETORNANDO", "MANUTENCAO"]
    return [f"- {status}: {counter.get(status, 0)}" for status in ordered_statuses if counter.get(status, 0)]


def build_report(
    rows: list[dict[str, Any]],
    items: list[dict[str, Any]],
    serials: list[dict[str, Any]],
    lotes: list[dict[str, Any]],
    maio_groups: dict[str, dict[str, Any]],
    raw_categories: set[str],
    warnings: list[str],
    duplicate_log: list[str],
    match_stats: dict[str, int],
    enrichment_stats: dict[str, int],
) -> str:
    total_rows = len(rows)
    lot_rows = sum(1 for row in rows if row["tracking_type"] == "LOTE")
    serial_rows = total_rows - lot_rows
    lot_quantity_total = sum(lote["quantidade"] for lote in lotes)
    current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    lines = [
        "Sprint 0 - Data Cleanup Report",
        f"Gerado em: {current_date}",
        f"Arquivo origem: {INPUT_WORKBOOK}",
        "",
        "Resumo da base observada",
        f"- Linhas fisicas uteis em EQUIPAMENTOS: {total_rows}",
        f"- Linhas individuais (serial_numbers): {serial_rows}",
        f"- Linhas agrupadas em lotes: {lot_rows}",
        f"- Items gerados: {len(items)}",
        f"- Serial numbers gerados: {len(serials)}",
        f"- Lotes gerados: {len(lotes)}",
        f"- Quantidade total agregada em lotes: {lot_quantity_total}",
        f"- Registros agrupados em EQUIPAMENTOS - MAIO: {len(maio_groups)}",
        "",
        "Divergencias relevantes do prompt",
        "- A aba atual se chama 'EQUIPAMENTOS - MAIO'.",
        "- A aba MAIO nao representa o inventario fisico completo. Ela foi usada como camada de enriquecimento por grupo de item.",
        f"- A aba antiga contem {total_rows} linhas uteis. O prompt antigo citava outro numero.",
        "- O agrupamento em lotes foi aplicado a CABO e EXTENSOR, porque a base atual trata ambos como estoque generico sem serial de fabrica.",
        "",
        "Categorias brutas observadas em EQUIPAMENTOS",
        "- " + ", ".join(sorted(repr(category) for category in raw_categories)),
        "",
        "Distribuicao de items por categoria",
        *summarize_category_counts(items),
        "",
        "Distribuicao de status nos serial_numbers",
        *summarize_status_counts(serials),
        "",
        "Match com EQUIPAMENTOS - MAIO",
        f"- Match exato: {match_stats['matched_exact']}",
        f"- Match fuzzy: {match_stats['matched_fuzzy']}",
        f"- Sem match: {match_stats['unmatched']}",
        f"- Valores preenchidos a partir do MAIO: {enrichment_stats['value_backfills']}",
        f"- Valores sobrescritos pelo MAIO: {enrichment_stats['value_overrides']}",
        f"- Seriais de fabrica preenchidos a partir do MAIO: {enrichment_stats['serial_backfills']}",
        f"- Slots sem serial mesmo apos MAIO: {enrichment_stats['serial_unmatched_slots']}",
        f"- Seriais MAIO nao usados: {enrichment_stats['unused_maio_serials']}",
        "",
        "Duplicatas de serial interno legado tratadas",
    ]

    if duplicate_log:
        lines.extend(f"- {entry}" for entry in duplicate_log)
    else:
        lines.append("- Nenhuma duplicata encontrada.")

    lines.extend(
        [
            "",
            "Warnings",
        ]
    )
    if warnings:
        lines.extend(f"- {warning}" for warning in warnings)
    else:
        lines.append("- Nenhum warning.")

    return "\n".join(lines) + "\n"


def prepare_output_records(items: list[dict[str, Any]], serials: list[dict[str, Any]], lotes: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    item_records = [
        {
            key: value
            for key, value in item.items()
            if key in {
                "id",
                "nome",
                "categoria",
                "subcategoria",
                "marca",
                "modelo",
                "tipo_rastreamento",
                "quantidade_total",
                "valor_mercado_unitario",
                "foto_url",
                "notas",
            }
        }
        for item in items
    ]

    serial_records = [
        {
            key: value
            for key, value in serial.items()
            if key in {
                "id",
                "item_id",
                "codigo_interno",
                "serial_fabrica",
                "tag_rfid",
                "qr_code",
                "status",
                "estado",
                "desgaste",
                "depreciacao_pct",
                "valor_atual",
                "localizacao",
                "notas",
                "legacy_internal_unique",
            }
        }
        for serial in serials
    ]

    lote_records = [
        {
            key: value
            for key, value in lote.items()
            if key in {
                "id",
                "item_id",
                "codigo_lote",
                "descricao",
                "quantidade",
                "tag_rfid",
                "qr_code",
                "status",
                "notas",
            }
        }
        for lote in lotes
    ]

    return item_records, serial_records, lote_records


def main() -> None:
    warnings: list[str] = []
    duplicate_log: list[str] = []

    rows, raw_categories = parse_old_rows(warnings, duplicate_log)
    maio_groups = parse_maio_groups(warnings)
    individual_groups = build_individual_item_groups(rows)
    match_stats = match_maio_groups(individual_groups, maio_groups, warnings)
    enrichment_stats = apply_maio_enrichment(individual_groups)

    serial_items, serials = build_serial_items_and_records(individual_groups)
    lot_items, lotes = build_lot_items_and_lotes(rows)
    items = serial_items + lot_items

    assign_serial_codes(serials)
    assign_lot_codes(lotes)

    item_records, serial_records, lote_records = prepare_output_records(items, serials, lotes)
    serialize_json(item_records, ITEMS_OUTPUT)
    serialize_json(serial_records, SERIALS_OUTPUT)
    serialize_json(lote_records, LOTES_OUTPUT)

    MIGRATION_OUTPUT.write_text(build_migration_sql(item_records, serial_records, lote_records), encoding="utf-8")
    REPORT_OUTPUT.write_text(
        build_report(
            rows=rows,
            items=item_records,
            serials=serial_records,
            lotes=lote_records,
            maio_groups=maio_groups,
            raw_categories=raw_categories,
            warnings=warnings,
            duplicate_log=duplicate_log,
            match_stats=match_stats,
            enrichment_stats=enrichment_stats,
        ),
        encoding="utf-8",
    )

    print(f"Itens gerados: {len(item_records)}")
    print(f"Serial numbers gerados: {len(serial_records)}")
    print(f"Lotes gerados: {len(lote_records)}")
    print(f"Warnings: {len(warnings)}")
    print(f"Duplicatas tratadas: {len(duplicate_log)}")


if __name__ == "__main__":
    main()
