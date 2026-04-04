#!/usr/bin/env python3
from __future__ import annotations

import csv
import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


PROJECT_DIR = Path(__file__).resolve().parent.parent
WORKBOOK_PATH = PROJECT_DIR / "data" / "inventario-limpo.xlsx"
MARKDOWN_PATH = PROJECT_DIR / "data" / "valores-para-importacao.md"
ACTIVE_SHEETS = ["ILUMINACAO", "AUDIO", "ENERGIA", "ESTRUTURA", "EFEITO", "VIDEO", "ACESSORIO", "FORA DE OPERACAO"]
CURRENCY_FORMAT = 'R$ #,##0.00'
PIE_CHART_ANCHOR = "J15"
PIE_CHART_WIDTH = 8.8
PIE_CHART_HEIGHT = 11.5
PIE_LABEL_COLUMN = "X"
PIE_COLORS = ["6366F1", "10B981", "F59E0B", "EF4444", "3B82F6", "8B5CF6", "14B8A6", "F97316"]


def depreciation_formula(row_idx: int, original_col: int, current_col: int) -> str:
    original_ref = f"{get_column_letter(original_col)}{row_idx}"
    current_ref = f"{get_column_letter(current_col)}{row_idx}"
    return f'=IF({original_ref}>0,({original_ref}-{current_ref})/{original_ref},"")'


def extract_csv_block(markdown_text: str) -> str:
    marker = "## CSV Para Importacao"
    start = markdown_text.find(marker)
    if start == -1:
        raise RuntimeError("Secao 'CSV Para Importacao' nao encontrada no markdown.")

    fence_start = markdown_text.find("```csv", start)
    if fence_start == -1:
        raise RuntimeError("Bloco CSV nao encontrado no markdown.")

    content_start = fence_start + len("```csv")
    fence_end = markdown_text.find("```", content_start)
    if fence_end == -1:
        raise RuntimeError("Fim do bloco CSV nao encontrado.")

    return markdown_text[content_start:fence_end].strip()


def load_price_map(markdown_path: Path) -> dict[str, dict[str, str]]:
    csv_text = extract_csv_block(markdown_path.read_text(encoding="utf-8"))
    reader = csv.DictReader(csv_text.splitlines())
    price_map: dict[str, dict[str, str]] = {}
    for row in reader:
        code = (row.get("Codigo") or "").strip()
        if not code:
            continue
        price_map[code] = row
    return price_map


def header_row_index(ws) -> int | None:
    for row_idx in range(1, 8):
        values = [cell.value for cell in ws[row_idx]]
        if "Codigo" in values and "Valor Original" in values:
            return row_idx
    return None


def safe_float(value) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return 0.0
    if "★" in text or "☆" in text:
        return float(text.count("★"))
    cleaned = text.replace("R$", "").replace(".", "").replace(",", ".")
    match = re.search(r"-?\d+(?:\.\d+)?", cleaned)
    if match:
        return float(match.group(0))
    return 0.0


def format_brl(value: float) -> str:
    return f"{value:,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")


def format_decimal_br(value: float, digits: int = 1) -> str:
    return f"{value:.{digits}f}".replace(".", ",")


def star_rating(value: float) -> str:
    rating = int(max(0, min(5, round(value))))
    return ("★" * rating) + ("☆" * (5 - rating))


def build_dashboard_pie_chart(ws) -> PieChart:
    pie = PieChart()
    labels = Reference(ws, min_col=24, min_row=17, max_row=24)
    data = Reference(ws, min_col=4, min_row=16, max_row=24)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Distribuicao do Patrimonio por Categoria"
    pie.varyColors = True
    pie.height = PIE_CHART_HEIGHT
    pie.width = PIE_CHART_WIDTH
    pie.legend = None
    pie.anchor = PIE_CHART_ANCHOR

    pie.dLbls = DataLabelList()
    pie.dLbls.showCatName = True
    pie.dLbls.showVal = False
    pie.dLbls.showPercent = False
    pie.dLbls.showLeaderLines = True

    if pie.ser:
        points = []
        for idx, color in enumerate(PIE_COLORS):
            point = DataPoint(idx=idx)
            point.graphicalProperties = GraphicalProperties(solidFill=color)
            points.append(point)
        pie.ser[0].dPt = points

    return pie


def refresh_dashboard_pie_chart(ws) -> None:
    ws.column_dimensions[PIE_LABEL_COLUMN].hidden = True

    for row in range(17, 25):
        category = ws.cell(row=row, column=2).value
        value = ws.cell(row=row, column=4).value
        if category in (None, "") or value in (None, ""):
            ws[f"{PIE_LABEL_COLUMN}{row}"] = None
            continue
        ws[f"{PIE_LABEL_COLUMN}{row}"] = f"{category}\nR$ {format_brl(float(value))}"

    ws._charts = [chart for chart in ws._charts if type(chart).__name__ != "PieChart"]
    ws.add_chart(build_dashboard_pie_chart(ws))


def collect_sheet_records(ws) -> list[dict[str, object]]:
    row_idx = header_row_index(ws)
    if row_idx is None:
        return []

    headers = [cell.value for cell in ws[row_idx]]
    idx = {name: headers.index(name) for name in headers if name}
    records = []
    for row in ws.iter_rows(min_row=row_idx + 1, values_only=True):
        codigo = row[idx["Codigo"]] if idx.get("Codigo") is not None else None
        nome = row[idx["Nome"]] if idx.get("Nome") is not None else None
        if not codigo:
            continue
        codigo_text = str(codigo).strip()
        if not codigo_text.startswith("MMD-"):
            continue
        subcategoria = (row[idx["Subcategoria"]] or "").strip() if idx.get("Subcategoria") is not None and row[idx["Subcategoria"]] else ""
        nome_text = str(nome).strip() if nome not in (None, "") else subcategoria or "SEM NOME"
        records.append(
            {
                "codigo": codigo_text,
                "nome": nome_text,
                "categoria": ws.title,
                "status": (row[idx["Status"]] or "").strip() if idx.get("Status") is not None and row[idx["Status"]] else "",
                "desgaste": safe_float(row[idx["Desgaste"]]) if idx.get("Desgaste") is not None else 0.0,
                "valor_original": safe_float(row[idx["Valor Original"]]) if idx.get("Valor Original") is not None else 0.0,
                "valor_atual": safe_float(row[idx["Valor Atual"]]) if idx.get("Valor Atual") is not None else 0.0,
            }
        )
    return records


def refresh_category_summary(ws, records: list[dict[str, object]]) -> None:
    if not records or ws.max_row < 2:
        return
    avg_wear = sum(record["desgaste"] for record in records) / len(records)
    total_value = sum(record["valor_original"] for record in records)
    ws["A2"] = f"{len(records)} itens   |   Valor: R$ {format_brl(total_value)}   |   Desgaste medio: {format_decimal_br(avg_wear)}/5"


def refresh_dashboard(wb) -> None:
    if "DASHBOARD" not in wb.sheetnames:
        return

    dashboard = wb["DASHBOARD"]
    active_data = {sheet: collect_sheet_records(wb[sheet]) for sheet in ACTIVE_SHEETS if sheet in wb.sheetnames}

    for sheet_name, records in active_data.items():
        refresh_category_summary(wb[sheet_name], records)

    active_records = [record for records in active_data.values() for record in records]
    lotes = []
    if "LOTES" in wb.sheetnames:
        lotes_ws = wb["LOTES"]
        for row in lotes_ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                lotes.append(row[0])

    total = len(active_records)
    disponivel = sum(1 for record in active_records if record["status"] == "DISPONIVEL")
    em_campo = sum(1 for record in active_records if record["status"] == "EM_CAMPO")
    manutencao = sum(1 for record in active_records if record["status"] == "MANUTENCAO")
    total_original = sum(record["valor_original"] for record in active_records)
    total_atual = sum(record["valor_atual"] for record in active_records)
    avg_wear = (sum(record["desgaste"] for record in active_records) / total) if total else 0.0
    criticos = sum(1 for record in active_records if record["desgaste"] <= 2)
    deprec_pct = round(((total_original - total_atual) / total_original) * 100, 2) if total_original else 0.0

    dashboard["B3"] = f"Inventario Inteligente  ·  {datetime.now().strftime('%d/%m/%Y')}"
    dashboard["B5"] = total
    dashboard["D5"] = disponivel
    dashboard["F5"] = em_campo
    dashboard["B8"] = manutencao
    dashboard["D8"] = round(total_original, 2)
    dashboard["F8"] = round(total_atual, 2)
    dashboard["B12"] = star_rating(avg_wear)
    dashboard["C12"] = criticos
    dashboard["D12"] = f"{deprec_pct:.0f}%"
    dashboard["E12"] = len(lotes)
    dashboard["D8"].number_format = CURRENCY_FORMAT
    dashboard["F8"].number_format = CURRENCY_FORMAT

    row = 17
    for sheet_name in ACTIVE_SHEETS:
        records = active_data.get(sheet_name, [])
        category_total = sum(record["valor_original"] for record in records)
        category_current = sum(record["valor_atual"] for record in records)
        dashboard.cell(row=row, column=2).value = sheet_name
        dashboard.cell(row=row, column=3).value = len(records)
        dashboard.cell(row=row, column=4).value = round(category_total, 2)
        dashboard.cell(row=row, column=5).value = round(category_current, 2)
        dashboard.cell(row=row, column=6).value = (category_total / total_original) if total_original else 0
        dashboard.cell(row=row, column=4).number_format = CURRENCY_FORMAT
        dashboard.cell(row=row, column=5).number_format = CURRENCY_FORMAT
        dashboard.cell(row=row, column=6).number_format = "0.0%"
        row += 1

    while row <= 24:
        for col in range(2, 7):
            dashboard.cell(row=row, column=col).value = None
        row += 1

    status_rows = {
        "DISPONIVEL": 28,
        "EM_CAMPO": 29,
        "PACKED": 30,
        "MANUTENCAO": 31,
    }
    for status, row_idx in status_rows.items():
        count = sum(1 for record in active_records if record["status"] == status)
        pct = (count / total) if total else 0
        dashboard.cell(row=row_idx, column=2).value = status
        dashboard.cell(row=row_idx, column=3).value = count
        dashboard.cell(row=row_idx, column=4).value = pct
        dashboard.cell(row=row_idx, column=4).number_format = "0.0%"

    top_records = sorted(active_records, key=lambda record: record["valor_original"], reverse=True)[:10]
    top_row = 35
    for record in top_records:
        dashboard.cell(row=top_row, column=2).value = record["nome"]
        dashboard.cell(row=top_row, column=3).value = record["categoria"]
        dashboard.cell(row=top_row, column=4).value = round(record["valor_original"], 2)
        dashboard.cell(row=top_row, column=5).value = star_rating(record["desgaste"])
        dashboard.cell(row=top_row, column=6).value = round(record["valor_atual"], 2)
        dashboard.cell(row=top_row, column=4).number_format = CURRENCY_FORMAT
        dashboard.cell(row=top_row, column=6).number_format = CURRENCY_FORMAT
        top_row += 1

    while top_row <= 44:
        for col in range(2, 7):
            dashboard.cell(row=top_row, column=col).value = None
        top_row += 1

    refresh_dashboard_pie_chart(dashboard)


def main() -> int:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Planilha nao encontrada: {WORKBOOK_PATH}")
    if not MARKDOWN_PATH.exists():
        raise FileNotFoundError(f"Markdown nao encontrado: {MARKDOWN_PATH}")

    price_map = load_price_map(MARKDOWN_PATH)

    wb = load_workbook(WORKBOOK_PATH)
    applied = 0
    unmatched_markdown = set(price_map.keys())

    for ws in wb.worksheets:
        row_idx = header_row_index(ws)
        if row_idx is None:
            continue

        headers = [cell.value for cell in ws[row_idx]]
        idx = {name: headers.index(name) for name in headers if name}
        orig_col = idx["Valor Original"] + 1
        curr_col = idx["Valor Atual"] + 1
        dep_col = idx["Deprec.%"] + 1
        code_col = idx["Codigo"]

        for data_row in ws.iter_rows(min_row=row_idx + 1, values_only=False):
            code = data_row[code_col].value
            if not code:
                continue
            code = str(code).strip()
            payload = price_map.get(code)
            if not payload:
                continue

            try:
                valor_original = safe_float(payload["Valor Original (R$)"])
                valor_atual = safe_float(payload["Valor Atual (R$)"])
            except (TypeError, ValueError) as exc:
                raise RuntimeError(f"Valores invalidos no markdown para {code}: {payload}") from exc

            ws.cell(row=data_row[0].row, column=orig_col).value = valor_original
            ws.cell(row=data_row[0].row, column=curr_col).value = valor_atual
            ws.cell(row=data_row[0].row, column=dep_col).value = depreciation_formula(data_row[0].row, orig_col, curr_col)
            ws.cell(row=data_row[0].row, column=orig_col).number_format = CURRENCY_FORMAT
            ws.cell(row=data_row[0].row, column=curr_col).number_format = CURRENCY_FORMAT
            ws.cell(row=data_row[0].row, column=dep_col).number_format = "0%"
            applied += 1
            unmatched_markdown.discard(code)

    backup_path = WORKBOOK_PATH.with_name(f"{WORKBOOK_PATH.stem}.backup-apply-{datetime.now().strftime('%Y%m%d-%H%M%S')}{WORKBOOK_PATH.suffix}")
    shutil.copy2(WORKBOOK_PATH, backup_path)
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcOnSave = True
    refresh_dashboard(wb)
    wb.save(WORKBOOK_PATH)

    print(f"Backup criado em: {backup_path}")
    print(f"Linhas aplicadas: {applied}")
    print(f"Codigos no markdown sem correspondencia na planilha atual: {len(unmatched_markdown)}")
    if unmatched_markdown:
        for code in sorted(unmatched_markdown)[:25]:
            print(f" - {code}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
