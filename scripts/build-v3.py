"""
MMD Inventario Inteligente v3.5
Max quality: star ratings, KPI cards, subtotals, hyperlinks, footer,
desgaste chart, conditional formatting, data validation.
"""
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, BarChart, BarChart3D, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from collections import defaultdict
import csv
import os, datetime
import re
import zipfile
import xml.etree.ElementTree as ET

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT = os.path.join(SCRIPT_DIR, '..', 'data', 'inventario-original.xlsx')
OUTPUT = os.path.join(SCRIPT_DIR, '..', 'data', 'inventario-limpo.xlsx')
PRICING_MARKDOWN = os.path.join(SCRIPT_DIR, '..', 'data', 'valores-para-importacao.md')
SHEET_XML_NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
ET.register_namespace('', SHEET_XML_NS)

# ── MODERN PALETTE ────────────────────────────────────────────
DARK = '0F172A'       # slate-900
MID = '334155'        # slate-700
LIGHT = '475569'      # slate-600
SUBTLE = 'CBD5E1'     # slate-300
BG_ALT = 'F1F5F9'     # slate-100
BG_WHITE = 'FFFFFF'
ACCENT = '4F46E5'     # indigo-600
ACCENT_LIGHT = 'E0E7FF'  # indigo-100

CAT_DESIGN = {
    'ILUMINACAO': {'tab':'6366F1','bg':'EEF2FF','tx':'4338CA','prefix':'ILU','label':'Iluminacao','desc':'Par LED, Ribalta, Moving, Strobo, Laser'},
    'AUDIO':      {'tab':'10B981','bg':'ECFDF5','tx':'047857','prefix':'AUD','label':'Audio','desc':'Caixas, Subs, Mesas, Mics, DJs'},
    'CABO':       {'tab':'F59E0B','bg':'FFFBEB','tx':'92400E','prefix':'CAB','label':'Cabo','desc':'XLR, DMX, P10, AC, HDMI, USB'},
    'ENERGIA':    {'tab':'EF4444','bg':'FEF2F2','tx':'B91C1C','prefix':'ENE','label':'Energia','desc':'Reguas, Centopeias, Extensoes'},
    'ESTRUTURA':  {'tab':'3B82F6','bg':'EFF6FF','tx':'1D4ED8','prefix':'EST','label':'Estrutura','desc':'Tripes, Box Truss, Praticaveis'},
    'EFEITO':     {'tab':'8B5CF6','bg':'F5F3FF','tx':'6D28D9','prefix':'EFE','label':'Efeito','desc':'Fumaca, Haze, CO2, Fluidos'},
    'VIDEO':      {'tab':'14B8A6','bg':'F0FDFA','tx':'0F766E','prefix':'VID','label':'Video','desc':'Projetores, Notebooks, Tablets'},
    'ACESSORIO':  {'tab':'F97316','bg':'FFF7ED','tx':'C2410C','prefix':'ACE','label':'Acessorio','desc':'Cases, Ferramentas, Roteadores'},
}
CAT_ORDER = ['ILUMINACAO','AUDIO','CABO','ENERGIA','ESTRUTURA','EFEITO','VIDEO','ACESSORIO']

# Cores mais saturadas e modernas (nada pastel fraco)
STATUS_STYLE = {
    'DISPONIVEL': {'bg':'BBF7D0','tx':'14532D'},
    'PACKED':     {'bg':'BFDBFE','tx':'1E3A8A'},
    'EM_CAMPO':   {'bg':'FDE68A','tx':'78350F'},
    'MANUTENCAO': {'bg':'FECACA','tx':'7F1D1D'},
    'EMPRESTADO': {'bg':'DDD6FE','tx':'4C1D95'},
    'VENDIDO':    {'bg':'E2E8F0','tx':'475569'},
    'BAIXA':      {'bg':'E2E8F0','tx':'475569'},
}
ESTADO_STYLE = {
    'NOVO':           {'bg':'BBF7D0','tx':'14532D'},
    'SEMI_NOVO':      {'bg':'BFDBFE','tx':'1E3A8A'},
    'USADO':          {'bg':'FDE68A','tx':'78350F'},
    'RECONDICIONADO': {'bg':'FED7AA','tx':'7C2D12'},
}
DESGASTE_BG = {5:'86EFAC',4:'93C5FD',3:'FDE047',2:'FDBA74',1:'FCA5A5'}
VALUE_ATUAL_BG = 'DCFCE7'
VALUE_ATUAL_TX = '166534'

BRAND_MAP = {
    'QSC':'QSC','SHOWTECH':'Showtech','BRIWAX':'Briwax','SENNHEISER':'Sennheiser',
    'BEHRINGER':'Behringer','PIONEER':'Pioneer','SHURE':'Shure','JBL':'JBL',
    'YAMAHA':'Yamaha','HOT MACHINE':'Hot Machine','PHENYX PRO':'Phenyx Pro',
    'WIRECONNEX':'Wireconnex','STANER AUDIO AMERICA':'Staner','RMV':'RMV',
    'BDL':'BDL','LEACS':'Leacs','SOG':'SOG','TOMATE':'Tomate',
    'MEGA LIGHT':'Mega Light','STAR LIGHT DIVISION':'Star Light Division',
    'SKYPIX':'Skypix','AURA':'Aura','CSR':'CSR','MACKIE':'Mackie','RCF':'RCF',
    'SOUNDBOX':'Soundbox','MARK AUDIO':'Mark Audio','NHL':'NHL',
    'CHINES':'','S/ MARCA':'','SEM MARCA':'','NONE':'','S/MARCA':'',
    'GATOR':'Gator','RODE':'Rode','BOSS':'Boss','ROLAND':'Roland',
}
CATEGORY_MAP = {
    'PAR LEDS':('ILUMINACAO','Par Led'),'PAR LEDS ':('ILUMINACAO','Par Led'),
    'RIBALTA':('ILUMINACAO','Ribalta'),'Ribalta':('ILUMINACAO','Ribalta'),
    'MINI MOVING':('ILUMINACAO','Mini Moving'),'Mini Moving':('ILUMINACAO','Mini Moving'),
    'MOVING BEAM':('ILUMINACAO','Moving Beam'),'Moving Beam':('ILUMINACAO','Moving Beam'),
    'MINI BRUTE':('ILUMINACAO','Mini Brute'),'OUTRAS LUZES':('ILUMINACAO','Outras Luzes'),
    'LUZ':('ILUMINACAO','Outras Luzes'),'LUZ NEGRA':('ILUMINACAO','Luz Negra'),
    'LASER':('ILUMINACAO','Laser'),'STROBO':('ILUMINACAO','Strobo'),
    'Iluminação':('ILUMINACAO','Geral'),
    'AUDIO':('AUDIO','Geral'),'Audio':('AUDIO','Geral'),'audio':('AUDIO','Geral'),
    'SOM':('AUDIO','Geral'),'Som':('AUDIO','Geral'),'som':('AUDIO','Geral'),
    'INSTRUMENTO':('AUDIO','Instrumento'),'CAPSULAS MIC':('AUDIO','Capsula Mic'),
    'MIDI':('AUDIO','Midi'),
    'CABO':('CABO','Geral'),'EXTENSOR':('CABO','Extensor'),
    'REGUA':('ENERGIA','Regua'),'ELETRICA':('ENERGIA','Eletrica'),
    'Eletrica':('ENERGIA','Eletrica'),'Energia':('ENERGIA','Geral'),'FONTE':('ENERGIA','Fonte'),
    'TRIPÉ':('ESTRUTURA','Tripe'),'SUPORTE':('ESTRUTURA','Suporte'),
    'Estrutura':('ESTRUTURA','Geral'),'PRATICÁVEL':('ESTRUTURA','Praticavel'),
    'Praticável':('ESTRUTURA','Praticavel'),
    'EFEITOS':('EFEITO','Geral'),'Efeitos':('EFEITO','Geral'),
    'FLUIDO':('EFEITO','Fluido'),'Fluido':('EFEITO','Fluido'),
    'GLOBO':('EFEITO','Globo'),'Globo':('EFEITO','Globo'),
    'VIDEO':('VIDEO','Geral'),'COMPUTADORES':('VIDEO','Computador'),
    'NOTEBOOK':('VIDEO','Notebook'),'Notebook':('VIDEO','Notebook'),
    'TABLET':('VIDEO','Tablet'),'Tablet':('VIDEO','Tablet'),
    'CASE':('ACESSORIO','Case'),'CASES':('ACESSORIO','Case'),
    'FERRAMENTAS':('ACESSORIO','Ferramenta'),
    'ROTEADOR':('ACESSORIO','Roteador'),'Roteador':('ACESSORIO','Roteador'),
    'REDE':('ACESSORIO','Rede'),'RADIO':('ACESSORIO','Radio'),'Radio':('ACESSORIO','Radio'),
    'Refrigeracao':('ACESSORIO','Refrigeracao'),'REFRIGERACAO':('ACESSORIO','Refrigeracao'),
}
ESTADO_FATOR = {'NOVO':1.0,'SEMI_NOVO':0.85,'USADO':0.65,'RECONDICIONADO':0.50}
BASE_WEAR = {'AUDIO':3,'ILUMINACAO':2,'EFEITO':2,'ESTRUTURA':3,'ENERGIA':3,'VIDEO':3,'ACESSORIO':3,'CABO':3}
HIGH_END_AUDIO = ('XDJ-XZ','XDJ-RR','DDJ-400','TRAKTOR KONTROL','K12.2','K10.2','KW181','DXR15','DXS12','RCF 8003')
MID_AUDIO = ('XDJ-RX','MG16XU','VENUE 360','HA8000','NU3000','CX3400','FBQ1502','DI600P','PHANTOM POWERED DI-BOX','PAS-225','DSM-600','PGA52','KORG MINI-KP')
LOW_AUDIO = ('K8/K10','LI2400','T500D','UB1202','Z4 STUDIO R')
LIGHT_CONTROL = ('COMMAND WING','PILOT 2000','PRO X4I','DMX DISTRIBUTOR','SPLITTER DMX')
LIGHT_LEGACY = ('ST-X251LAY','RIBALTA LASER','BX-570','SOG-610A','WASH SOG-837','ST-S206','BRY-1912','BX-325','BX-432','BX-353','BRY-1903','HW-LED267')
ENTRY_EFFECT = ('SK-FM900','LK-Y4-600W','YJ-600W','F-950')

# ── STYLE HELPERS ─────────────────────────────────────────────
def FN(sz=10, bold=False, color=MID, name='Arial'):
    return Font(name=name, size=sz, bold=bold, color=color)

def BG(color):
    return PatternFill(start_color=color, end_color=color, fill_type='solid')

def BD(color=SUBTLE):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(ws, row, ncols):
    for c in range(1, ncols+1):
        cl = ws.cell(row=row, column=c)
        cl.font = FN(10, True, 'F8FAFC')
        cl.fill = BG(DARK)
        cl.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cl.border = BD(DARK)

def drow(ws, row, ncols, alt=False):
    bg = BG(BG_ALT) if alt else BG(BG_WHITE)
    for c in range(1, ncols+1):
        cl = ws.cell(row=row, column=c)
        cl.fill = bg
        cl.border = BD()
        cl.alignment = Alignment(vertical='center')
        cl.font = FN(10)

def W(ws, ws_list):
    for i, w in enumerate(ws_list, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def norm_brand(raw):
    if not raw: return ''
    key = raw.strip().upper()
    if key in BRAND_MAP: return BRAND_MAP[key]
    if key in ('','NONE'): return ''
    words = raw.strip().split()
    return ' '.join(w.upper() if len(w) <= 3 else w.title() for w in words)

def trunc(text, mx=60):
    if not text or text == 'None': return '', ''
    text = text.strip()
    if len(text) <= mx: return text, ''
    return text[:mx-3] + '...', text

def trunc_serial(text, mx=18):
    if not text or text == 'None': return ''
    text = text.strip()
    if len(text) <= mx: return text
    return text[:mx-3] + '...'

def merge_notes(*parts):
    merged = []
    seen = set()
    for part in parts:
        if not part:
            continue
        clean = ' '.join(str(part).split()).strip()
        if not clean:
            continue
        key = clean.upper()
        if key in seen:
            continue
        seen.add(key)
        merged.append(clean)
    return ' | '.join(merged)

def normalize_subcategory(subcat_raw, fallback):
    value = subcat_raw or fallback or 'OUTRO'
    value = ' '.join(str(value).strip().split())
    value = re.sub(r'\s*/\s*', ' / ', value)
    value = re.sub(r'\s+', ' ', value)
    value = value.upper()
    typo_fixes = {
        'MIFROFONE': 'MICROFONE',
        'MIROFONE': 'MICROFONE',
        'MICROFONSE': 'MICROFONE',
    }
    for wrong, right in typo_fixes.items():
        value = value.replace(wrong, right)

    aliases = {
        'MICROFONE SEM FIO': 'MIC S/ FIO',
        'MICROFONE WIRELESS': 'MIC S/ FIO',
        'MICROFONE COM FIO': 'MIC C/ FIO',
        'MICROFONE LAPELA': 'MIC LAPELA',
        'MICROFONE': 'MIC',
        'RECEPTOR MICROFONE': 'RX MIC',
        'BASE DE MICROFONE': 'BASE MIC',
        'CAIXA DE SOM': 'CAIXA SOM',
        'MESA DE SOM': 'MESA SOM',
        'MESA DE LUZ': 'MESA LUZ',
        'CONTROLADOR DMX': 'CTRL DMX',
        'AMPLIFICADOR DE FONES': 'AMP FONES',
        'AMPLIFICADOR RF': 'AMP RF',
        'AMPLIFICADOR': 'AMP',
        'SUBWOOFER': 'SUB',
        'DIRECT BOX': 'DI',
        'TRIPÉ DE MICROFONE': 'TRIPE MIC',
        'TRIPE DE MICROFONE': 'TRIPE MIC',
        'TRIPÉ DE CAIXA': 'TRIPE CAIXA',
        'TRIPE DE CAIXA': 'TRIPE CAIXA',
        'MÁQUINA DE FUMAÇA': 'MAQ FUMACA',
        'MAQUINA DE FUMAÇA': 'MAQ FUMACA',
        'MAQUINA DE FUMACA': 'MAQ FUMACA',
        'DISTRIBUIDOR DE ENERGIA': 'DIST ENERGIA',
        'PROCESSADOR DE AUDIO': 'PROC AUDIO',
        'PROCESSADOR DE ÁUDIO': 'PROC AUDIO',
        'SUPORTE DE TV': 'SUP TV',
        'BATERIA DE MAO': 'BAT MAO',
        'BATERIA DE MÃO': 'BAT MAO',
        'CAPTADOR MIDI': 'CAPT MIDI',
        'FONE DE OUVIDO': 'FONE',
        'EXTENSAO 1 PONTA': 'EXT 1 PTA',
        'EXTENSOR AC': 'EXT AC',
        'EXTENSOR P10': 'EXT P10',
        'EXTENSOR P2': 'EXT P2',
        'EXTENSOR HDMI': 'EXT HDMI',
        'EXTENSOR USB': 'EXT USB',
    }
    return aliases.get(value, value)

def normalize_serial(serial_fab):
    if not serial_fab:
        return '', ''

    serial = ' '.join(str(serial_fab).replace('\n', ' ').replace('\r', ' ').split()).strip()
    if not serial or serial == ',':
        return '', ''

    serial_up = serial.upper()
    invalid_markers = ('S/ SERIAL', 'SEM SERIAL', 'SEM SERIAL APARENTE', 'SEM ID', 'NONE')
    if any(marker in serial_up for marker in invalid_markers):
        return '', ''

    serial_notes = []

    label_match = re.match(r'^(BODYPACK|BASE|NA CAIXA)\s*:\s*(.+)$', serial, flags=re.IGNORECASE)
    if label_match:
        label, serial = label_match.groups()
        serial_notes.append(f'Serial Fab.: {label.strip().upper()}')

    if ' - ' in serial:
        serial, detail = serial.split(' - ', 1)
        serial_notes.append(f'Serial Fab.: {detail.strip()}')

    if '*' in serial:
        serial, detail = serial.split('*', 1)
        serial_notes.append(f'Serial Fab.: {detail.strip()}')

    zero_prefix_match = re.match(r'^\((\d+\s+ZEROS?|ZERO)\)\s*(.+)$', serial, flags=re.IGNORECASE)
    if zero_prefix_match:
        zero_hint, serial = zero_prefix_match.groups()
        serial_notes.append(f'Serial Fab.: {zero_hint.strip().upper()}')

    trailing_note_match = re.match(r'^(.+?)\s*\(([^)]+)\)$', serial)
    if trailing_note_match:
        candidate, detail = trailing_note_match.groups()
        detail_up = detail.upper()
        if any(marker in detail_up for marker in ('CAIXA', 'CODIGO', 'RISCAD', 'ETIQUETA')):
            serial = candidate
            serial_notes.append(f'Serial Fab.: {detail.strip()}')

    serial = serial.strip(' -:;')
    if not serial:
        return '', merge_notes(*serial_notes)

    if re.fullmatch(r'\d+\.0+', serial):
        serial = serial.split('.', 1)[0]

    serial = re.sub(r'\s*/\s*', ' / ', serial)
    serial = re.sub(r'\s+', ' ', serial)

    return serial.upper(), merge_notes(*serial_notes)

def has_real_serial(serial_fab):
    if not serial_fab:
        return False
    serial_up = serial_fab.strip().upper()
    invalid_markers = ('S/ SERIAL', 'SEM SERIAL', 'SEM ID', 'NONE')
    return not any(marker in serial_up for marker in invalid_markers)

def has_any(text, hints):
    return any(hint in text for hint in hints)

def infer_condition(cat, subcat, brand, model, serial_fab, notes_raw, status, valor_num):
    text = ' '.join(filter(None, [cat, subcat, brand, model, serial_fab, notes_raw])).upper()
    wear = BASE_WEAR.get(cat, 3)
    state = 'USADO'
    is_light_control = has_any(text, LIGHT_CONTROL)

    if has_any(text, HIGH_END_AUDIO):
        wear = 4
        state = 'SEMI_NOVO'
    elif has_any(text, MID_AUDIO):
        wear = max(wear, 3)
    elif has_any(text, LOW_AUDIO):
        wear = min(wear, 2)

    if is_light_control:
        wear = max(wear, 3)
    elif has_any(text, LIGHT_LEGACY):
        wear = 2

    if has_any(text, ENTRY_EFFECT):
        wear = 2

    if cat == 'AUDIO' and ('CAIXA DE SOM' in text or 'SUBWOOFER' in text):
        if has_any(text, ('K12.2','K10.2','KW181','DXR15','DXS12','RCF 8003')):
            wear = max(wear, 4)
            state = 'SEMI_NOVO'
        else:
            wear = max(wear, 3)

    if cat == 'ESTRUTURA' and ('PRATICAVEL' in text or 'BOX TRUSS' in text or 'Q20' in text):
        wear = 3

    if 'COMPROU' in text:
        wear = 5
        state = 'SEMI_NOVO'

    if 'SEM PLACA' in text:
        wear -= 1
    if 'ADESIVO' in text or 'RISCAD' in text:
        wear -= 1

    if cat == 'ILUMINACAO' and not is_light_control and ('S/ SERIAL' in text or not serial_fab):
        wear = min(wear, 2 if (valor_num or 0) < 5000 else 3)

    if cat == 'AUDIO' and (valor_num or 0) >= 3000 and ('S/ SERIAL' in text or 'SEM SERIAL APARENTE' in text or not serial_fab):
        wear -= 1

    if status == 'BAIXA':
        wear = 1

    wear = max(1, min(5, int(wear)))
    if state == 'SEMI_NOVO' and wear < 4:
        state = 'USADO'
    return state, wear

def stars(n):
    return '\u2605' * n + '\u2606' * (5 - n)  # ★★★☆☆

def load_price_map():
    if not os.path.exists(PRICING_MARKDOWN):
        return {}
    with open(PRICING_MARKDOWN, encoding='utf-8') as fh:
        markdown = fh.read()
    marker = '## CSV Para Importacao'
    start = markdown.find(marker)
    if start == -1:
        return {}
    fence_start = markdown.find('```csv', start)
    if fence_start == -1:
        return {}
    content_start = fence_start + len('```csv')
    fence_end = markdown.find('```', content_start)
    if fence_end == -1:
        return {}
    csv_text = markdown[content_start:fence_end].strip()
    reader = csv.DictReader(csv_text.splitlines())
    price_map = {}
    for row in reader:
        code = (row.get('Codigo') or '').strip()
        value = (row.get('Valor Original (R$)') or '').strip()
        current_value = (row.get('Valor Atual (R$)') or '').strip()
        method = (row.get('Metodo') or '').strip()
        if not code:
            continue
        record = {}
        try:
            if value:
                record['valor_original'] = float(value)
        except ValueError:
            pass
        try:
            if current_value and method and method != 'valor_original_existente':
                record['valor_atual'] = float(current_value)
        except ValueError:
            pass
        if record:
            price_map[code] = record
    return price_map

def estimate_current_value(valor_original, estado, desgaste):
    if not isinstance(valor_original, (int, float)) or not valor_original:
        return None
    if not estado or desgaste in (None, ''):
        return None
    factor = ESTADO_FATOR.get(str(estado).strip().upper(), 0.65)
    try:
        desgaste_num = float(desgaste)
    except (TypeError, ValueError):
        return None
    return round(float(valor_original) * (desgaste_num / 5.0) * factor, 2)

def display_name(brand, model):
    model_name = ' '.join(str(model or '').split()).strip()
    brand_name = ' '.join(str(brand or '').split()).strip()
    if brand_name and model_name.upper().startswith(f'{brand_name.upper()} '):
        model_name = model_name[len(brand_name):].strip(' -/')
    return model_name or brand_name

def current_value_for_item(item):
    manual_value = item.get('valor_atual_manual')
    if isinstance(manual_value, (int, float)):
        return round(float(manual_value), 2)
    return estimate_current_value(item.get('valor_mercado'), item.get('estado'), item.get('desgaste'))

def kpi_card(ws, row, col, value, label, color, fmt='#,##0', span=2):
    """Render a KPI card with colored background block."""
    # Background block
    for c in range(col, col + span):
        for r in [row, row+1, row+2]:
            ws.cell(row=r, column=c).fill = BG(ACCENT_LIGHT)
            ws.cell(row=r, column=c).border = BD('C7D2FE')
    # Value
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    cv = ws.cell(row=row, column=col)
    cv.value = value
    cv.font = FN(18, True, color)
    cv.number_format = fmt
    cv.alignment = Alignment(horizontal='center', vertical='center')
    # Label
    ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+span-1)
    cl = ws.cell(row=row+1, column=col)
    cl.value = label
    cl.font = FN(8, True, LIGHT)
    cl.alignment = Alignment(horizontal='center', vertical='center')

def add_footer(ws, row, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1)
    c.value = f'MMD Eventos  |  Inventario Inteligente  |  {datetime.date.today().strftime("%d/%m/%Y")}'
    c.font = FN(8, False, SUBTLE)
    c.alignment = Alignment(horizontal='center')
    ws.row_dimensions[row].height = 18

def format_brl_label(value):
    return f'R$ {value:,.0f}'.replace(',', 'X').replace('.', ',').replace('X', '.')

# ── DATA VALIDATION + CONDITIONAL FORMATTING ──────────────────
def add_validations(ws, start_row, end_row, status_col='G', estado_col='H', desgaste_col='I'):
    dv_status = DataValidation(type='list', formula1='"DISPONIVEL,PACKED,EM_CAMPO,MANUTENCAO,EMPRESTADO,VENDIDO,BAIXA"', allow_blank=True)
    dv_status.error = 'Selecione um status valido'
    dv_status.errorTitle = 'Status invalido'
    dv_estado = DataValidation(type='list', formula1='"NOVO,SEMI_NOVO,USADO,RECONDICIONADO"', allow_blank=True)
    dv_desgaste = DataValidation(type='list', formula1=f'"{stars(1)},{stars(2)},{stars(3)},{stars(4)},{stars(5)}"', allow_blank=True)
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_estado)
    ws.add_data_validation(dv_desgaste)
    dv_status.add(f'{status_col}{start_row}:{status_col}{end_row}')
    dv_estado.add(f'{estado_col}{start_row}:{estado_col}{end_row}')
    dv_desgaste.add(f'{desgaste_col}{start_row}:{desgaste_col}{end_row}')

def add_cond_fmt(ws, start_row, end_row, status_col='G', estado_col='H', desgaste_col='I', deprec_col='L'):
    rng_s = f'{status_col}{start_row}:{status_col}{end_row}'
    for st, sty in STATUS_STYLE.items():
        ws.conditional_formatting.add(rng_s, CellIsRule(operator='equal', formula=[f'"{st}"'], fill=BG(sty['bg']), font=FN(10, True, sty['tx'])))
    rng_e = f'{estado_col}{start_row}:{estado_col}{end_row}'
    for est, sty in ESTADO_STYLE.items():
        ws.conditional_formatting.add(rng_e, CellIsRule(operator='equal', formula=[f'"{est}"'], fill=BG(sty['bg']), font=FN(10, False, sty['tx'])))
    rng_d = f'{desgaste_col}{start_row}:{desgaste_col}{end_row}'
    for val, bg_c in DESGASTE_BG.items():
        ws.conditional_formatting.add(rng_d, FormulaRule(formula=[f'{desgaste_col}{start_row}="{stars(val)}"'], fill=BG(bg_c), font=FN(10, True, DARK)))
    rng_p = f'{deprec_col}{start_row}:{deprec_col}{end_row}'
    ws.conditional_formatting.add(rng_p, CellIsRule(operator='lessThan', formula=['0.2'], fill=BG('BBF7D0'), font=FN(10, True, '14532D')))
    ws.conditional_formatting.add(rng_p, CellIsRule(operator='between', formula=['0.2', '0.4'], fill=BG('FDE68A'), font=FN(10, True, '78350F')))
    ws.conditional_formatting.add(rng_p, CellIsRule(operator='between', formula=['0.4', '0.6'], fill=BG('FDBA74'), font=FN(10, True, '7C2D12')))
    ws.conditional_formatting.add(rng_p, CellIsRule(operator='greaterThanOrEqual', formula=['0.6'], fill=BG('FCA5A5'), font=FN(10, True, '7F1D1D')))

# ── READ DATA ─────────────────────────────────────────────────
def read_data():
    wb = load_workbook(INPUT, data_only=True)
    ws = wb['EQUIPAMENTOS']
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=False):
        cat_raw = str(row[2].value).strip() if row[2].value else ''
        modelo = str(row[5].value).strip() if row[5].value else ''
        if not cat_raw and not modelo: continue
        ok = row[1].value
        subcat_raw = str(row[3].value).strip() if row[3].value else ''
        marca_raw = str(row[4].value).strip() if row[4].value else ''
        serial_fab_raw = str(row[9].value).strip() if row[9].value else ''
        valor = row[10].value
        notas_raw = str(row[13].value).strip() if len(row) > 13 and row[13].value else ''

        mapped = CATEGORY_MAP.get(cat_raw, ('ACESSORIO','Outro'))
        cat, subcat = mapped
        if subcat_raw and subcat_raw.upper() not in ('SEM SUB','NONE',''):
            subcat = subcat_raw
        subcat = normalize_subcategory(subcat_raw, subcat)

        serial_fab, serial_note = normalize_serial(serial_fab_raw)
        notas_raw = merge_notes(notas_raw, serial_note)
        has_serial = has_real_serial(serial_fab)
        valor_num = None
        if valor:
            try: valor_num = float(valor)
            except: pass

        marca = norm_brand(marca_raw)
        ok_str = str(ok).strip().upper() if ok else ''
        notas_up = notas_raw.upper() if notas_raw else ''

        if 'VENDID' in ok_str or 'VENDID' in notas_up:
            status, motivo = 'VENDIDO', notas_raw or 'Vendido'
        elif 'BIGODE' in ok_str:
            status, motivo = 'EMPRESTADO', 'Emprestado - Seu Bigode'
        elif 'QUEBRAD' in ok_str:
            status, motivo = 'BAIXA', 'Quebrado'
        else:
            status, motivo = 'DISPONIVEL', ''

        estado, desgaste = infer_condition(cat, subcat, marca, modelo, serial_fab if has_serial else '', notas_raw, status, valor_num)
        notas_short, notas_full = trunc(notas_raw)

        rows.append({
            'categoria':cat, 'subcategoria':subcat, 'marca':marca,
            'modelo':modelo, 'serial_fabrica':serial_fab if has_serial else '',
            'valor_mercado':valor_num, 'status':status, 'motivo':motivo,
            'valor_atual_manual':None,
            'estado':estado, 'desgaste':desgaste,
            'notas_short':notas_short, 'notas_full':notas_full,
        })
    return rows

def assign_codes(rows):
    counters = defaultdict(int)
    for r in rows:
        counters[r['categoria']] += 1
        pfx = CAT_DESIGN[r['categoria']]['prefix']
        r['codigo'] = f"MMD-{pfx}-{counters[r['categoria']]:04d}"
    return rows

def apply_price_overrides(rows, price_map):
    applied = 0
    if not price_map:
        return applied
    for row in rows:
        payload = price_map.get(row['codigo'])
        if payload is None:
            continue
        if payload.get('valor_original') is not None:
            row['valor_mercado'] = payload['valor_original']
        if payload.get('valor_atual') is not None:
            row['valor_atual_manual'] = payload['valor_atual']
        applied += 1
    return applied

def build_formula_cache_map(wb):
    cache_map = {}
    for ws in wb.worksheets:
        if ws.title not in CAT_ORDER and ws.title != 'FORA DE OPERACAO':
            continue
        ws_cache = {}
        start_row = 2 if ws.title == 'FORA DE OPERACAO' else 5
        for r in range(start_row, ws.max_row + 1):
            valor_original = ws.cell(r, 10).value
            valor_atual = ws.cell(r, 11).value
            if not isinstance(valor_original, (int, float)) or not isinstance(valor_atual, (int, float)):
                continue
            deprec_pct = round((float(valor_original) - valor_atual) / float(valor_original), 6) if valor_original else 0
            ws_cache[f'L{r}'] = deprec_pct
        if ws_cache:
            cache_map[ws.title] = ws_cache
    return cache_map

def patch_formula_caches(output_path, sheetnames, cache_map):
    if not cache_map:
        return

    sheet_file_map = {f'xl/worksheets/sheet{i}.xml': name for i, name in enumerate(sheetnames, 1)}
    temp_output = f'{output_path}.tmp'

    with zipfile.ZipFile(output_path, 'r') as zin, zipfile.ZipFile(temp_output, 'w', zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            sheet_name = sheet_file_map.get(item.filename)
            if sheet_name and sheet_name in cache_map:
                root = ET.fromstring(data)
                for cell in root.iter(f'{{{SHEET_XML_NS}}}c'):
                    ref = cell.attrib.get('r')
                    cached_value = cache_map[sheet_name].get(ref)
                    if cached_value is None:
                        continue
                    value_node = cell.find(f'{{{SHEET_XML_NS}}}v')
                    if value_node is None:
                        value_node = ET.SubElement(cell, f'{{{SHEET_XML_NS}}}v')
                    value_node.text = str(cached_value)
                data = ET.tostring(root, encoding='utf-8', xml_declaration=False)
            zout.writestr(item, data)

    os.replace(temp_output, output_path)

def make_lots(rows):
    groups = defaultdict(list)
    for r in rows:
        if r['categoria'] == 'CABO' and r['status'] == 'DISPONIVEL':
            groups[(r['marca'],r['modelo'],r['subcategoria'])].append(r)
    lots = []
    for i, ((marca,modelo,subcat), items) in enumerate(sorted(groups.items()), 1):
        nome_base = display_name(marca, modelo)
        nome = f"{marca} {nome_base}".strip() if marca else nome_base
        lots.append({'codigo':f"MMD-LOT-{i:03d}", 'descricao':f"Lote {len(items)}x {nome}",
                     'subcategoria':subcat, 'quantidade':len(items), 'status':'DISPONIVEL'})
    return lots

# ── MANUAL TAB ────────────────────────────────────────────────
def build_manual(wb):
    ws = wb.active
    ws.title = 'MANUAL'
    ws.sheet_properties.tabColor = DARK
    W(ws, [3, 16, 50, 3])

    def big_section(row, title, bg_c=DARK, tx_c='F8FAFC'):
        ws.merge_cells(f'B{row}:C{row}')
        c = ws.cell(row=row, column=2)
        c.value = f'  {title}'
        c.font = FN(14, True, tx_c)
        c.fill = BG(bg_c)
        c.alignment = Alignment(vertical='center')
        ws.cell(row=row, column=3).fill = BG(bg_c)
        ws.row_dimensions[row].height = 40

    def line(row, text, color=MID, bold=False):
        ws.merge_cells(f'B{row}:C{row}')
        c = ws.cell(row=row, column=2)
        c.value = f'  {text}'
        c.font = FN(10, bold, color)
        c.alignment = Alignment(vertical='center')
        ws.row_dimensions[row].height = 24

    def badge(row, label, desc, bg_c, tx_c):
        c1 = ws.cell(row=row, column=2)
        c1.value = f'  {label}'
        c1.font = FN(11, True, tx_c)
        c1.fill = BG(bg_c)
        c1.border = BD()
        c1.alignment = Alignment(vertical='center')
        c2 = ws.cell(row=row, column=3)
        c2.value = f'  {desc}'
        c2.font = FN(10, False, MID)
        c2.alignment = Alignment(vertical='center')
        ws.row_dimensions[row].height = 28

    r = 2
    # Hero header
    ws.merge_cells(f'B{r}:C{r}')
    c = ws.cell(row=r, column=2)
    c.value = 'MMD EVENTOS'
    c.font = FN(24, True, ACCENT)
    ws.row_dimensions[r].height = 50
    r += 1
    ws.merge_cells(f'B{r}:C{r}')
    ws.cell(row=r, column=2).value = 'INVENTARIO INTELIGENTE  —  MANUAL DE USO'
    ws.cell(row=r, column=2).font = FN(12, True, LIGHT)
    ws.row_dimensions[r].height = 28
    r += 1
    ws.merge_cells(f'B{r}:C{r}')
    ws.cell(row=r, column=2).value = f'  Atualizado em {datetime.date.today().strftime("%d/%m/%Y")}'
    ws.cell(row=r, column=2).font = FN(9, False, SUBTLE)
    r += 2

    # NAVEGACAO
    big_section(r, 'NAVEGACAO'); r += 1
    line(r, 'Cada aba na parte inferior corresponde a uma categoria de equipamento.'); r += 1
    line(r, 'DASHBOARD mostra resumo geral com graficos e numeros.'); r += 1
    line(r, 'Use os FILTROS no header de cada coluna para buscar itens.'); r += 1
    line(r, 'LOTES agrupa cabos em kits. FORA DE OPERACAO = vendidos/emprestados.'); r += 2

    # STATUS
    big_section(r, 'STATUS — Onde o item esta', '1E293B', 'F8FAFC'); r += 1
    for st, desc in [('DISPONIVEL','No galpao, pronto para uso'),
                      ('EM_CAMPO','Saiu para evento, em uso'),
                      ('PACKED','Separado para evento, ainda no galpao'),
                      ('MANUTENCAO','Em reparo ou revisao'),
                      ('EMPRESTADO','Com terceiro (retorna)'),
                      ('VENDIDO','Vendido, fora do estoque'),
                      ('BAIXA','Descartado ou perdido')]:
        badge(r, st, desc, STATUS_STYLE[st]['bg'], STATUS_STYLE[st]['tx']); r += 1
    r += 1

    # ESTADO
    big_section(r, 'ESTADO — Ciclo de vida do equipamento', '1E293B', 'F8FAFC'); r += 1
    for est, desc in [('NOVO','Nunca usado em evento. Embalagem original.'),
                       ('SEMI_NOVO','Usado 1 a 5 vezes. Sem marcas visiveis.'),
                       ('USADO','Uso regular. Marcas normais de operacao.'),
                       ('RECONDICIONADO','Foi reparado ou restaurado. Funcional.')]:
        badge(r, est, desc, ESTADO_STYLE[est]['bg'], ESTADO_STYLE[est]['tx']); r += 1
    r += 1

    # DESGASTE
    big_section(r, 'DESGASTE — Condicao fisica (1 a 5)', '1E293B', 'F8FAFC'); r += 1
    for d, label, desc in [(5,f'{stars(5)} EXCELENTE','Como novo, sem marcas'),
                            (4,f'{stars(4)} BOM','Marcas leves, 100% funcional'),
                            (3,f'{stars(3)} REGULAR','Uso visivel, funcional'),
                            (2,f'{stars(2)} DESGASTADO','Problemas cosmeticos'),
                            (1,f'{stars(1)} CRITICO','Precisa reparo/trocar')]:
        badge(r, label, desc, DESGASTE_BG[d], DARK); r += 1
    r += 1

    # DEPRECIACAO
    big_section(r, 'DEPRECIACAO — Calculo automatico', ACCENT, 'F8FAFC'); r += 1
    line(r, 'Valor Atual prioriza o levantamento externo quando existir. Sem isso, sai uma sugestao inicial para ajuste manual.', MID, True); r += 1
    line(r, 'Deprec.% = 100% - (Valor Atual / Valor Original)', MID); r += 1
    line(r, 'Mudar Estado ou Desgaste nao altera a depreciacao. So Valor Atual altera.', MID); r += 1
    line(r, 'Estado e desgaste iniciais ja vem sugeridos por item, usando tipo, modelo e observacoes.', LIGHT); r += 1
    line(r, 'Valor Atual aparece em verde para revisao rapida. Deprec.% acompanha esse valor.', ACCENT, True); r += 1
    line(r, 'Serial Fab. e Subcategoria saem padronizados para leitura mais limpa.', LIGHT); r += 2

    # COMO ATUALIZAR
    big_section(r, 'COMO ATUALIZAR O INVENTARIO'); r += 1
    line(r, f'ITEM NOVO: adicionar na aba da categoria. Estado=NOVO, Desgaste={stars(5)}.', MID, True); r += 1
    line(r, 'APOS EVENTO: revisar desgaste dos itens que foram para o campo.'); r += 1
    line(r, 'DEFEITO: mudar status para MANUTENCAO e ajustar desgaste.'); r += 1
    line(r, 'VENDIDO: mover para aba FORA DE OPERACAO com status VENDIDO.'); r += 2

    # CATEGORIAS
    big_section(r, 'CATEGORIAS DE EQUIPAMENTO'); r += 1
    for cat in CAT_ORDER:
        d = CAT_DESIGN[cat]
        badge(r, f"{cat} ({d['prefix']})", d['desc'], d['bg'], d['tx']); r += 1

    # Footer
    r += 2
    add_footer(ws, r, 3)

# ── DASHBOARD TAB ─────────────────────────────────────────────
def build_dashboard(wb, activos, lots):
    ws = wb.create_sheet('DASHBOARD')
    ws.sheet_properties.tabColor = ACCENT
    W(ws, [3, 18, 14, 14, 14, 14, 14, 3, 3, 18, 14, 14, 14])
    ws.column_dimensions['N'].hidden = True

    total = len(activos)
    disponivel = sum(1 for s in activos if s['status']=='DISPONIVEL')
    em_campo = sum(1 for s in activos if s['status']=='EM_CAMPO')
    manut = sum(1 for s in activos if s['status']=='MANUTENCAO')
    val_orig = sum(s['valor_mercado'] or 0 for s in activos)
    val_atual = sum(current_value_for_item(s) or 0 for s in activos)
    degs = [s['desgaste'] for s in activos]
    avg_d = sum(degs)/len(degs) if degs else 0
    criticos = sum(1 for d in degs if d <= 2)
    deprec_pct = ((val_orig - val_atual)/val_orig*100) if val_orig > 0 else 0

    # Title
    ws.merge_cells('B2:G2')
    ws['B2'].value = 'MMD EVENTOS'
    ws['B2'].font = FN(22, True, DARK)
    ws.row_dimensions[2].height = 38
    ws.merge_cells('B3:G3')
    ws['B3'].value = f'Inventario Inteligente  ·  {datetime.date.today().strftime("%d/%m/%Y")}'
    ws['B3'].font = FN(10, False, SUBTLE)

    # KPI Cards (with background)
    ws.row_dimensions[5].height = 36
    ws.row_dimensions[6].height = 16
    ws.row_dimensions[7].height = 4
    kpi_card(ws, 5, 2, total, 'TOTAL ITENS', ACCENT)
    kpi_card(ws, 5, 4, disponivel, 'DISPONIVEL', '16A34A')
    kpi_card(ws, 5, 6, em_campo, 'EM CAMPO', 'CA8A04')

    ws.row_dimensions[8].height = 36
    ws.row_dimensions[9].height = 16
    kpi_card(ws, 8, 2, manut, 'MANUTENCAO', 'DC2626')
    kpi_card(ws, 8, 4, val_orig, 'VALOR ORIGINAL', ACCENT, 'R$ #,##0')
    kpi_card(ws, 8, 6, val_atual, 'VALOR ATUAL', '16A34A', 'R$ #,##0')

    # Health row
    ws.row_dimensions[10].height = 4
    ws.merge_cells('B11:G11')
    ws['B11'].value = 'SAUDE DO PATRIMONIO'
    ws['B11'].font = FN(9, True, SUBTLE)
    ws.row_dimensions[12].height = 30
    ws.row_dimensions[13].height = 16
    health_data = [
        (stars(round(avg_d)), 'DESGASTE', ACCENT),
        (str(criticos), 'CRITICOS', 'DC2626'),
        (f'{deprec_pct:.0f}%', 'DEPREC.', 'CA8A04'),
        (str(len(lots)), 'LOTES', 'F59E0B'),
    ]
    for i, (val, lbl, clr) in enumerate(health_data):
        col = 2 + i
        ws.cell(row=12, column=col).value = val
        ws.cell(row=12, column=col).font = FN(14, True, clr)
        ws.cell(row=12, column=col).alignment = Alignment(horizontal='center')
        ws.cell(row=13, column=col).value = lbl
        ws.cell(row=13, column=col).font = FN(8, True, SUBTLE)
        ws.cell(row=13, column=col).alignment = Alignment(horizontal='center')

    # ── LEFT: POR CATEGORIA (table + numbers) ──
    r = 15
    ws.cell(row=r, column=2).value = 'POR CATEGORIA'
    ws.cell(row=r, column=2).font = FN(9, True, SUBTLE)
    r += 1
    for i, h in enumerate(['Categoria', 'Qtd', 'Valor Orig.', 'Valor Atual', '% Patr.']):
        cl = ws.cell(row=r, column=2+i)
        cl.value = h
        cl.font = FN(9, True, 'F8FAFC')
        cl.fill = BG(DARK)
        cl.alignment = Alignment(horizontal='center', vertical='center')
        cl.border = BD(DARK)
    ws.row_dimensions[r].height = 26
    r += 1

    cat_data = defaultdict(lambda: {'count':0,'original':0,'current':0})
    for s in activos:
        cat_data[s['categoria']]['count'] += 1
        cat_data[s['categoria']]['original'] += s['valor_mercado'] or 0
        cat_data[s['categoria']]['current'] += current_value_for_item(s) or 0

    cat_start = r
    for cat in CAT_ORDER:
        d = cat_data.get(cat, {'count':0,'original':0,'current':0})
        ws.cell(row=r, column=2).value = cat
        ws.cell(row=r, column=2).font = FN(10, True, CAT_DESIGN[cat]['tx'])
        ws.cell(row=r, column=3).value = d['count']
        ws.cell(row=r, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=3).font = FN(10, False, MID)
        ws.cell(row=r, column=4).value = d['original']
        ws.cell(row=r, column=4).number_format = 'R$ #,##0'
        ws.cell(row=r, column=4).font = FN(10, False, MID)
        ws.cell(row=r, column=5).value = d['current']
        ws.cell(row=r, column=5).number_format = 'R$ #,##0'
        ws.cell(row=r, column=5).font = FN(10, True, '166534')
        pct = d['current']/val_atual if val_atual else 0
        ws.cell(row=r, column=6).value = pct
        ws.cell(row=r, column=6).number_format = '0%'
        ws.cell(row=r, column=6).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=6).font = FN(10, False, MID)
        ws.cell(row=r, column=14).value = f'{cat} {format_brl_label(d["current"])}'
        drow(ws, r, 6, alt=(r%2==0))
        ws.cell(row=r, column=2).font = FN(10, True, CAT_DESIGN[cat]['tx'])
        ws.cell(row=r, column=3).font = FN(10, False, MID)
        ws.cell(row=r, column=4).font = FN(10, False, MID)
        ws.cell(row=r, column=5).font = FN(10, True, '166534')
        ws.cell(row=r, column=6).font = FN(10, False, MID)
        r += 1
    cat_end = r - 1

    # Pie Chart (positioned to the RIGHT, no overlap)
    pie = PieChart()
    pie.style = 10
    pie.title = 'Distribuicao do Patrimonio Atual'
    pie.width = 14
    pie.height = 10
    cats_ref = Reference(ws, min_col=14, min_row=cat_start, max_row=cat_end)
    vals_ref = Reference(ws, min_col=5, min_row=cat_start, max_row=cat_end)
    pie.add_data(vals_ref)
    pie.set_categories(cats_ref)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showCatName = True
    pie.dataLabels.showVal = False
    pie.dataLabels.showLeaderLines = True
    # Color each slice
    cat_colors = [CAT_DESIGN[c]['tab'] for c in CAT_ORDER]
    for idx, color in enumerate(cat_colors):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = color
        pie.series[0].data_points.append(pt)
    ws.add_chart(pie, 'G15')

    # ── STATUS TABLE (below category table) ──
    r += 1
    ws.cell(row=r, column=2).value = 'POR STATUS'
    ws.cell(row=r, column=2).font = FN(9, True, SUBTLE)
    r += 1
    for i, h in enumerate(['Status', 'Qtd', '%']):
        cl = ws.cell(row=r, column=2+i)
        cl.value = h
        cl.font = FN(9, True, 'F8FAFC')
        cl.fill = BG(DARK)
        cl.alignment = Alignment(horizontal='center', vertical='center')
        cl.border = BD(DARK)
    ws.row_dimensions[r].height = 26
    r += 1

    status_counts = defaultdict(int)
    for s in activos: status_counts[s['status']] += 1

    for st in ['DISPONIVEL','EM_CAMPO','PACKED','MANUTENCAO']:
        cnt = status_counts.get(st, 0)
        ws.cell(row=r, column=2).value = st
        sty = STATUS_STYLE.get(st, {})
        ws.cell(row=r, column=2).font = FN(10, True, sty.get('tx', MID))
        ws.cell(row=r, column=2).fill = BG(sty.get('bg', BG_WHITE))
        ws.cell(row=r, column=2).border = BD()
        ws.cell(row=r, column=3).value = cnt
        ws.cell(row=r, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=3).font = FN(10, False, MID)
        ws.cell(row=r, column=3).border = BD()
        ws.cell(row=r, column=4).value = cnt/total if total else 0
        ws.cell(row=r, column=4).number_format = '0%'
        ws.cell(row=r, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=4).font = FN(10, False, MID)
        ws.cell(row=r, column=4).border = BD()
        r += 1

    # ── TOP 10 (below status, clear separation) ──
    r += 1
    ws.cell(row=r, column=2).value = 'TOP 10 MAIS VALIOSOS'
    ws.cell(row=r, column=2).font = FN(9, True, SUBTLE)
    r += 1
    for i, h in enumerate(['Nome','Categoria','Valor (R$)','Desgaste','Valor Atual']):
        cl = ws.cell(row=r, column=2+i)
        cl.value = h
        cl.font = FN(9, True, 'F8FAFC')
        cl.fill = BG(DARK)
        cl.alignment = Alignment(horizontal='center', vertical='center')
        cl.border = BD(DARK)
    ws.row_dimensions[r].height = 26
    r += 1

    top10 = sorted([s for s in activos if s['valor_mercado']], key=lambda x: -(x['valor_mercado'] or 0))[:10]
    for idx, s in enumerate(top10):
        nome = f"{s['marca']} {display_name(s['marca'], s['modelo'])}".strip() if s['marca'] else display_name('', s['modelo'])
        v_atual = current_value_for_item(s) or 0
        ws.cell(row=r, column=2).value = nome
        ws.cell(row=r, column=2).font = FN(10, True, MID)
        ws.cell(row=r, column=3).value = s['categoria']
        ws.cell(row=r, column=3).font = FN(9, False, CAT_DESIGN[s['categoria']]['tx'])
        ws.cell(row=r, column=4).value = s['valor_mercado']
        ws.cell(row=r, column=4).number_format = 'R$ #,##0'
        ws.cell(row=r, column=4).font = FN(10, False, MID)
        ws.cell(row=r, column=5).value = stars(s['desgaste'])
        ws.cell(row=r, column=5).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=5).fill = BG(DESGASTE_BG.get(s['desgaste'],'FEF9C3'))
        ws.cell(row=r, column=5).font = FN(10, True, DARK)
        ws.cell(row=r, column=6).value = v_atual
        ws.cell(row=r, column=6).number_format = 'R$ #,##0'
        ws.cell(row=r, column=6).font = FN(10, True, '16A34A')
        drow(ws, r, 6, alt=(idx%2==0))
        ws.cell(row=r, column=2).font = FN(10, True, MID)
        ws.cell(row=r, column=3).font = FN(9, False, CAT_DESIGN[s['categoria']]['tx'])
        ws.cell(row=r, column=5).fill = BG(DESGASTE_BG.get(s['desgaste'],'FEF9C3'))
        ws.cell(row=r, column=5).font = FN(10, True, DARK)
        ws.cell(row=r, column=6).font = FN(10, True, '16A34A')
        r += 1

    # ── PERDA PATRIMONIAL POR CATEGORIA (table + bar chart) ──
    r += 1
    ws.cell(row=r, column=2).value = 'PERDA PATRIMONIAL POR CATEGORIA'
    ws.cell(row=r, column=2).font = FN(9, True, SUBTLE)
    r += 1
    for i, h in enumerate(['Categoria','Perda (R$)','Deprec.%']):
        cl = ws.cell(row=r, column=2+i)
        cl.value = h
        cl.font = FN(9, True, 'F8FAFC')
        cl.fill = BG(DARK)
        cl.alignment = Alignment(horizontal='center', vertical='center')
        cl.border = BD(DARK)
    ws.row_dimensions[r].height = 26
    r += 1
    loss_start = r

    loss_rows = []
    for cat in CAT_ORDER:
        original = cat_data.get(cat, {}).get('original', 0)
        current = cat_data.get(cat, {}).get('current', 0)
        loss = max(0, original - current)
        loss_pct = (loss / original) if original else 0
        loss_rows.append((cat, loss, loss_pct))
    loss_rows.sort(key=lambda item: item[1], reverse=True)

    for cat, loss, loss_pct in loss_rows:
        ws.cell(row=r, column=2).value = cat
        ws.cell(row=r, column=2).font = FN(10, True, CAT_DESIGN[cat]['tx'])
        ws.cell(row=r, column=3).value = loss
        ws.cell(row=r, column=3).number_format = 'R$ #,##0'
        ws.cell(row=r, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=3).font = FN(10, True, MID)
        ws.cell(row=r, column=4).value = loss_pct
        ws.cell(row=r, column=4).number_format = '0%'
        ws.cell(row=r, column=4).font = FN(10, False, 'CA8A04')
        for c in range(2, 5):
            ws.cell(row=r, column=c).border = BD()
        drow(ws, r, 4, alt=(r%2==0))
        ws.cell(row=r, column=2).font = FN(10, True, CAT_DESIGN[cat]['tx'])
        ws.cell(row=r, column=3).font = FN(10, True, MID)
        ws.cell(row=r, column=4).font = FN(10, False, 'CA8A04')
        r += 1

    # Footer
    r += 2
    add_footer(ws, r, 7)

    # Print
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1

# ── CATEGORY TABS ─────────────────────────────────────────────
COLS = ['Codigo','Nome','Marca','Subcategoria','Serial Fab.','Tag RFID','Status','Estado','Desgaste','Valor Original','Valor Atual','Deprec.%','Notas']
COL_W = [18, 34, 16, 20, 22, 16, 14, 16, 12, 14, 16, 10, 36]
NC = len(COLS)

def build_cat_tab(wb, cat, items):
    design = CAT_DESIGN[cat]
    ws = wb.create_sheet(cat)
    ws.sheet_properties.tabColor = design['tab']
    W(ws, COL_W)

    vals = [i['valor_mercado'] or 0 for i in items]
    degs = [i['desgaste'] for i in items]
    avg_d = sum(degs)/len(degs) if degs else 0

    ws.merge_cells(f'A1:{get_column_letter(NC)}1')
    ws['A1'].value = f"{design['label'].upper()}  —  {design['desc']}"
    ws['A1'].font = FN(14, True, design['tx'])
    ws['A1'].fill = BG(design['bg'])
    ws.row_dimensions[1].height = 38

    ws.merge_cells(f'A2:{get_column_letter(NC)}2')
    ws['A2'].value = f'{len(items)} itens   |   Valor: R$ {sum(vals):,.0f}   |   Desgaste medio: {avg_d:.1f}/5'
    ws['A2'].font = FN(10, False, LIGHT)
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 6

    for i, h in enumerate(COLS):
        ws.cell(row=4, column=i+1).value = h
    hdr(ws, 4, NC)
    ws.row_dimensions[4].height = 28
    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f'A4:{get_column_letter(NC)}4'

    by_sub = defaultdict(list)
    for item in items: by_sub[item['subcategoria']].append(item)

    r = 5
    data_idx = 0
    first_data_row = r
    for subcat in sorted(by_sub.keys()):
        # Separator
        ws.merge_cells(f'A{r}:{get_column_letter(NC)}{r}')
        ws.cell(row=r, column=1).value = f'  {subcat}  ({len(by_sub[subcat])})'
        ws.cell(row=r, column=1).font = FN(10, True, design['tx'])
        ws.cell(row=r, column=1).fill = BG(design['bg'])
        ws.cell(row=r, column=1).alignment = Alignment(vertical='center')
        for c in range(1, NC+1):
            ws.cell(row=r, column=c).fill = BG(design['bg'])
        ws.row_dimensions[r].height = 24
        r += 1

        sub_items = sorted(by_sub[subcat], key=lambda x: x['codigo'])
        for item in sub_items:
            nome = display_name(item['marca'], item['modelo'])
            valor_atual_inicial = current_value_for_item(item)
            # Codigo as hyperlink
            c_code = ws.cell(row=r, column=1)
            c_code.value = item['codigo']
            c_code.hyperlink = f"https://mmd.app/s/{item['codigo']}"
            ws.cell(row=r, column=2).value = nome
            ws.cell(row=r, column=3).value = item['marca']
            ws.cell(row=r, column=4).value = item['subcategoria']
            ws.cell(row=r, column=5).value = trunc_serial(item['serial_fabrica'])
            if item['serial_fabrica'] and len(item['serial_fabrica']) > 18:
                ws.cell(row=r, column=5).comment = Comment(item['serial_fabrica'], 'MMD')
            ws.cell(row=r, column=6).value = ''
            ws.cell(row=r, column=7).value = item['status']
            ws.cell(row=r, column=8).value = item['estado']
            ws.cell(row=r, column=9).value = stars(item['desgaste'])
            ws.cell(row=r, column=9).alignment = Alignment(horizontal='center')
            ws.cell(row=r, column=10).value = item['valor_mercado']
            if item['valor_mercado']:
                ws.cell(row=r, column=10).number_format = '#,##0'
                ws.cell(row=r, column=11).value = valor_atual_inicial
                ws.cell(row=r, column=11).number_format = '#,##0.00'
                ws.cell(row=r, column=12).value = f'=IF(J{r}>0,(J{r}-K{r})/J{r},"")'
                ws.cell(row=r, column=12).number_format = '0%'
                ws.cell(row=r, column=12).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=r, column=13).value = item['notas_short']
            ws.cell(row=r, column=13).alignment = Alignment(vertical='center', wrap_text=True)
            if item['notas_full']:
                ws.cell(row=r, column=13).comment = Comment(item['notas_full'], 'MMD')

            drow(ws, r, NC, alt=(data_idx%2==0))
            ws.cell(row=r, column=1).font = FN(10, True, design['tx'])
            ws.cell(row=r, column=2).font = FN(10, True, MID)
            ws.cell(row=r, column=9).font = FN(10, True, DARK)
            ws.cell(row=r, column=9).fill = BG(DESGASTE_BG.get(item['desgaste'], 'FEF9C3'))
            ws.cell(row=r, column=9).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=r, column=11).font = FN(10, True, VALUE_ATUAL_TX)
            ws.cell(row=r, column=11).fill = BG(VALUE_ATUAL_BG)
            ws.cell(row=r, column=11).alignment = Alignment(horizontal='right', vertical='center')
            ws.cell(row=r, column=12).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=r, column=13).alignment = Alignment(vertical='top', wrap_text=True)
            ws.row_dimensions[r].height = 36 if item['notas_short'] else 22
            data_idx += 1
            r += 1

    # Total row
    r += 1
    ws.merge_cells(f'A{r}:F{r}')
    ws.cell(row=r, column=1).value = f'  TOTAL: {len(items)} itens'
    ws.cell(row=r, column=1).font = FN(10, True, DARK)
    ws.cell(row=r, column=1).fill = BG(BG_ALT)
    ws.cell(row=r, column=1).alignment = Alignment(vertical='center')
    for c in range(1, NC+1):
        ws.cell(row=r, column=c).fill = BG(BG_ALT)
        ws.cell(row=r, column=c).border = BD()
    total_val = sum(i['valor_mercado'] or 0 for i in items)
    if total_val > 0:
        ws.cell(row=r, column=10).value = total_val
        ws.cell(row=r, column=10).number_format = 'R$ #,##0'
        ws.cell(row=r, column=10).font = FN(10, True, DARK)
    ws.row_dimensions[r].height = 26

    # Footer
    r += 2
    add_footer(ws, r, NC)

    last_data_row = max(r, first_data_row + 1)
    add_validations(ws, first_data_row, last_data_row + 200)
    add_cond_fmt(ws, first_data_row, last_data_row + 200)

    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.print_title_rows = '4:4'

# ── LOTES TAB ─────────────────────────────────────────────────
def build_lotes(wb, lots):
    ws = wb.create_sheet('LOTES')
    ws.sheet_properties.tabColor = 'F59E0B'
    headers = ['Codigo','Descricao','Subcategoria','Qtd','Status','Tag RFID','QR Code']
    W(ws, [18, 42, 18, 10, 14, 16, 32])
    for i, h in enumerate(headers):
        ws.cell(row=1, column=i+1).value = h
    hdr(ws, 1, len(headers))
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

    for idx, lot in enumerate(lots):
        r = idx + 2
        ws.cell(row=r, column=1).value = lot['codigo']
        ws.cell(row=r, column=2).value = lot['descricao']
        ws.cell(row=r, column=3).value = lot['subcategoria']
        ws.cell(row=r, column=4).value = lot['quantidade']
        ws.cell(row=r, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=5).value = lot['status']
        ws.cell(row=r, column=6).value = ''
        ws.cell(row=r, column=7).value = f"https://mmd.app/l/{lot['codigo']}"
        drow(ws, r, len(headers), alt=(idx%2==0))
        ws.cell(row=r, column=1).font = FN(10, True, '92400E')
        ws.row_dimensions[r].height = 22

    # Conditional formatting for status
    rng = f'E2:E{len(lots)+100}'
    for st, sty in STATUS_STYLE.items():
        ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=[f'"{st}"'], fill=BG(sty['bg']), font=FN(10, True, sty['tx'])))

# ── FORA DE OPERACAO TAB ──────────────────────────────────────
def build_fora(wb, items):
    ws = wb.create_sheet('FORA DE OPERACAO')
    ws.sheet_properties.tabColor = '64748B'
    headers = COLS + ['Motivo']
    W(ws, COL_W + [30])
    for i, h in enumerate(headers):
        ws.cell(row=1, column=i+1).value = h
    hdr(ws, 1, len(headers))
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = 'A2'

    for idx, item in enumerate(items):
        r = idx + 2
        nome = display_name(item['marca'], item['modelo'])
        valor_atual = current_value_for_item(item)
        ws.cell(row=r, column=1).value = item['codigo']
        ws.cell(row=r, column=2).value = nome
        ws.cell(row=r, column=3).value = item['marca']
        ws.cell(row=r, column=4).value = item['subcategoria']
        ws.cell(row=r, column=5).value = trunc_serial(item['serial_fabrica'])
        if item['serial_fabrica'] and len(item['serial_fabrica']) > 18:
            ws.cell(row=r, column=5).comment = Comment(item['serial_fabrica'], 'MMD')
        ws.cell(row=r, column=6).value = ''
        ws.cell(row=r, column=7).value = item['status']
        ws.cell(row=r, column=8).value = item['estado']
        ws.cell(row=r, column=9).value = stars(item['desgaste'])
        ws.cell(row=r, column=9).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=r, column=10).value = item['valor_mercado']
        if item['valor_mercado']:
            ws.cell(row=r, column=10).number_format = '#,##0'
            ws.cell(row=r, column=11).value = valor_atual
            ws.cell(row=r, column=11).number_format = '#,##0.00'
            ws.cell(row=r, column=12).value = f'=IF(J{r}>0,(J{r}-K{r})/J{r},"")'
            ws.cell(row=r, column=12).number_format = '0%'
            ws.cell(row=r, column=12).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=r, column=13).value = item['notas_short']
        ws.cell(row=r, column=13).alignment = Alignment(vertical='top', wrap_text=True)
        ws.cell(row=r, column=14).value = item['motivo']
        ws.cell(row=r, column=14).font = FN(10, False, 'DC2626')
        drow(ws, r, len(headers), alt=(idx%2==0))
        ws.cell(row=r, column=1).font = FN(10, True, '64748B')
        ws.cell(row=r, column=9).font = FN(10, True, DARK)
        ws.cell(row=r, column=9).fill = BG(DESGASTE_BG.get(item['desgaste'], 'FEF9C3'))
        ws.cell(row=r, column=9).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=r, column=11).font = FN(10, True, VALUE_ATUAL_TX)
        ws.cell(row=r, column=11).fill = BG(VALUE_ATUAL_BG)
        ws.cell(row=r, column=11).alignment = Alignment(horizontal='right', vertical='center')
        ws.cell(row=r, column=13).alignment = Alignment(vertical='top', wrap_text=True)
        ws.row_dimensions[r].height = 36 if item['notas_short'] else 22

    rng = f'G2:G{len(items)+100}'
    for st, sty in STATUS_STYLE.items():
        ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=[f'"{st}"'], fill=BG(sty['bg']), font=FN(10, True, sty['tx'])))

# ── MAIN ──────────────────────────────────────────────────────
if __name__ == '__main__':
    print('Lendo inventario original...')
    rows = read_data()
    print(f'  {len(rows)} itens lidos')

    rows = assign_codes(rows)
    price_map = load_price_map()
    applied_prices = apply_price_overrides(rows, price_map)
    if price_map:
        print(f'  {applied_prices} valores reaproveitados de valores-para-importacao.md')

    activos = [r for r in rows if r['status'] not in ('VENDIDO','EMPRESTADO','BAIXA')]
    fora = [r for r in rows if r['status'] in ('VENDIDO','EMPRESTADO','BAIXA')]
    print(f'  {len(activos)} ativos, {len(fora)} fora de operacao')

    # Filter CABO tab: only cables with serial or value
    cabo_especial = [r for r in activos if r['categoria']=='CABO' and (r['serial_fabrica'] or r['valor_mercado'])]
    cabo_generico = [r for r in activos if r['categoria']=='CABO' and not r['serial_fabrica'] and not r['valor_mercado']]
    print(f'  Cabos: {len(cabo_especial)} especiais (aba CABO), {len(cabo_generico)} genericos (so LOTES)')

    lots = make_lots(rows)
    print(f'  {len(lots)} lotes')

    print('\nConstruindo planilha v3...')
    wb = Workbook()
    wb.calculation.calcMode = 'auto'
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcOnSave = True

    build_manual(wb)
    print('  [1/12] MANUAL')

    build_dashboard(wb, activos, lots)
    print('  [2/12] DASHBOARD')

    for i, cat in enumerate(CAT_ORDER):
        if cat == 'CABO':
            cat_items = cabo_especial
        else:
            cat_items = [r for r in activos if r['categoria'] == cat]
        build_cat_tab(wb, cat, cat_items)
        print(f'  [{i+3}/12] {cat} ({len(cat_items)})')

    build_lotes(wb, lots)
    print(f'  [11/12] LOTES ({len(lots)})')

    build_fora(wb, fora)
    print(f'  [12/12] FORA DE OPERACAO ({len(fora)})')

    wb.save(OUTPUT)
    patch_formula_caches(OUTPUT, wb.sheetnames, build_formula_cache_map(wb))
    print(f'\nSalvo: {OUTPUT}')
    print('Pronto!')
