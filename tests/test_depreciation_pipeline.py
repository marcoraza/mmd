import importlib.util
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]


def load_module(relative_path: str, module_name: str):
    module_path = REPO_ROOT / relative_path
    spec = importlib.util.spec_from_file_location(module_name, module_path)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    sys.modules[module_name] = module
    spec.loader.exec_module(module)
    return module


class DepreciationPipelineTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.apply_module = load_module("scripts/apply-pricing-markdown.py", "apply_pricing_markdown_test")
        cls.export_module = load_module("scripts/export-pricing-markdown.py", "export_pricing_markdown_test")

    def test_exported_depreciation_is_loss_percentage(self):
        valor_atual, dep_pct = self.export_module.compute_current_value(950.0, "USADO", 2)

        self.assertEqual(valor_atual, 247.0)
        self.assertEqual(dep_pct, 74.0)

    def test_apply_rewrites_depreciation_as_formula_and_dashboard_loss(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            tmpdir_path = Path(tmpdir)
            workbook_path = tmpdir_path / "inventario-limpo.xlsx"
            markdown_path = tmpdir_path / "valores-para-importacao.md"

            wb = Workbook()
            ws = wb.active
            ws.title = "ILUMINACAO"
            dashboard = wb.create_sheet("DASHBOARD")
            dashboard["A1"] = "placeholder"

            headers = ["Codigo", "Nome", "Subcategoria", "Status", "Desgaste", "Valor Original", "Valor Atual", "Deprec.%"]
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=4, column=col_idx).value = header

            ws["A5"] = "MMD-ILU-0001"
            ws["B5"] = "Par Teste"
            ws["C5"] = "Par Led"
            ws["D5"] = "DISPONIVEL"
            ws["E5"] = "★★☆☆☆"
            ws["F5"] = 950.0
            ws["G5"] = 247.0
            ws["H5"] = ""
            wb.save(workbook_path)

            markdown_path.write_text(
                "\n".join(
                    [
                        "## CSV Para Importacao",
                        "```csv",
                        "Aba,Codigo,Nome,Subcategoria,Marca,Valor Original (R$),Valor Atual (R$),Deprec. (%),Metodo,Confianca,Fonte Titulo,Fonte URL",
                        "ILUMINACAO,MMD-ILU-0001,Par Teste,Par Led,,950.00,247.00,26.00,manual_web_curado,0.950,Fonte Teste,https://example.com/preco",
                        "```",
                    ]
                ),
                encoding="utf-8",
            )

            original_workbook_path = self.apply_module.WORKBOOK_PATH
            original_markdown_path = self.apply_module.MARKDOWN_PATH
            original_active_sheets = self.apply_module.ACTIVE_SHEETS
            try:
                self.apply_module.WORKBOOK_PATH = workbook_path
                self.apply_module.MARKDOWN_PATH = markdown_path
                self.apply_module.ACTIVE_SHEETS = ["ILUMINACAO"]

                result = self.apply_module.main()
                self.assertEqual(result, 0)
            finally:
                self.apply_module.WORKBOOK_PATH = original_workbook_path
                self.apply_module.MARKDOWN_PATH = original_markdown_path
                self.apply_module.ACTIVE_SHEETS = original_active_sheets

            updated_wb = load_workbook(workbook_path, data_only=False)
            updated_ws = updated_wb["ILUMINACAO"]
            updated_dashboard = updated_wb["DASHBOARD"]

            self.assertEqual(updated_ws["H5"].value, '=IF(F5>0,(F5-G5)/F5,"")')
            self.assertEqual(updated_dashboard["D12"].value, "74%")
            self.assertEqual(updated_dashboard["D8"].value, 950.0)
            self.assertEqual(updated_dashboard["F8"].value, 247.0)


if __name__ == "__main__":
    unittest.main()
