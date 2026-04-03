"""
MMD Inventario Inteligente v3.5
Max quality: star ratings, KPI cards, subtotals, hyperlinks, footer,
desgaste chart, conditional formatting, data validation.
"""
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, BarChart, BarChart3D, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from collections import defaultdict
import os, datetime

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT = os.path.join(SCRIPT_DIR, '..', 'data', 'inventario-original.xlsx')
OUTPUT = os.path.join(SCRIPT_DIR, '..', 'data', 'inventario-limpo.xlsx')

# ── MODERN PALETTE ────────────────────────────────────────────
DARK = '0F172A'       # slate-900
MID = '334155'        # slate-700
LIGHT = '475569'      # slate-600
SUBTLE = 'CBD5E1'     # slate-300
BG_ALT = 'F1F5F9'     # slate-100
BG_WHITE = 'FFFFFF'
ACCENT = '4F46E5'     # indigo-600
ACCENT_LIGHT = 'E0E7FF'  # indigo-100

CAT_DESIGN = {
    'ILUMINACAO': {'tab':'6366F1','bg':'EEF2FF','tx':'4338CA','prefix':'ILU','label':'Iluminacao','desc':'Par LED, Ribalta, Moving, Strobo, Laser'},
    'AUDIO':      {'tab':'10B981','bg':'ECFDF5','tx':'047857','prefix':'AUD','label':'Audio','desc':'Caixas, Subs, Mesas, Mics, DJs'},
    'CABO':       {'tab':'F59E0B','bg':'FFFBEB','tx':'92400E','prefix':'CAB','label':'Cabo','desc':'XLR, DMX, P10, AC, HDMI, USB'},
    'ENERGIA':    {'tab':'EF4444','bg':'FEF2F2','tx':'B91C1C','prefix':'ENE','label':'Energia','desc':'Reguas, Centopeias, Extensoes'},
    'ESTRUTURA':  {'tab':'3B82F6','bg':'EFF6FF','tx':'1D4ED8','prefix':'EST','label':'Estrutura','desc':'Tripes, Box Truss, Praticaveis'},
    'EFEITO':     {'tab':'8B5CF6','bg':'F5F3FF','tx':'6D28D9','prefix':'EFE','label':'Efeito','desc':'Fumaca, Haze, CO2, Fluidos'},
    'VIDEO':      {'tab':'14B8A6','bg':'F0FDFA','tx':'0F766E','prefix':'VID','label':'Video','desc':'Projetores, Notebooks, Tablets'},
    'ACESSORIO':  {'tab':'F97316','bg':'FFF7ED','tx':'C2410C','prefix':'ACE','label':'Acessorio','desc':'Cases, Ferramentas, Roteadores'},
}
CAT_ORDER = ['ILUMINACAO','AUDIO','CABO','ENERGIA','ESTRUTURA','EFEITO','VIDEO','ACESSORIO']

# Cores mais saturadas e modernas (nada pastel fraco)
STATUS_STYLE = {
    'DISPONIVEL': {'bg':'BBF7D0','tx':'14532D'},
    'PACKED':     {'bg':'BFDBFE','tx':'1E3A8A'},
    'EM_CAMPO':   {'bg':'FDE68A','tx':'78350F'},
    'MANUTENCAO': {'bg':'FECACA','tx':'7F1D1D'},
    'EMPRESTADO': {'bg':'DDD6FE','tx':'4C1D95'},
    'VENDIDO':    {'bg':'E2E8F0','tx':'475569'},
    'BAIXA':      {'bg':'E2E8F0','tx':'475569'},
}
ESTADO_STYLE = {
    'NOVO':           {'bg':'BBF7D0','tx':'14532D'},
    'SEMI_NOVO':      {'bg':'BFDBFE','tx':'1E3A8A'},
    'USADO':          {'bg':'FDE68A','tx':'78350F'},
    'RECONDICIONADO': {'bg':'FED7AA','tx':'7C2D12'},
}
DESGASTE_BG = {5:'86EFAC',4:'93C5FD',3:'FDE047',2:'FDBA74',1:'FCA5A5'}

BRAND_MAP = {
    'QSC':'QSC','SHOWTECH':'Showtech','BRIWAX':'Briwax','SENNHEISER':'Sennheiser',
    'BEHRINGER':'Behringer','PIONEER':'Pioneer','SHURE':'Shure','JBL':'JBL',
    'YAMAHA':'Yamaha','HOT MACHINE':'Hot Machine','PHENYX PRO':'Phenyx Pro',
    'WIRECONNEX':'Wireconnex','STANER AUDIO AMERICA':'Staner','RMV':'RMV',
    'BDL':'BDL','LEACS':'Leacs','SOG':'SOG','TOMATE':'Tomate',
    'MEGA LIGHT':'Mega Light','STAR LIGHT DIVISION':'Star Light Division',
    'SKYPIX':'Skypix','AURA':'Aura','CSR':'CSR','MACKIE':'Mackie','RCF':'RCF',
    'SOUNDBOX':'Soundbox','MARK AUDIO':'Mark Audio','NHL':'NHL',
    'CHINES':'','S/ MARCA':'','SEM MARCA':'','NONE':'','S/MARCA':'',
    'GATOR':'Gator','RODE':'Rode','BOSS':'Boss','ROLAND':'Roland',
}
CATEGORY_MAP = {
    'PAR LEDS':('ILUMINACAO','Par Led'),'PAR LEDS ':('ILUMINACAO','Par Led'),
    'RIBALTA':('ILUMINACAO','Ribalta'),'Ribalta':('ILUMINACAO','Ribalta'),
    'MINI MOVING':('ILUMINACAO','Mini Moving'),'Mini Moving':('ILUMINACAO','Mini Moving'),
    'MOVING BEAM':('ILUMINACAO','Moving Beam'),'Moving Beam':('ILUMINACAO','Moving Beam'),
    'MINI BRUTE':('ILUMINACAO','Mini Brute'),'OUTRAS LUZES':('ILUMINACAO','Outras Luzes'),
    'LUZ':('ILUMINACAO','Outras Luzes'),'LUZ NEGRA':('ILUMINACAO','Luz Negra'),
    'LASER':('ILUMINACAO','Laser'),'STROBO':('ILUMINACAO','Strobo'),
    'Iluminação':('ILUMINACAO','Geral'),
    'AUDIO':('AUDIO','Geral'),'Audio':('AUDIO','Geral'),'audio':('AUDIO','Geral'),
    'SOM':('AUDIO','Geral'),'Som':('AUDIO','Geral'),'som':('AUDIO','Geral'),
    'INSTRUMENTO':('AUDIO','Instrumento'),'CAPSULAS MIC':('AUDIO','Capsula Mic'),
    'MIDI':('AUDIO','Midi'),
    'CABO':('CABO','Geral'),'EXTENSOR':('CABO','Extensor'),
    'REGUA':('ENERGIA','Regua'),'ELETRICA':('ENERGIA','Eletrica'),
    'Eletrica':('ENERGIA','Eletrica'),'Energia':('ENERGIA','Geral'),'FONTE':('ENERGIA','Fonte'),
    'TRIPÉ':('ESTRUTURA','Tripe'),'SUPORTE':('ESTRUTURA','Suporte'),
    'Estrutura':('ESTRUTURA','Geral'),'PRATICÁVEL':('ESTRUTURA','Praticavel'),
    'Praticável':('ESTRUTURA','Praticavel'),
    'EFEITOS':('EFEITO','Geral'),'Efeitos':('EFEITO','Geral'),
    'FLUIDO':('EFEITO','Fluido'),'Fluido':('EFEITO','Fluido'),
    'GLOBO':('EFEITO','Globo'),'Globo':('EFEITO','Globo'),
    'VIDEO':('VIDEO','Geral'),'COMPUTADORES':('VIDEO','Computador'),
    'NOTEBOOK':('VIDEO','Notebook'),'Notebook':('VIDEO','Notebook'),
    'TABLET':('VIDEO','Tablet'),'Tablet':('VIDEO','Tablet'),
    'CASE':('ACESSORIO','Case'),'CASES':('ACESSORIO','Case'),
    'FERRAMENTAS':('ACESSORIO','Ferramenta'),
    'ROTEADOR':('ACESSORIO','Roteador'),'Roteador':('ACESSORIO','Roteador'),
    'REDE':('ACESSORIO','Rede'),'RADIO':('ACESSORIO','Radio'),'Radio':('ACESSORIO','Radio'),
    'Refrigeracao':('ACESSORIO','Refrigeracao'),'REFRIGERACAO':('ACESSORIO','Refrigeracao'),
}
ESTADO_FATOR = {'NOVO':1.0,'SEMI_NOVO':0.85,'USADO':0.65,'RECONDICIONADO':0.50}

# ── STYLE HELPERS ─────────────────────────────────────────────
def FN(sz=10, bold=False, color=MID, name='Arial'):
    return Font(name=name, size=sz, bold=bold, color=color)

def BG(color):
    return PatternFill(start_color=color, end_color=color, fill_type='solid')

def BD(color=SUBTLE):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(ws, row, ncols):
    for c in range(1, ncols+1):
        cl = ws.cell(row=row, column=c)
        cl.font = FN(10, True, 'F8FAFC')
        cl.fill = BG(DARK)
        cl.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cl.border = BD(DARK)

def drow(ws, row, ncols, alt=False):
    bg = BG(BG_ALT) if alt else BG(BG_WHITE)
    for c in range(1, ncols+1):
        cl = ws.cell(row=row, column=c)
        cl.fill = bg
        cl.border = BD()
        cl.alignment = Alignment(vertical='center')
        cl.font = FN(10)

def W(ws, ws_list):
    for i, w in enumerate(ws_list, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def norm_brand(raw):
    if not raw: return ''
    key = raw.strip().upper()
    if key in BRAND_MAP: return BRAND_MAP[key]
    if key in ('','NONE'): return ''
    words = raw.strip().split()
    return ' '.join(w.upper() if len(w) <= 3 else w.title() for w in words)

def trunc(text, mx=60):
    if not text or text == 'None': return '', ''
    text = text.strip()
    if len(text) <= mx: return text, ''
    return text[:mx-3] + '...', text

def trunc_serial(text, mx=18):
    if not text or text == 'None': return ''
    text = text.strip()
    if len(text) <= mx: return text
    return text[:mx-3] + '...'

def stars(n):
    return '\u2605' * n + '\u2606' * (5 - n)  # ★★★☆☆

def kpi_card(ws, row, col, value, label, color, fmt='#,##0', span=2):
    """Render a KPI card with colored background block."""
    # Background block
    for c in range(col, col + span):
        for r in [row, row+1, row+2]:
            ws.cell(row=r, column=c).fill = BG(ACCENT_LIGHT)
            ws.cell(row=r, column=c).border = BD('C7D2FE')
    # Value
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    cv = ws.cell(row=row, column=col)
    cv.value = value
    cv.font = FN(18, True, color)
    cv.number_format = fmt
    cv.alignment = Alignment(horizontal='center', vertical='center')
    # Label
    ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+span-1)
    cl = ws.cell(row=row+1, column=col)
    cl.value = label
    cl.font = FN(8, True, LIGHT)
    cl.alignment = Alignment(horizontal='center', vertical='center')

def add_footer(ws, row, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1)
    c.value = f'MMD Eventos  |  Inventario Inteligente  |  {datetime.date.today().strftime("%d/%m/%Y")}'
    c.font = FN(8, False, SUBTLE)
    c.alignment = Alignment(horizontal='center')
    ws.row_dimensions[row].height = 18

# ── DATA VALIDATION + CONDITIONAL FORMATTING ──────────────────
def add_validations(ws, start_row, end_row, status_col='G', estado_col='H', desgaste_col='I'):
    dv_status = DataValidation(type='list', formula1='"DISPONIVEL,PACKED,EM_CAMPO,MANUTENCAO,EMPRESTADO,VENDIDO,BAIXA"', allow_blank=True)
    dv_status.error = 'Selecione um status valido'
    dv_status.errorTitle = 'Status invalido'
    dv_estado = DataValidation(type='list', formula1='"NOVO,SEMI_NOVO,USADO,RECONDICIONADO"', allow_blank=True)
    dv_desgaste = DataValidation(type='list', formula1='"1,2,3,4,5"', allow_blank=True)
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_estado)
    ws.add_data_validation(dv_desgaste)
    dv_status.add(f'{status_col}{start_row}:{status_col}{end_row}')
    dv_estado.add(f'{estado_col}{start_row}:{estado_col}{end_row}')
    dv_desgaste.add(f'{desgaste_col}{start_row}:{desgaste_col}{end_row}')

def add_cond_fmt(ws, start_row, end_row, status_col='G', estado_col='H', desgaste_col='I', deprec_col='L'):
    rng_s = f'{status_col}{start_row}:{status_col}{end_row}'
    for st, sty in STATUS_STYLE.items():
        ws.conditional_formatting.add(rng_s, CellIsRule(operator='equal', formula=[f'"{st}"'], fill=BG(sty['bg']), font=FN(10, True, sty['tx'])))
    rng_e = f'{estado_col}{start_row}:{estado_col}{end_row}'
    for est, sty in ESTADO_STYLE.items():
        ws.conditional_formatting.add(rng_e, CellIsRule(operator='equal', formula=[f'"{est}"'], fill=BG(sty['bg']), font=FN(10, False, sty['tx'])))
    rng_d = f'{desgaste_col}{start_row}:{desgaste_col}{end_row}'
    for val, bg_c in DESGASTE_BG.items():
        ws.conditional_formatting.add(rng_d, CellIsRule(operator='equal', formula=[f'{val}'], fill=BG(bg_c), font=FN(10, True, DARK)))
    rng_p = f'{deprec_col}{start_row}:{deprec_col}{end_row}'
    ws.conditional_formatting.add(rng_p, CellIsRule(operator='lessThan', formula=['0.2'], fill=BG('BBF7D0'), font=FN(10, True, '14532D')))
    ws.conditional_formatting.add(rng_p, CellIsRule(operator='between', formula=['0.2', '0.4'], fill=BG('FDE68A'), font=FN(10, True, '78350F')))
    ws.conditional_formatting.add(rng_p, CellIsRule(operator='between', formula=['0.4', '0.6'], fill=BG('FDBA74'), font=FN(10, True, '7C2D12')))
    ws.conditional_formatting.add(rng_p, CellIsRule(operator='greaterThanOrEqual', formula=['0.6'], fill=BG('FCA5A5'), font=FN(10, True, '7F1D1D')))

# ── READ DATA ─────────────────────────────────────────────────
def read_data():
    wb = load_workbook(INPUT, data_only=True)
    ws = wb['EQUIPAMENTOS']
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=False):
        cat_raw = str(row[2].value).strip() if row[2].value else ''
        modelo = str(row[5].value).strip() if row[5].value else ''
        if not cat_raw and not modelo: continue
        ok = row[1].value
        subcat_raw = str(row[3].value).strip() if row[3].value else ''
        marca_raw = str(row[4].value).strip() if row[4].value else ''
        serial_fab = str(row[9].value).strip() if row[9].value else ''
        valor = row[10].value
        notas_raw = str(row[13].value).strip() if len(row) > 13 and row[13].value else ''

        mapped = CATEGORY_MAP.get(cat_raw, ('ACESSORIO','Outro'))
        cat, subcat = mapped
        if subcat_raw and subcat_raw.upper() not in ('SEM SUB','NONE',''):
            subcat = subcat_raw.title()

        has_serial = serial_fab and 'S/ SERIAL' not in serial_fab.upper() and 'SEM' not in serial_fab.upper() and serial_fab.upper() != 'NONE'
        valor_num = None
        if valor:
            try: valor_num = float(valor)
            except: pass

        marca = norm_brand(marca_raw)
        ok_str = str(ok).strip().upper() if ok else ''
        notas_up = notas_raw.upper() if notas_raw else ''

        if 'VENDID' in ok_str or 'VENDID' in notas_up:
            status, motivo = 'VENDIDO', notas_raw or 'Vendido'
        elif 'BIGODE' in ok_str:
            status, motivo = 'EMPRESTADO', 'Emprestado - Seu Bigode'
        elif 'QUEBRAD' in ok_str:
            status, motivo = 'BAIXA', 'Quebrado'
        else:
            status, motivo = 'DISPONIVEL', ''

        notas_short, notas_full = trunc(notas_raw)

        rows.append({
            'categoria':cat, 'subcategoria':subcat, 'marca':marca,
            'modelo':modelo, 'serial_fabrica':serial_fab if has_serial else '',
            'valor_mercado':valor_num, 'status':status, 'motivo':motivo,
            'estado':'USADO', 'desgaste':3,
            'notas_short':notas_short, 'notas_full':notas_full,
        })
    return rows

def assign_codes(rows):
    counters = defaultdict(int)
    for r in rows:
        counters[r['categoria']] += 1
        pfx = CAT_DESIGN[r['categoria']]['prefix']
        r['codigo'] = f"MMD-{pfx}-{counters[r['categoria']]:04d}"
    return rows

def make_lots(rows):
    groups = defaultdict(list)
    for r in rows:
        if r['categoria'] == 'CABO' and r['status'] == 'DISPONIVEL':
            groups[(r['marca'],r['modelo'],r['subcategoria'])].append(r)
    lots = []
    for i, ((marca,modelo,subcat), items) in enumerate(sorted(groups.items()), 1):
        nome = f"{marca} {modelo}".strip() if marca else modelo
        lots.append({'codigo':f"MMD-LOT-{i:03d}", 'descricao':f"Lote {len(items)}x {nome}",
                     'subcategoria':subcat, 'quantidade':len(items), 'status':'DISPONIVEL'})
    return lots

# ── MANUAL TAB ────────────────────────────────────────────────
def build_manual(wb):
    ws = wb.active
    ws.title = 'MANUAL'
    ws.sheet_properties.tabColor = DARK
    W(ws, [3, 16, 50, 3])

    def big_section(row, title, bg_c=DARK, tx_c='F8FAFC'):
        ws.merge_cells(f'B{row}:C{row}')
        c = ws.cell(row=row, column=2)
        c.value = f'  {title}'
        c.font = FN(14, True, tx_c)
        c.fill = BG(bg_c)
        c.alignment = Alignment(vertical='center')
        ws.cell(row=row, column=3).fill = BG(bg_c)
        ws.row_dimensions[row].height = 40

    def line(row, text, color=MID, bold=False):
        ws.merge_cells(f'B{row}:C{row}')
        c = ws.cell(row=row, column=2)
        c.value = f'  {text}'
        c.font = FN(10, bold, color)
        c.alignment = Alignment(vertical='center')
        ws.row_dimensions[row].height = 24

    def badge(row, label, desc, bg_c, tx_c):
        c1 = ws.cell(row=row, column=2)
        c1.value = f'  {label}'
        c1.font = FN(11, True, tx_c)
        c1.fill = BG(bg_c)
        c1.border = BD()
        c1.alignment = Alignment(vertical='center')
        c2 = ws.cell(row=row, column=3)
        c2.value = f'  {desc}'
        c2.font = FN(10, False, MID)
        c2.alignment = Alignment(vertical='center')
        ws.row_dimensions[row].height = 28

    r = 2
    # Hero header
    ws.merge_cells(f'B{r}:C{r}')
    c = ws.cell(row=r, column=2)
    c.value = 'MMD EVENTOS'
    c.font = FN(24, True, ACCENT)
    ws.row_dimensions[r].height = 50
    r += 1
    ws.merge_cells(f'B{r}:C{r}')
    ws.cell(row=r, column=2).value = 'INVENTARIO INTELIGENTE  —  MANUAL DE USO'
    ws.cell(row=r, column=2).font = FN(12, True, LIGHT)
    ws.row_dimensions[r].height = 28
    r += 1
    ws.merge_cells(f'B{r}:C{r}')
    ws.cell(row=r, column=2).value = f'  Atualizado em {datetime.date.today().strftime("%d/%m/%Y")}'
    ws.cell(row=r, column=2).font = FN(9, False, SUBTLE)
    r += 2

    # NAVEGACAO
    big_section(r, 'NAVEGACAO'); r += 1
    line(r, 'Cada aba na parte inferior corresponde a uma categoria de equipamento.'); r += 1
    line(r, 'DASHBOARD mostra resumo geral com graficos e numeros.'); r += 1
    line(r, 'Use os FILTROS no header de cada coluna para buscar itens.'); r += 1
    line(r, 'LOTES agrupa cabos em kits. FORA DE OPERACAO = vendidos/emprestados.'); r += 2

    # STATUS
    big_section(r, 'STATUS — Onde o item esta', '1E293B', 'F8FAFC'); r += 1
    for st, desc in [('DISPONIVEL','No galpao, pronto para uso'),
                      ('EM_CAMPO','Saiu para evento, em uso'),
                      ('PACKED','Separado para evento, ainda no galpao'),
                      ('MANUTENCAO','Em reparo ou revisao'),
                      ('EMPRESTADO','Com terceiro (retorna)'),
                      ('VENDIDO','Vendido, fora do estoque'),
                      ('BAIXA','Descartado ou perdido')]:
        badge(r, st, desc, STATUS_STYLE[st]['bg'], STATUS_STYLE[st]['tx']); r += 1
    r += 1

    # ESTADO
    big_section(r, 'ESTADO — Ciclo de vida do equipamento', '1E293B', 'F8FAFC'); r += 1
    for est, desc in [('NOVO','Nunca usado em evento. Embalagem original.'),
                       ('SEMI_NOVO','Usado 1 a 5 vezes. Sem marcas visiveis.'),
                       ('USADO','Uso regular. Marcas normais de operacao.'),
                       ('RECONDICIONADO','Foi reparado ou restaurado. Funcional.')]:
        badge(r, est, desc, ESTADO_STYLE[est]['bg'], ESTADO_STYLE[est]['tx']); r += 1
    r += 1

    # DESGASTE
    big_section(r, 'DESGASTE — Condicao fisica (1 a 5)', '1E293B', 'F8FAFC'); r += 1
    for d, label, desc in [(5,'5 EXCELENTE','Como novo, sem marcas'),
                            (4,'4 BOM','Marcas leves, 100% funcional'),
                            (3,'3 REGULAR','Uso visivel, funcional'),
                            (2,'2 DESGASTADO','Problemas cosmeticos'),
                            (1,'1 CRITICO','Precisa reparo/trocar')]:
        badge(r, label, desc, DESGASTE_BG[d], DARK); r += 1
    r += 1

    # DEPRECIACAO
    big_section(r, 'DEPRECIACAO — Calculo automatico', ACCENT, 'F8FAFC'); r += 1
    line(r, 'Valor Atual = Valor Original x (Desgaste / 5) x Fator do Estado', MID, True); r += 1
    line(r, 'Deprec.% = 100% - (Valor Atual / Valor Original)', MID); r += 1
    line(r, 'NOVO = 100%  |  SEMI_NOVO = 85%  |  USADO = 65%  |  RECONDICIONADO = 50%', LIGHT); r += 1
    line(r, 'Mude o Desgaste ou Estado e o Valor Atual recalcula automaticamente.', ACCENT, True); r += 2

    # COMO ATUALIZAR
    big_section(r, 'COMO ATUALIZAR O INVENTARIO'); r += 1
    line(r, 'ITEM NOVO: adicionar na aba da categoria. Estado=NOVO, Desgaste=5.', MID, True); r += 1
    line(r, 'APOS EVENTO: revisar desgaste dos itens que foram para o campo.'); r += 1
    line(r, 'DEFEITO: mudar status para MANUTENCAO e ajustar desgaste.'); r += 1
    line(r, 'VENDIDO: mover para aba FORA DE OPERACAO com status VENDIDO.'); r += 2

    # CATEGORIAS
    big_section(r, 'CATEGORIAS DE EQUIPAMENTO'); r += 1
    for cat in CAT_ORDER:
        d = CAT_DESIGN[cat]
        badge(r, f"{cat} ({d['prefix']})", d['desc'], d['bg'], d['tx']); r += 1

    # Footer
    r += 2
    add_footer(ws, r, 3)

# ── DASHBOARD TAB ─────────────────────────────────────────────
def build_dashboard(wb, activos, lots):
    ws = wb.create_sheet('DASHBOARD')
    ws.sheet_properties.tabColor = ACCENT
    W(ws, [3, 18, 14, 14, 14, 14, 14, 3, 3, 18, 14, 14, 14])

    total = len(activos)
    disponivel = sum(1 for s in activos if s['status']=='DISPONIVEL')
    em_campo = sum(1 for s in activos if s['status']=='EM_CAMPO')
    manut = sum(1 for s in activos if s['status']=='MANUTENCAO')
    val_orig = sum(s['valor_mercado'] or 0 for s in activos)
    val_atual = sum((s['valor_mercado'] or 0)*(s['desgaste']/5)*ESTADO_FATOR[s['estado']] for s in activos)
    degs = [s['desgaste'] for s in activos]
    avg_d = sum(degs)/len(degs) if degs else 0
    criticos = sum(1 for d in degs if d <= 2)
    deprec_pct = ((val_orig - val_atual)/val_orig*100) if val_orig > 0 else 0

    # Title
    ws.merge_cells('B2:G2')
    ws['B2'].value = 'MMD EVENTOS'
    ws['B2'].font = FN(22, True, DARK)
    ws.row_dimensions[2].height = 38
    ws.merge_cells('B3:G3')
    ws['B3'].value = f'Inventario Inteligente  ·  {datetime.date.today().strftime("%d/%m/%Y")}'
    ws['B3'].font = FN(10, False, SUBTLE)

    # KPI Cards (with background)
    ws.row_dimensions[5].height = 36
    ws.row_dimensions[6].height = 16
    ws.row_dimensions[7].height = 4
    kpi_card(ws, 5, 2, total, 'TOTAL ITENS', ACCENT)
    kpi_card(ws, 5, 4, disponivel, 'DISPONIVEL', '16A34A')
    kpi_card(ws, 5, 6, em_campo, 'EM CAMPO', 'CA8A04')

    ws.row_dimensions[8].height = 36
    ws.row_dimensions[9].height = 16
    kpi_card(ws, 8, 2, manut, 'MANUTENCAO', 'DC2626')
    kpi_card(ws, 8, 4, val_orig, 'VALOR ORIGINAL', ACCENT, 'R$ #,##0')
    kpi_card(ws, 8, 6, val_atual, 'VALOR ATUAL', '16A34A', 'R$ #,##0')

    # Health row
    ws.row_dimensions[10].height = 4
    ws.merge_cells('B11:G11')
    ws['B11'].value = 'SAUDE DO PATRIMONIO'
    ws['B11'].font = FN(9, True, SUBTLE)
    ws.row_dimensions[12].height = 30
    ws.row_dimensions[13].height = 16
    health_data = [
        (stars(round(avg_d)), 'DESGASTE', ACCENT),
        (str(criticos), 'CRITICOS', 'DC2626'),
        (f'{deprec_pct:.0f}%', 'DEPREC.', 'CA8A04'),
        (str(len(lots)), 'LOTES', 'F59E0B'),
    ]
    for i, (val, lbl, clr) in enumerate(health_data):
        col = 2 + i
        ws.cell(row=12, column=col).value = val
        ws.cell(row=12, column=col).font = FN(14, True, clr)
        ws.cell(row=12, column=col).alignment = Alignment(horizontal='center')
        ws.cell(row=13, column=col).value = lbl
        ws.cell(row=13, column=col).font = FN(8, True, SUBTLE)
        ws.cell(row=13, column=col).alignment = Alignment(horizontal='center')

    # ── LEFT: POR CATEGORIA (table + numbers) ──
    r = 15
    ws.cell(row=r, column=2).value = 'POR CATEGORIA'
    ws.cell(row=r, column=2).font = FN(9, True, SUBTLE)
    r += 1
    for i, h in enumerate(['Categoria', 'Qtd', 'Valor (R$)', '%']):
        cl = ws.cell(row=r, column=2+i)
        cl.value = h
        cl.font = FN(9, True, 'F8FAFC')
        cl.fill = BG(DARK)
        cl.alignment = Alignment(horizontal='center', vertical='center')
        cl.border = BD(DARK)
    ws.row_dimensions[r].height = 26
    r += 1

    cat_data = defaultdict(lambda: {'count':0,'value':0})
    for s in activos:
        cat_data[s['categoria']]['count'] += 1
        cat_data[s['categoria']]['value'] += s['valor_mercado'] or 0

    cat_start = r
    for cat in CAT_ORDER:
        d = cat_data.get(cat, {'count':0,'value':0})
        ws.cell(row=r, column=2).value = cat
        ws.cell(row=r, column=2).font = FN(10, True, CAT_DESIGN[cat]['tx'])
        ws.cell(row=r, column=3).value = d['count']
        ws.cell(row=r, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=3).font = FN(10, False, MID)
        ws.cell(row=r, column=4).value = d['value']
        ws.cell(row=r, column=4).number_format = '#,##0'
        ws.cell(row=r, column=4).font = FN(10, False, MID)
        pct = d['count']/total if total else 0
        ws.cell(row=r, column=5).value = pct
        ws.cell(row=r, column=5).number_format = '0%'
        ws.cell(row=r, column=5).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=5).font = FN(10, False, MID)
        drow(ws, r, 5, alt=(r%2==0))
        ws.cell(row=r, column=2).font = FN(10, True, CAT_DESIGN[cat]['tx'])
        ws.cell(row=r, column=3).font = FN(10, False, MID)
        ws.cell(row=r, column=4).font = FN(10, False, MID)
        ws.cell(row=r, column=5).font = FN(10, False, MID)
        r += 1
    cat_end = r - 1

    # Pie Chart (positioned to the RIGHT, no overlap)
    pie = PieChart()
    pie.style = 10
    pie.title = 'Distribuicao por Categoria'
    pie.width = 14
    pie.height = 10
    cats_ref = Reference(ws, min_col=2, min_row=cat_start, max_row=cat_end)
    vals_ref = Reference(ws, min_col=3, min_row=cat_start, max_row=cat_end)
    pie.add_data(vals_ref)
    pie.set_categories(cats_ref)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showCatName = True
    pie.dataLabels.showVal = True
    pie.dataLabels.showLeaderLines = True
    pie.dataLabels.numFmt = '#,##0'
    # Color each slice
    cat_colors = [CAT_DESIGN[c]['tab'] for c in CAT_ORDER]
    for idx, color in enumerate(cat_colors):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = color
        pie.series[0].data_points.append(pt)
    ws.add_chart(pie, 'G15')

    # ── STATUS TABLE (below category table) ──
    r += 1
    ws.cell(row=r, column=2).value = 'POR STATUS'
    ws.cell(row=r, column=2).font = FN(9, True, SUBTLE)
    r += 1
    for i, h in enumerate(['Status', 'Qtd', '%']):
        cl = ws.cell(row=r, column=2+i)
        cl.value = h
        cl.font = FN(9, True, 'F8FAFC')
        cl.fill = BG(DARK)
        cl.alignment = Alignment(horizontal='center', vertical='center')
        cl.border = BD(DARK)
    ws.row_dimensions[r].height = 26
    r += 1

    status_counts = defaultdict(int)
    for s in activos: status_counts[s['status']] += 1

    for st in ['DISPONIVEL','EM_CAMPO','PACKED','MANUTENCAO']:
        cnt = status_counts.get(st, 0)
        ws.cell(row=r, column=2).value = st
        sty = STATUS_STYLE.get(st, {})
        ws.cell(row=r, column=2).font = FN(10, True, sty.get('tx', MID))
        ws.cell(row=r, column=2).fill = BG(sty.get('bg', BG_WHITE))
        ws.cell(row=r, column=2).border = BD()
        ws.cell(row=r, column=3).value = cnt
        ws.cell(row=r, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=3).font = FN(10, False, MID)
        ws.cell(row=r, column=3).border = BD()
        ws.cell(row=r, column=4).value = cnt/total if total else 0
        ws.cell(row=r, column=4).number_format = '0%'
        ws.cell(row=r, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=4).font = FN(10, False, MID)
        ws.cell(row=r, column=4).border = BD()
        r += 1

    # ── TOP 10 (below status, clear separation) ──
    r += 1
    ws.cell(row=r, column=2).value = 'TOP 10 MAIS VALIOSOS'
    ws.cell(row=r, column=2).font = FN(9, True, SUBTLE)
    r += 1
    for i, h in enumerate(['Nome','Categoria','Valor (R$)','Desgaste','Valor Atual']):
        cl = ws.cell(row=r, column=2+i)
        cl.value = h
        cl.font = FN(9, True, 'F8FAFC')
        cl.fill = BG(DARK)
        cl.alignment = Alignment(horizontal='center', vertical='center')
        cl.border = BD(DARK)
    ws.row_dimensions[r].height = 26
    r += 1

    top10 = sorted([s for s in activos if s['valor_mercado']], key=lambda x: -(x['valor_mercado'] or 0))[:10]
    for idx, s in enumerate(top10):
        nome = f"{s['marca']} {s['modelo']}".strip() if s['marca'] else s['modelo']
        v_atual = (s['valor_mercado'] or 0) * (s['desgaste']/5) * ESTADO_FATOR[s['estado']]
        ws.cell(row=r, column=2).value = nome
        ws.cell(row=r, column=2).font = FN(10, True, MID)
        ws.cell(row=r, column=3).value = s['categoria']
        ws.cell(row=r, column=3).font = FN(9, False, CAT_DESIGN[s['categoria']]['tx'])
        ws.cell(row=r, column=4).value = s['valor_mercado']
        ws.cell(row=r, column=4).number_format = 'R$ #,##0'
        ws.cell(row=r, column=4).font = FN(10, False, MID)
        ws.cell(row=r, column=5).value = s['desgaste']
        ws.cell(row=r, column=5).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=5).fill = BG(DESGASTE_BG.get(s['desgaste'],'FEF9C3'))
        ws.cell(row=r, column=5).font = FN(10, True, DARK)
        ws.cell(row=r, column=6).value = v_atual
        ws.cell(row=r, column=6).number_format = 'R$ #,##0'
        ws.cell(row=r, column=6).font = FN(10, True, '16A34A')
        drow(ws, r, 6, alt=(idx%2==0))
        ws.cell(row=r, column=2).font = FN(10, True, MID)
        ws.cell(row=r, column=3).font = FN(9, False, CAT_DESIGN[s['categoria']]['tx'])
        ws.cell(row=r, column=5).fill = BG(DESGASTE_BG.get(s['desgaste'],'FEF9C3'))
        ws.cell(row=r, column=5).font = FN(10, True, DARK)
        ws.cell(row=r, column=6).font = FN(10, True, '16A34A')
        r += 1

    # ── DESGASTE POR CATEGORIA (table + bar chart) ──
    r += 1
    ws.cell(row=r, column=2).value = 'DESGASTE MEDIO POR CATEGORIA'
    ws.cell(row=r, column=2).font = FN(9, True, SUBTLE)
    r += 1
    for i, h in enumerate(['Categoria','Media','Visual']):
        cl = ws.cell(row=r, column=2+i)
        cl.value = h
        cl.font = FN(9, True, 'F8FAFC')
        cl.fill = BG(DARK)
        cl.alignment = Alignment(horizontal='center', vertical='center')
        cl.border = BD(DARK)
    ws.row_dimensions[r].height = 26
    r += 1
    desg_start = r

    cat_desg = defaultdict(list)
    for s in activos:
        cat_desg[s['categoria']].append(s['desgaste'])
    for cat in CAT_ORDER:
        degs_cat = cat_desg.get(cat, [3])
        avg = sum(degs_cat)/len(degs_cat)
        ws.cell(row=r, column=2).value = cat
        ws.cell(row=r, column=2).font = FN(10, True, CAT_DESIGN[cat]['tx'])
        ws.cell(row=r, column=3).value = round(avg, 1)
        ws.cell(row=r, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=3).font = FN(10, True, MID)
        ws.cell(row=r, column=4).value = stars(round(avg))
        ws.cell(row=r, column=4).font = FN(10, False, 'CA8A04')
        for c in range(2, 5):
            ws.cell(row=r, column=c).border = BD()
        drow(ws, r, 4, alt=(r%2==0))
        ws.cell(row=r, column=2).font = FN(10, True, CAT_DESIGN[cat]['tx'])
        ws.cell(row=r, column=3).font = FN(10, True, MID)
        ws.cell(row=r, column=4).font = FN(10, False, 'CA8A04')
        r += 1

    # Bar chart for desgaste
    bar_d = BarChart()
    bar_d.type = 'bar'
    bar_d.style = 10
    bar_d.title = None
    bar_d.width = 14
    bar_d.height = 8
    bar_d.y_axis.scaling.min = 0
    bar_d.y_axis.scaling.max = 5
    bar_d.y_axis.delete = True
    bar_d.x_axis.delete = False
    vals_d = Reference(ws, min_col=3, min_row=desg_start, max_row=desg_start+len(CAT_ORDER)-1)
    cats_d = Reference(ws, min_col=2, min_row=desg_start, max_row=desg_start+len(CAT_ORDER)-1)
    bar_d.add_data(vals_d)
    bar_d.set_categories(cats_d)
    bar_d.legend = None
    bar_d.series[0].graphicalProperties.solidFill = ACCENT
    bar_d.dataLabels = DataLabelList()
    bar_d.dataLabels.showVal = True
    ws.add_chart(bar_d, f'F{desg_start-1}')

    # Footer
    r += 2
    add_footer(ws, r, 7)

    # Print
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1

# ── CATEGORY TABS ─────────────────────────────────────────────
COLS = ['Codigo','Nome','Subcategoria','Marca','Serial Fab.','Tag RFID','Status','Estado','Desgaste','Valor Original','Valor Atual','Deprec.%','Notas']
COL_W = [18, 38, 18, 16, 20, 16, 14, 16, 12, 14, 14, 10, 28]
NC = len(COLS)

def build_cat_tab(wb, cat, items):
    design = CAT_DESIGN[cat]
    ws = wb.create_sheet(cat)
    ws.sheet_properties.tabColor = design['tab']
    W(ws, COL_W)

    vals = [i['valor_mercado'] or 0 for i in items]
    degs = [i['desgaste'] for i in items]
    avg_d = sum(degs)/len(degs) if degs else 0

    ws.merge_cells(f'A1:{get_column_letter(NC)}1')
    ws['A1'].value = f"{design['label'].upper()}  —  {design['desc']}"
    ws['A1'].font = FN(14, True, design['tx'])
    ws['A1'].fill = BG(design['bg'])
    ws.row_dimensions[1].height = 38

    ws.merge_cells(f'A2:{get_column_letter(NC)}2')
    ws['A2'].value = f'{len(items)} itens   |   Valor: R$ {sum(vals):,.0f}   |   Desgaste medio: {avg_d:.1f}/5'
    ws['A2'].font = FN(10, False, LIGHT)
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 6

    for i, h in enumerate(COLS):
        ws.cell(row=4, column=i+1).value = h
    hdr(ws, 4, NC)
    ws.row_dimensions[4].height = 28
    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f'A4:{get_column_letter(NC)}4'

    by_sub = defaultdict(list)
    for item in items: by_sub[item['subcategoria']].append(item)

    r = 5
    data_idx = 0
    first_data_row = r
    for subcat in sorted(by_sub.keys()):
        # Separator
        ws.merge_cells(f'A{r}:{get_column_letter(NC)}{r}')
        ws.cell(row=r, column=1).value = f'  {subcat}  ({len(by_sub[subcat])})'
        ws.cell(row=r, column=1).font = FN(10, True, design['tx'])
        ws.cell(row=r, column=1).fill = BG(design['bg'])
        ws.cell(row=r, column=1).alignment = Alignment(vertical='center')
        for c in range(1, NC+1):
            ws.cell(row=r, column=c).fill = BG(design['bg'])
        ws.row_dimensions[r].height = 24
        r += 1

        sub_items = sorted(by_sub[subcat], key=lambda x: x['codigo'])
        for item in sub_items:
            nome = f"{item['marca']} {item['modelo']}".strip() if item['marca'] else item['modelo']
            # Codigo as hyperlink
            c_code = ws.cell(row=r, column=1)
            c_code.value = item['codigo']
            c_code.hyperlink = f"https://mmd.app/s/{item['codigo']}"
            ws.cell(row=r, column=2).value = nome
            ws.cell(row=r, column=3).value = item['subcategoria']
            ws.cell(row=r, column=4).value = item['marca']
            ws.cell(row=r, column=5).value = trunc_serial(item['serial_fabrica'])
            if item['serial_fabrica'] and len(item['serial_fabrica']) > 18:
                ws.cell(row=r, column=5).comment = Comment(item['serial_fabrica'], 'MMD')
            ws.cell(row=r, column=6).value = ''
            ws.cell(row=r, column=7).value = item['status']
            ws.cell(row=r, column=8).value = item['estado']
            ws.cell(row=r, column=9).value = item['desgaste']
            ws.cell(row=r, column=9).alignment = Alignment(horizontal='center')
            ws.cell(row=r, column=10).value = item['valor_mercado']
            if item['valor_mercado']:
                ws.cell(row=r, column=10).number_format = '#,##0'
                ws.cell(row=r, column=11).value = f'=J{r}*(I{r}/5)*IF(H{r}="NOVO",1,IF(H{r}="SEMI_NOVO",0.85,IF(H{r}="USADO",0.65,0.5)))'
                ws.cell(row=r, column=11).number_format = '#,##0'
                ws.cell(row=r, column=12).value = f'=IF(J{r}>0,(J{r}-K{r})/J{r},"")'
                ws.cell(row=r, column=12).number_format = '0%'
                ws.cell(row=r, column=12).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=r, column=13).value = item['notas_short']
            ws.cell(row=r, column=13).alignment = Alignment(vertical='center', wrap_text=True)
            if item['notas_full']:
                ws.cell(row=r, column=13).comment = Comment(item['notas_full'], 'MMD')

            drow(ws, r, NC, alt=(data_idx%2==0))
            ws.cell(row=r, column=1).font = FN(10, True, design['tx'])
            ws.cell(row=r, column=2).font = FN(10, True, MID)
            ws.cell(row=r, column=12).alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[r].height = 22
            data_idx += 1
            r += 1

    # Total row
    r += 1
    ws.merge_cells(f'A{r}:F{r}')
    ws.cell(row=r, column=1).value = f'  TOTAL: {len(items)} itens'
    ws.cell(row=r, column=1).font = FN(10, True, DARK)
    ws.cell(row=r, column=1).fill = BG(BG_ALT)
    ws.cell(row=r, column=1).alignment = Alignment(vertical='center')
    for c in range(1, NC+1):
        ws.cell(row=r, column=c).fill = BG(BG_ALT)
        ws.cell(row=r, column=c).border = BD()
    total_val = sum(i['valor_mercado'] or 0 for i in items)
    if total_val > 0:
        ws.cell(row=r, column=10).value = total_val
        ws.cell(row=r, column=10).number_format = 'R$ #,##0'
        ws.cell(row=r, column=10).font = FN(10, True, DARK)
    ws.row_dimensions[r].height = 26

    # Footer
    r += 2
    add_footer(ws, r, NC)

    last_data_row = max(r, first_data_row + 1)
    add_validations(ws, first_data_row, last_data_row + 200)
    add_cond_fmt(ws, first_data_row, last_data_row + 200)

    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.print_title_rows = '4:4'

# ── LOTES TAB ─────────────────────────────────────────────────
def build_lotes(wb, lots):
    ws = wb.create_sheet('LOTES')
    ws.sheet_properties.tabColor = 'F59E0B'
    headers = ['Codigo','Descricao','Subcategoria','Qtd','Status','Tag RFID','QR Code']
    W(ws, [18, 42, 18, 10, 14, 16, 32])
    for i, h in enumerate(headers):
        ws.cell(row=1, column=i+1).value = h
    hdr(ws, 1, len(headers))
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

    for idx, lot in enumerate(lots):
        r = idx + 2
        ws.cell(row=r, column=1).value = lot['codigo']
        ws.cell(row=r, column=2).value = lot['descricao']
        ws.cell(row=r, column=3).value = lot['subcategoria']
        ws.cell(row=r, column=4).value = lot['quantidade']
        ws.cell(row=r, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=5).value = lot['status']
        ws.cell(row=r, column=6).value = ''
        ws.cell(row=r, column=7).value = f"https://mmd.app/l/{lot['codigo']}"
        drow(ws, r, len(headers), alt=(idx%2==0))
        ws.cell(row=r, column=1).font = FN(10, True, '92400E')
        ws.row_dimensions[r].height = 22

    # Conditional formatting for status
    rng = f'E2:E{len(lots)+100}'
    for st, sty in STATUS_STYLE.items():
        ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=[f'"{st}"'], fill=BG(sty['bg']), font=FN(10, True, sty['tx'])))

# ── FORA DE OPERACAO TAB ──────────────────────────────────────
def build_fora(wb, items):
    ws = wb.create_sheet('FORA DE OPERACAO')
    ws.sheet_properties.tabColor = '64748B'
    headers = COLS + ['Motivo']
    W(ws, COL_W + [30])
    for i, h in enumerate(headers):
        ws.cell(row=1, column=i+1).value = h
    hdr(ws, 1, len(headers))
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = 'A2'

    for idx, item in enumerate(items):
        r = idx + 2
        nome = f"{item['marca']} {item['modelo']}".strip() if item['marca'] else item['modelo']
        ws.cell(row=r, column=1).value = item['codigo']
        ws.cell(row=r, column=2).value = nome
        ws.cell(row=r, column=3).value = item['subcategoria']
        ws.cell(row=r, column=4).value = item['marca']
        ws.cell(row=r, column=5).value = item['serial_fabrica']
        ws.cell(row=r, column=6).value = ''
        ws.cell(row=r, column=7).value = item['status']
        ws.cell(row=r, column=8).value = item['estado']
        ws.cell(row=r, column=9).value = item['desgaste']
        ws.cell(row=r, column=10).value = item['valor_mercado']
        if item['valor_mercado']:
            ws.cell(row=r, column=10).number_format = '#,##0'
        ws.cell(row=r, column=13).value = item['notas_short']
        ws.cell(row=r, column=14).value = item['motivo']
        ws.cell(row=r, column=14).font = FN(10, False, 'DC2626')
        drow(ws, r, len(headers), alt=(idx%2==0))
        ws.cell(row=r, column=1).font = FN(10, True, '64748B')
        ws.row_dimensions[r].height = 22

    rng = f'G2:G{len(items)+100}'
    for st, sty in STATUS_STYLE.items():
        ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=[f'"{st}"'], fill=BG(sty['bg']), font=FN(10, True, sty['tx'])))

# ── MAIN ──────────────────────────────────────────────────────
if __name__ == '__main__':
    print('Lendo inventario original...')
    rows = read_data()
    print(f'  {len(rows)} itens lidos')

    rows = assign_codes(rows)

    activos = [r for r in rows if r['status'] not in ('VENDIDO','EMPRESTADO','BAIXA')]
    fora = [r for r in rows if r['status'] in ('VENDIDO','EMPRESTADO','BAIXA')]
    print(f'  {len(activos)} ativos, {len(fora)} fora de operacao')

    # Filter CABO tab: only cables with serial or value
    cabo_especial = [r for r in activos if r['categoria']=='CABO' and (r['serial_fabrica'] or r['valor_mercado'])]
    cabo_generico = [r for r in activos if r['categoria']=='CABO' and not r['serial_fabrica'] and not r['valor_mercado']]
    print(f'  Cabos: {len(cabo_especial)} especiais (aba CABO), {len(cabo_generico)} genericos (so LOTES)')

    lots = make_lots(rows)
    print(f'  {len(lots)} lotes')

    print('\nConstruindo planilha v3...')
    wb = Workbook()

    build_manual(wb)
    print('  [1/12] MANUAL')

    build_dashboard(wb, activos, lots)
    print('  [2/12] DASHBOARD')

    for i, cat in enumerate(CAT_ORDER):
        if cat == 'CABO':
            cat_items = cabo_especial
        else:
            cat_items = [r for r in activos if r['categoria'] == cat]
        build_cat_tab(wb, cat, cat_items)
        print(f'  [{i+3}/12] {cat} ({len(cat_items)})')

    build_lotes(wb, lots)
    print(f'  [11/12] LOTES ({len(lots)})')

    build_fora(wb, fora)
    print(f'  [12/12] FORA DE OPERACAO ({len(fora)})')

    wb.save(OUTPUT)
    print(f'\nSalvo: {OUTPUT}')
    print('Pronto!')
